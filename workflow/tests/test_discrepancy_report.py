from workflow import discrepancy_report
from factories import workflow_models
from workflow import program, models, utils
from datetime import date
from django import test
from unittest import skip
import openpyxl
import json


# Prevents an exception when ran on github
try:
    msr_country_codes_list = utils.AccessMSR().countrycode_list()
    msr_gaitid_list = utils.AccessMSR().gaitid_list()
except AttributeError:
    pass


@skip('Tests will fail on GitHub without the secret_keys')
class TestDiscrepancyReport(test.TestCase):
    idaa_sample_data_path = 'workflow/tests/idaa_sample_data/idaa_invalid_sample.json'
    discrepancy_report_path = f'workflow/discrepancy_report_{date.today().isoformat()}.xlsx'
    idaa_json = None
    duplicated_gaitids = ["1111.0000000000"]

    def setUp(self):
        with open(self.idaa_sample_data_path) as file:
            self.idaa_json = json.load(file)

        # Need to create tola programs based on pre_create_idaa_programs
        colombia = workflow_models.CountryFactory(country='Colombia', code='CO')
        haiti = workflow_models.CountryFactory(country='Haiti', code='HT')

        idaa_program_multiple_discrepancies_index = 2
        idaa_program_mulitple_discrepancies_json = self.idaa_json['value'][idaa_program_multiple_discrepancies_index]['fields']
        tola_program = models.Program(
            start_date="2025-01-01", end_date="2027-01-01", name="Tola program test",
            funding_status=idaa_program_mulitple_discrepancies_json['ProgramStatus']
        )
        tola_program.save()
        tola_program.country.add(colombia)
        tola_program_gaitid = models.GaitID(gaitid=program.clean_idaa_gaitid(idaa_program_mulitple_discrepancies_json['GaitIDs'][0]['LookupValue']), program_id=tola_program.id)
        tola_program_gaitid.save()

        idaa_program_matches_multiple_index = 3
        idaa_program_matches_multiple_json = self.idaa_json['value'][idaa_program_matches_multiple_index]['fields']
        for number in range(2):
            new_program = models.Program(
                start_date=program.convert_date(idaa_program_matches_multiple_json['ProgramStartDate']),
                end_date=program.convert_date(idaa_program_matches_multiple_json['ProgramEndDate']),
                name=f"{idaa_program_matches_multiple_json['ProgramName']}-{number}", funding_status=idaa_program_matches_multiple_json['ProgramStatus']
            )
            new_program.save()
            new_program.country.add(colombia)
            gaitid = models.GaitID(gaitid=program.clean_idaa_gaitid(idaa_program_matches_multiple_json['GaitIDs'][0]['LookupValue']), program_id=new_program.id)
            gaitid.save()

        tola_program_mismatched_countries_index = 4
        tola_program_mismatched_countries_json = self.idaa_json['value'][tola_program_mismatched_countries_index]['fields']
        tola_program_mismatched_countries = models.Program(
            start_date=program.convert_date(tola_program_mismatched_countries_json['ProgramStartDate']),
            end_date=program.convert_date(tola_program_mismatched_countries_json['ProgramEndDate']),
            name=tola_program_mismatched_countries_json['ProgramName'], funding_status=tola_program_mismatched_countries_json['ProgramStatus'],
        )
        tola_program_mismatched_countries.save()
        tola_program_mismatched_countries.country.add(haiti, colombia)
        tola_program_mismatched_countries_gaitid = models.GaitID(
            gaitid=program.clean_idaa_gaitid(tola_program_mismatched_countries_json['GaitIDs'][0]['LookupValue']), program_id=tola_program_mismatched_countries.id
        )
        tola_program_mismatched_countries_gaitid.save()

    def test_discrepancy_report(self):
        for idaa_program in self.idaa_json['value']:
            upload_program = program.ProgramUpload(idaa_program=idaa_program['fields'], 
                msr_country_codes_list=msr_country_codes_list, msr_gaitid_list=msr_gaitid_list, duplicated_gaitids=self.duplicated_gaitids
            )

            valid_program = upload_program.is_valid()

            # multiple_programs and countries discrepancies are still valid since we are going to update everything but the programs country
            if upload_program.has_discrepancy('multiple_programs') or upload_program.has_discrepancy('countries'):
                self.assertTrue(valid_program)
            else:
                self.assertFalse(valid_program)

            upload_program.create_discrepancies()

        discrepancy_report_obj = discrepancy_report.GenerateDiscrepancyReport()
        discrepancy_report_obj.generate()

        workbook = openpyxl.load_workbook(self.discrepancy_report_path)
        self.assert_workbook(workbook)

    def assert_workbook(self, workbook):
        worksheets = [
            {'class': discrepancy_report.OverviewTab, 'expected_rows': 20, 'id_col': None, 'expected_ids': None},
            {'class': discrepancy_report.MultipleProgramsTab, 'expected_rows': 3, 'id_col': 'F', 'expected_ids': [3, 3]},
            {'class': discrepancy_report.MismatchingFieldsTab, 'expected_rows': 3, 'id_col': 'K', 'expected_ids': [2, 4]},
            {'class': discrepancy_report.IDAAInvalidFieldsTab, 'expected_rows': 4, 'id_col': 'C', 'expected_ids': [1, 5, 6]},
            {'class': discrepancy_report.DuplicateIDAAProgramsTab, 'expected_rows': 3, 'id_col': 'B', 'expected_ids': [8, 9]}
        ]

        self.assertEquals(len(worksheets), len(workbook.worksheets))

        for index, worksheet in enumerate(workbook.worksheets):
            worksheet_tab = worksheets[index]['class']
            expected_rows = worksheets[index]['expected_rows']
            id_col = worksheets[index]['id_col']
            expected_ids = worksheets[index]['expected_ids']

            self.assertEquals(worksheet.title, worksheet_tab.title)
            self.assertEquals(worksheet.max_column, len(worksheet_tab.columns))
            self.assertEquals(worksheet.max_row, expected_rows)

            if id_col:
                self.assert_programs_displayed(worksheet, id_col, expected_ids)

    def assert_programs_displayed(self, worksheet, id_col, expected_ids):
        # Checks that programs are displayed on the correct worksheet tab
        for index in range(2, worksheet.max_row + 1):
            self.assertEquals(worksheet[f'{id_col}{index}'].value, expected_ids[index - 2])
