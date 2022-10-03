from workflow.models import ProgramDiscrepancy
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from workflow.program import convert_date
from openpyxl.utils import get_column_letter
from openpyxl import workbook
from datetime import date
from os import path


class DiscrepancyReportTab:
    discrepancies = []
    columns = []
    title = ''
    discrepancy_to_columns = []
    wide_cell_width = 60
    standard_cell_width = 25
    small_cell_width = 15
    column_font = Font(bold=True, size='12')

    @staticmethod
    def comma_separate_list(separate_list):
        try:
            if type(separate_list[0]) is int:
                # Need to convert int to str to use .join
                separate_list = [str(x) for x in separate_list]
        except IndexError:
            pass

        return ','.join(separate_list)

    def get_idaa_gaitids(self, idaa_json):
        return self.comma_separate_list([str(gaitid['LookupValue']).split('.')[0] for gaitid in idaa_json['GaitIDs']])

    def get_idaa_countries(self, idaa_json):
        return self.comma_separate_list([country['LookupValue'] for country in idaa_json['Country']])

    def get_discrepancy_reasons(self, worksheet_discrepancies):
        return ','.join([ProgramDiscrepancy.DISCREPANCY_REASONS[discrepancy] for discrepancy in worksheet_discrepancies])

    def format_worksheet(self, worksheet, wide_cells=None, small_cells=None):
        """
        Method for formatting the worksheet
        """
        if wide_cells is None:
            wide_cells = [0, 1]

        if small_cells is None:
            small_cells = [2]

        for index in range(len(self.columns)):
            if index in wide_cells:
                width = self.wide_cell_width
            elif index in small_cells:
                width = self.small_cell_width
            else:
                width = self.standard_cell_width
            
            column_letter = get_column_letter(index + 1)
            worksheet.column_dimensions[column_letter].width = width

        # Add custom font to column headers
        for column in worksheet["1:1"]:
            column.font = self.column_font

    def after_populate_format(self, worksheet):
        """
        Some cells can't be formatted until data is inserted. Override this method in the tab that needs special formatting
        """
        pass

    def create_worksheet(self, wb):
        worksheet = wb.create_sheet(self.title)
        worksheet.title = self.title

        if type(self.columns) is list:
            worksheet.append(self.columns)

        self.format_worksheet(worksheet)

    def populate(self, program_discrepancy, *args, **kwargs):
        """
        Returns a list to be appended to the worksheet
        """
        idaa_json = program_discrepancy.idaa_json
        idaa_gaitids = self.get_idaa_gaitids(idaa_json)

        if len(idaa_json['Country']) == 0:
            country = ""
        else:
            country = self.get_idaa_countries(idaa_json)

        return [
            idaa_json.get('ProgramName', 'N/A'), idaa_json['id'], idaa_gaitids, country, convert_date(idaa_json.get('ProgramStartDate', None), readable=True), 
            convert_date(idaa_json.get('ProgramEndDate', None), readable=True), idaa_json.get('ProgramStatus', 'N/A'), ''
        ]


class IDAAInvalidFieldsTab(DiscrepancyReportTab):
    discrepancies = ["ProgramName", "id", "ProgramStartDate", "ProgramEndDate", "Country", "ProgramStatus", "funded", "gaitid"]
    columns = [
        "Discrepancy Reasons", "IDAA Program Name", "IDAA Program ID", "IDAA GAIT IDs", "IDAA Countries", "IDAA Start Date", 
        "IDAA End Date", "IDAA Program Status", "Notes"
    ]
    title = "IDAA Program Has Missing Data"
    discrepancy_to_columns = [
        {"discrepancy": "funded", "columns": ['IDAA Program Status']},
        {"discrepancy": "gaitid", "columns": ['IDAA GAIT IDs']},
        {"discrepancy": "ProgramName", "columns": ['IDAA Program Name']},
        {"discrepancy": "ID", "columns": ["IDAA Program ID"]},
        {"discrepancy": "ProgramStartDate", "columns": ['IDAA Start Date']},
        {"discrepancy": "ProgramEndDate", "columns": ['IDAA End Date']},
        {"discrepancy": "Country", "columns": ["IDAA Countries"]},
        {"discrepancy": "ProgramStatus", "columns": ['IDAA Program Status']}
    ] 

    def after_populate_format(self, worksheet):
        last_row = worksheet.max_row
        coord = f"A{last_row}"

        worksheet[coord].alignment = Alignment(wrap_text=True)

    def populate(self, program_discrepancy, worksheet_discrepancies):
        reasons = self.get_discrepancy_reasons(worksheet_discrepancies)

        row = [reasons]

        row.extend(super().populate(program_discrepancy, worksheet_discrepancies))

        return row


class MismatchingFieldsTab(DiscrepancyReportTab):
    discrepancies = ["funding_status", "end_date", "start_date", "countries", "out_of_bounds_tracking_dates"]
    columns = [
        "Discrepancy Reasons", "TolaData Program Name", "TolaData GAIT IDs", "TolaData Countries", "TolaData Start Date", "TolaData End Date",
        "TolaData Indicator Tracking Start Date", "TolaData Indicator Tracking End Date", "TolaData Funding Status", "IDAA Program Name", "IDAA Program ID", 
        "IDAA GAIT IDs", "IDAA Countries", "IDAA Start Date", "IDAA End Date", "IDAA Program Status", "Notes"
    ]
    title = "MisMatching Fields"
    discrepancy_to_columns = [
        {"discrepancy": "funding_status", "columns": ['IDAA Program Status', 'TolaData Funding Status']},
        {"discrepancy": "end_date", "columns": ['IDAA End Date', 'TolaData End Date']},
        {"discrepancy": "start_date", "columns": ['IDAA Start Date', 'TolaData Start Date']},
        {"discrepancy": "countries", "columns": ['IDAA Countries', 'TolaData Countries']},
        {"discrepancy": "out_of_bounds_tracking_dates", "columns": [
            'IDAA Start Date', 'IDAA End Date', 'TolaData Indicator Tracking Start Date', 'TolaData Indicator Tracking End Date'
            ]}
    ]

    def format_worksheet(self, worksheet):
        wide_cells = [
            self.columns.index('Discrepancy Reasons'),
            self.columns.index('IDAA Program Name'),
            self.columns.index('TolaData Program Name')
        ]

        super().format_worksheet(worksheet, wide_cells)

    def after_populate_format(self, worksheet):
        last_row = worksheet.max_row
        coord = f"A{last_row}"

        worksheet[coord].alignment = Alignment(wrap_text=True)

    def populate(self, program_discrepancy, worksheet_discrepancies):
        tola_program = program_discrepancy.program.first()
        reasons = self.get_discrepancy_reasons(worksheet_discrepancies)

        row = [
            reasons, tola_program.name, self.comma_separate_list(tola_program.gaitids), tola_program.countries,
            convert_date(str(tola_program.start_date), readable=True), convert_date(str(tola_program.end_date), readable=True),
            convert_date(str(tola_program.reporting_period_start), readable=True), convert_date(str(tola_program.reporting_period_end), readable=True),
            tola_program.funding_status
        ]

        row.extend(super().populate(program_discrepancy, worksheet_discrepancies))

        return row


class MultipleProgramsTab(DiscrepancyReportTab):
    discrepancies = ["multiple_programs"]
    columns = [
        "TolaData Program Name", "TolaData GAIT IDs", "TolaData Countries", "TolaData Funding Status", 
        "IDAA Program Name", "IDAA Program ID", "IDAA GAIT IDs", "IDAA Countries", "IDAA Funding Status", "Notes"
    ]
    title = "IDAA Program to Multiple TolaDa"  # Max character length of 31

    def format_worksheet(self, worksheet):
        wide_cells = [self.columns.index('TolaData Program Name'), self.columns.index('IDAA Program Name')]
        small_cells = [self.columns.index('IDAA Program ID')]

        super().format_worksheet(worksheet, wide_cells, small_cells)

    def populate(self, idaa_json, tola_program):
        return [
            tola_program.name, self.comma_separate_list(tola_program.gaitids), tola_program.countries, tola_program.funding_status,
            idaa_json['ProgramName'], idaa_json['id'], self.get_idaa_gaitids(idaa_json), self.get_idaa_countries(idaa_json),
            idaa_json['ProgramStatus'], ''
        ]


class DuplicateIDAAProgramsTab(DiscrepancyReportTab):
    columns = [
        "IDAA Program Name", "IDAA Program ID", "IDAA GAIT IDs", "IDAA Countries", "IDAA Start Date", 
        "IDAA End Date", "IDAA Program Status", "Notes"
    ]
    title = "Duplicate IDAA Programs"
    discrepancies = ["duplicate_gaitid"]

    def format_worksheet(self, worksheet, wide_cells=None, small_cells=None):
        wide_cells = [self.columns.index('IDAA Program Name')]
        small_cells = [self.columns.index('IDAA Program ID')]
        return super().format_worksheet(worksheet, wide_cells, small_cells)

    def sort_by_gaitids(self, programs):
        """
        Takes a list of programs formatted from super().populate()
        Sorts the programs to have duplicates show as the next list item
        """
        sorted_programs = []
        checked_ids = []
        gaitid_key = self.columns.index('IDAA GAIT IDs')
        id_key = self.columns.index('IDAA Program ID')
        for program in programs:
            gaitids = [int(gaitid) for gaitid in program[gaitid_key].split(',')]
            if program[id_key] in checked_ids:
                continue
            checked_ids.append(program[1])
            sorted_programs.append(program)
            for gaitid in gaitids:
                for other_program in programs:
                    if other_program[id_key] in checked_ids:
                        continue
                    other_gaitids = [int(gaitid) for gaitid in other_program[gaitid_key].split(',')]
                    if gaitid in other_gaitids:
                        sorted_programs.append(other_program)
                        checked_ids.append(other_program[1])

        return sorted_programs


class OverviewTab(DiscrepancyReportTab):
    columns = range(2)  # columns as a range to work with inherited formating methods
    wide_cell_width = 120
    title = "Discrepancy Report Overview"
    static_text = [
        {
            "header": "ORIENTATION",
            "body": [
                "The Discrepancy Report is divided into 5 tabs:",
                "The first tab, titled Discrepancy Report Overview, provides explanation and instruction on what this Discrepancy Report is and how to use this to improve program data quality and consistency in TolaData (and IDAA).",
                "The second tab, titled IDAA Program to Multiple TolaData, identifies duplicated programs in TolaData, where there is one program in IDAA but this single program is broken out into multiple programs in TolaData. This often happens with multi-country programs, where each individual country decided to make their own version of the program in TolaData.",
                "The third tab, titled MisMatching Fields, identifies existing programs in TolaData whose Countries and/or Indicator Tracking dates do not match the data and information in IDAA. These discrepancies are not automatically corrected by the system given their sensitive nature. As a result, these discrepancies must be dealt with manually on a case-by-case basis.",
                "The fourth tab, titled IDAA Program Has Missing Data, identifies programs in IDAA, which cannot be added or updated in TolaData due to data quality issues. These issues need to be addressed directly in IDAA before these programs can be added or updated in TolaData.",
                "The fifth tab, titled Duplicate IDAA Programs, identifies programs in IDAA that have the same GAIT ID(s) assigned to them. A GAIT ID should only ever be assigned to one program. A single program may have multiple GAIT IDs associated to it, but a single GAIT ID should never be assigned to multiple programs. This check also attempts to address or prevent two possible undesirable scenarios: 1) many IDAA programs to one TolaData program and 2) many IDAA programs to many TolaData programs."
            ]
        },
        {
            "header": "PURPOSE & USE",
            "body": [
                "The purpose of this Discrepancy Report is to provide HQ and Regional MEL/PAQ/Standards Advisors, Regional Program Team (RPT) members, and Country and Program Teams with data and information regarding discrepancies found across IDAA and TolaData with the goal of correcting these discrepancies.",
                "Most discrepancies between IDAA and TolaData will be addressed by the systems automatically, but there are a handful of sensitive issues that merit special review and handling by people, not systems. This report attempts to identify those issues that require manual review and correction by team members. For example, the countries assigned to a program in TolaData determines in which country portfolio the program appears and, as a result, which users see and have access to that program. Changing Indicator Tracking Dates impacts time-based indicators with periodic targets and any results associated with periodic targets. Therefore, updating these dates should only be done by or with authorization from the program team.",
                "The Discrepancy Report is generated by the TolaData Development Team at the beginning and middle of every month and is shared with the HQ MEL and Standards Teams for communication and dissemination to the relevant teams. Issues in TolaData may be addressed by anyone with authority and authorization to do so, such as the Country or Program Teams with support from the HQ and Regional MEL Advisors. Issues in IDAA will need to be addressed by those responsible for data entry and maintenance in that system, namely the RPT and/or the Country or Program Teams with support from the HQ Standards Advisors and IDAA Product Owner. It is ultimately up to these respective teams on when, how, and if these issues are addressed."
            ]
        }
    ]

    def format_worksheet(self, worksheet):
        wide_cells = [1]
        small_cells = [0]

        worksheet.sheet_view.showGridLines = False

        super().format_worksheet(worksheet, wide_cells, small_cells)

    def after_populate_format(self, worksheet):
        last_row = worksheet.max_row
        coord = f"B{last_row}"

        worksheet[coord].alignment = Alignment(wrap_text=True)

    def populate_static(self, worksheet):
        """
        Method for populating the worksheet with static data.
        """
        for static_text in self.static_text:
            worksheet.append([static_text['header']])
            for body_text in static_text['body']:
                worksheet.append(['', body_text])
                self.after_populate_format(worksheet)
                worksheet.append([])


class GenerateDiscrepancyReport:
    file_path = path.join(path.dirname(path.abspath(__file__)), f'discrepancy_report_{date.today().isoformat()}.xlsx')
    _worksheet_mapper = [
        OverviewTab, MultipleProgramsTab, MismatchingFieldsTab, IDAAInvalidFieldsTab, DuplicateIDAAProgramsTab
    ]
    discrepancy_highlight = '40 % - Accent2'
    discrepancy_border = Border(left=Side(style='thin', color='fff00000'), 
                                right=Side(style='thin', color='fff00000'), 
                                top=Side(style='thin', color='fff00000'), 
                                bottom=Side(style='thin', color='fff00000'))
    
    def __init__(self):
        self.wb = workbook.Workbook()

        self._create_worksheets()

    def _create_worksheets(self):
        # Delete the default worksheet
        del self.wb['Sheet']

        for worksheet_object in self._worksheet_mapper:
            worksheet_object().create_worksheet(self.wb)

    def discrepancies_to_worksheets(self, discrepancies):
        """
        returns a list of worksheets that a list of discrepancies are shown in
        """
        used_worksheets = set()

        for worksheet_object in self._worksheet_mapper:
            for discrepancy in discrepancies:
                if discrepancy in worksheet_object.discrepancies:
                    used_worksheets.add(worksheet_object.title)

        return list(used_worksheets)


    def get_discrepancy_column_number(self, discrepancies, worksheet_object):
        """
        Returns the column number a discrepancy is located at
        """
        highlight_indexes = []

        for discrepancy in discrepancies:
            for discrepancy_column in worksheet_object.discrepancy_to_columns:
                if discrepancy_column['discrepancy'] == discrepancy:
                    for column in discrepancy_column['columns']:
                        highlight_indexes.append(
                            worksheet_object.columns.index(column) + 1  # Excel coords are not zero indexed
                        )
        
        return highlight_indexes

    def highlight_discrepancies(self, worksheet, discrepancies, worksheet_object):
        """
        Highlights discrepancies on the worksheet
        """
        last_row = worksheet.max_row
        column_indexes = self.get_discrepancy_column_number(discrepancies, worksheet_object)
        
        for column_index in column_indexes:
            column_letter = get_column_letter(column_index)
            coord = f"{column_letter}{last_row}"
            worksheet[coord].style = self.discrepancy_highlight
            worksheet[coord].border = self.discrepancy_border

    def generate(self):
        """
        Method to generate the discrepancy report
        """
        duplicated_idaa_programs = []
        program_discrepancies = ProgramDiscrepancy.objects.all()
        overview_tab = OverviewTab()
        overview_tab.populate_static(self.wb[overview_tab.title])
        duplicate_idaa_tab = DuplicateIDAAProgramsTab()
        duplicate_idaa_worksheet = self.wb[duplicate_idaa_tab.title]

        for program_discrepancy in program_discrepancies:
            for worksheet_object in self._worksheet_mapper:
                worksheet_discrepancies = []
                worksheet = self.wb[worksheet_object.title]

                for discrepancy in program_discrepancy.discrepancies:
                    if discrepancy in worksheet_object.discrepancies:
                        worksheet_discrepancies.append(discrepancy)

                if len(worksheet_discrepancies) == 0:
                    continue

                worksheet_tab = worksheet_object()

                if isinstance(worksheet_tab, MultipleProgramsTab):
                    for tola_program in program_discrepancy.program.all():
                        row = worksheet_tab.populate(program_discrepancy.idaa_json, tola_program)

                        worksheet.append(row)
                elif isinstance(worksheet_tab, DuplicateIDAAProgramsTab):
                    row = worksheet_tab.populate(program_discrepancy, worksheet_discrepancies)
                    duplicated_idaa_programs.append(row)
                else:
                    row = worksheet_tab.populate(program_discrepancy, worksheet_discrepancies)

                    worksheet.append(row)

                    self.highlight_discrepancies(worksheet, worksheet_discrepancies, worksheet_object)

                worksheet_tab.after_populate_format(worksheet)

        rows = duplicate_idaa_tab.sort_by_gaitids(duplicated_idaa_programs)

        for row in rows:
            duplicate_idaa_worksheet.append(row)

        self.wb.save(self.file_path)
