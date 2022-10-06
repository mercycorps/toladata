
from collections import OrderedDict
import pytz
from django.db import transaction
from django.db import models
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.translation import ugettext as _
from rest_framework import viewsets, permissions, pagination
from rest_framework.response import Response
from rest_framework import status as httpstatus
from rest_framework.decorators import action
from rest_framework.serializers import (
    ModelSerializer,
    Serializer,
    CharField,
    IntegerField,
    ValidationError,
    BooleanField,
    DateField,
    DateTimeField,
    SerializerMethodField,
    JSONField
)

from openpyxl import Workbook, utils
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill

from dateutil.relativedelta import relativedelta
from datetime import date

from workflow.models import (
    Program,
    TolaUser,
    Organization,
    Country,
    Sector,
    ProgramAccess,
    CountryAccess,
    GaitID,
    FundCode,
    IDAASector
)

from indicators.models import (
    Indicator,
    Level,
    IDAAOutcomeTheme
)

from tola_management.models import (
    ProgramAdminAuditLog,
    ProgramAuditLog
)

from tola_management.permissions import HasProgramAdminAccess


def get_audit_log_workbook(ws, program):

    def _indicator_number(indicator):
        if indicator.results_aware_number:
            return f"{_('Indicator')} {indicator.results_aware_number}"
        else:
            return _('Indicator')

    # helper for result level column and result level diff row text
    def _result_level(row):
        if row.indicator:
            if row.indicator.leveltier_name and row.indicator.level_display_ontology:
                return u'{} {}'.format(
                    str(row.indicator.leveltier_name),
                    str(row.indicator.level_display_ontology),
                )
            elif row.indicator.leveltier_name:
                return str(row.indicator.leveltier_name)
        if row.level:
            return f"{row.level.leveltier} {row.level.display_ontology}"
        else:
            return None


    # helper for result level column
    def _result_disaggregation_serializer(raw_disaggs):
        disaggs = {}
        for item in raw_disaggs.values():
            # default value of "unknown" for data stored before we were also logging 'type'
            disaggregation_type = item.get('type', 'unknown')
            try:
                disaggs[disaggregation_type].append(item)
            except KeyError:
                disaggs[disaggregation_type] = [item]

        output_string = ""
        for disagg_type in sorted(list(disaggs.keys())):
            if disagg_type and disagg_type != 'unknown':  # don't include "unknown" placeholder in output
                output_string += f"\r\n{disagg_type}\r\n"
            else:
                output_string += "\r\n"
            disaggs[disagg_type].sort(key=lambda item: item.get('customsort', ''))

            for item in disaggs[disagg_type]:
                output_string += f"{item['name']}: {item['value']}\r\n"

        return output_string


    header = [
        # Translators: The date and time of the change made to a piece of data
        Cell(ws, value=_("Date and time")),
        # Translators: Number of the indicator being shown
        Cell(ws, value=_('Result level')),
        Cell(ws, value=_('Indicator')),
        # Translators: The name of the user who carried out an action
        Cell(ws, value=_('User')),
        Cell(ws, value=_('Organization')),
        # Translators: Part of change log, indicates the type of change being made to a particular piece of data
        Cell(ws, value=_('Change type')),
        # Translators: Part of change log, shows what the data looked like before the changes
        Cell(ws, value=_('Previous entry')),
        # Translators: Part of change log, shows what the data looks like after the changes
        Cell(ws, value=_('New entry')),
        # Translators: Part of change log, reason for the change as entered by the user
        Cell(ws, value=_('Rationale'))
    ]

    # Translators: Page title for a log of all changes to indicators over a program's history
    title = Cell(ws, value=_("Program change log"))
    title.font = Font(size=18)
    ws.append([title,])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(header))
    subtitle = Cell(ws, value=program.name)
    subtitle.font = Font(size=18)
    ws.append([subtitle,])
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(header))

    header_font = Font(bold=True)
    header_fill = PatternFill('solid', 'EEEEEE')

    for h in header:
        h.font = header_font
        h.fill = header_fill

    ws.append(header)

    alignment = Alignment(
        horizontal='general',
        vertical='top',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0
    )

    for row in program.audit_logs.all().order_by('-date'):
        indicator_name = f'{_indicator_number(row.indicator)}: {row.indicator.name}' if row.indicator else "N/A"
        prev_string = ''
        new_string = ''
        for entry in row.diff_list:
            if entry['name'] == "id":
                continue
            elif entry['name'] == 'targets':
                for target in entry['prev'].values():
                    prev_string += str(target['name']) + ": " + str(target['value']) + "\r\n"
                for target in entry['new'].values():
                    new_string += str(target['name']) + ": " + str(target['value']) + "\r\n"
            elif entry['name'] == 'disaggregation_values':
                prev_string += _result_disaggregation_serializer(entry['prev']) + "\r\n"
                new_string += _result_disaggregation_serializer(entry['new']) + "\r\n"
            elif row.change_type == "indicator_changed" and entry["name"] == "name":
                prev_string += f"{_indicator_number(row.indicator)}: {entry['prev']} \r\n"
                new_string += f"{_indicator_number(row.indicator)}: {entry['new']} \r\n"
            elif row.change_type == "level_changed" and entry["name"] == "name":
                prev_string += f"{_result_level(row)}: {entry['prev']} \r\n"
                new_string += f"{_result_level(row)}: {entry['new']} \r\n"
            else:
                prev_string += str(entry['pretty_name']) + ": "
                prev_string += str(entry['prev'] if entry['prev'] else "") + "\r\n"
                new_string += str(entry['pretty_name']) + ": "
                new_string += str(entry['new'] if entry['new'] else "") + "\r\n"

        xl_row = [
            Cell(ws, value=row.date.strftime("%Y-%m-%d %H:%M:%S (UTC)")),
            Cell(ws, value=_result_level(row) or _('N/A')),
            Cell(ws, value=indicator_name),
            Cell(ws, value=str(row.user.name)),
            Cell(ws, value=str(row.organization.name)),
            Cell(ws, value=str(row.pretty_change_type)),
            Cell(ws, value=str(prev_string)),
            Cell(ws, value=str(new_string)),
            Cell(ws, value=str(row.rationale))
        ]
        for cell in xl_row:
            cell.alignment = alignment
        ws.append(xl_row)

    for rd in ws.row_dimensions:
        rd.auto_size = True

    for cd in ws.column_dimensions:
        cd.auto_size = True
    widths = [23, 12, 50, 20, 15, 20, 40, 40, 40]
    for col_no, width in enumerate(widths):
        ws.column_dimensions[utils.get_column_letter(col_no + 1)].width = width
    return ws


class Paginator(pagination.PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 50

    def get_paginated_response(self, data):
        response = Response(OrderedDict([
            ('count', self.page.paginator.count),
            ('page_count', self.page.paginator.num_pages),
            ('next', self.get_next_link()),
            ('previous', self.get_previous_link()),
            ('results', data),
        ]))
        return response


class NestedIDAASectorSerializer(Serializer):
    def to_representation(self, sector):
        return sector.id

    def to_internal_value(self, data):
        sector = IDAASector.objects.get(pk=data)
        return sector


class NestedCountrySerializer(Serializer):

    def to_representation(self, country):
        return country.id

    def to_internal_value(self, data):
        country = Country.objects.get(pk=data)
        return country


class NestedFundCodeSerializer(ModelSerializer):
    fund_code = IntegerField(max_value=99999)

    class Meta:
        model = FundCode
        fields = ('fund_code',)

    def to_representation(self, instance):
        return instance.fund_code

    def to_internal_value(self, data):
        return data


class NestedGaitIDSerializer(ModelSerializer):
    gaitid = IntegerField(max_value=99999)
    donor = CharField(max_length=255)
    donor_dept = CharField()
    fund_code = NestedFundCodeSerializer(many=True)

    class Meta:
        model = GaitID
        fields = ('gaitid', 'donor', 'donor_dept', 'fund_code')

    def to_representation(self, instance):
        return {
            'gaitid': instance.gaitid,
            'donor': instance.donor,
            'donor_dept': instance.donor_dept,
            'fund_code': [fc.fund_code for fc in instance.fundcode_set.all()]
        }


class NestedIDAAOutcomeThemeSerializer(Serializer):

    def to_representation(self, outcome_theme):
        return outcome_theme.id

    def to_internal_value(self, data):
        if type(data) is list:
            pk = data[0]
        else:
            pk = data
        idaa_outcome_theme = IDAAOutcomeTheme.objects.get(pk=pk)
        return idaa_outcome_theme


class ProgramAdminSerializer(ModelSerializer):
    id = IntegerField(read_only=True)
    external_program_id = IntegerField(required=True, max_value=9999)
    name = CharField(required=True, max_length=255)
    funding_status = CharField(required=True)
    gaitid = NestedGaitIDSerializer(required=True, many=True)
    fundCode = CharField(required=False, allow_blank=True, allow_null=True, source='cost_center')
    idaa_sector = NestedIDAASectorSerializer(required=True, many=True)
    country = NestedCountrySerializer(required=True, many=True)
    auto_number_indicators = BooleanField(required=False)
    #organizations = IntegerField(source='organization_count', read_only=True)
    #program_users = IntegerField(source='program_users_count', read_only=True)
    #onlyOrganizationId = IntegerField(source='only_organization_id', read_only=True)
    organizations = SerializerMethodField()
    program_users = SerializerMethodField()
    onlyOrganizationId = SerializerMethodField()
    _using_results_framework = IntegerField(required=False, allow_null=True)
    start_date = DateField(required=True)
    end_date = DateField(required=True)
    idaa_outcome_theme = NestedIDAAOutcomeThemeSerializer(required=True, many=True)

    def validate_country(self, values):
        if not values:
            raise ValidationError(_("This field may not be blank."))
        return values

    def validate_dates(self, start_date, end_date):
        if start_date and end_date:
            if start_date > end_date:
                raise ValidationError(_("The program start date may not be after the program end date."))
            if end_date < start_date:
                raise ValidationError(_("The program end date may not be before the program start date."))

    def validate(self, data):
        self.validate_dates(data['start_date'], data['end_date'])
        return super().validate(data)

    def validate_start_date(self, value):
        past_limit = date.today() - relativedelta(years=10)
        if value < past_limit:
            raise ValidationError(_("The program start date may not be more than 10 years in the past."))
        return value

    def validate_end_date(self, value):
        future_limit = date.today() + relativedelta(years=10)
        if value > future_limit:
            raise ValidationError(_("The program end date may not be more than 10 years in the future."))

    def validate_gaitid(self, values):
        checked_gaitids = []
        for gid in values:
            if gid['gaitid'] in checked_gaitids:
                # error
                raise ValidationError(_("Duplicate GAIT ID numbers are not allowed."))
            else:
                checked_gaitids.append(gid['gaitid'])

        return values

    class Meta:
        model = Program
        fields = (
            'id',
            'external_program_id',
            'name',
            'funding_status',
            'gaitid',
            'fundCode',
            'idaa_sector',
            'country',
            'organizations',
            'program_users',
            'onlyOrganizationId',
            'auto_number_indicators',
            '_using_results_framework',
            'start_date',
            'end_date',
            'idaa_outcome_theme'
        )

    @staticmethod
    def get_program_users(program):
        tolauser_ids = set(
            [ca.tolauser_id for country in getattr(program, 'country_with_users', [])
             for ca in getattr(country, 'country_users', [])] +
            [pa.tolauser_id for pa in getattr(program, 'program_users', [])]
        )
        return len(tolauser_ids) + getattr(program, 'su_count', 0)

    @staticmethod
    def _get_org_ids(program):
        org_ids = set(
            [ca.tolauser.organization_id for country in getattr(program, 'country_with_users', [])
             for ca in getattr(country, 'country_users', [])] +
            [pa.tolauser.organization_id for pa in getattr(program, 'program_users', [])] +
            ([Organization.MERCY_CORPS_ID] if getattr(program, 'su_count', 0) else [])
        )
        return org_ids

    def get_organizations(self, program):
        org_ids = self._get_org_ids(program)
        return len(org_ids)

    def get_onlyOrganizationId(self, program):
        org_ids = self._get_org_ids(program)
        if len(org_ids) == 1:
            return list(org_ids)[0]
        return None

    def to_representation(self, program, with_aggregates=True):
        ret = super(ProgramAdminSerializer, self).to_representation(program)
        if not with_aggregates:
            return ret
        if ret['_using_results_framework'] == Program.RF_ALWAYS:
            ret.pop('_using_results_framework')
        return ret

    @transaction.atomic
    def create(self, validated_data):
        if '_using_results_framework' in validated_data and \
                validated_data['_using_results_framework'] is None:
            validated_data.pop('_using_results_framework')
        country = validated_data.pop('country')
        sector = validated_data.pop('idaa_sector')
        gaitid_data = validated_data.pop('gaitid')
        idaa_outcome_theme = validated_data.pop('idaa_outcome_theme')
        program = super(ProgramAdminSerializer, self).create(validated_data)
        program.country.add(*country)
        program.idaa_sector.add(*sector)
        program.idaa_outcome_theme.add(*idaa_outcome_theme)
        program.save()

        for gid in gaitid_data:
            gaitid = GaitID(gaitid=gid['gaitid'], donor=gid['donor'], program=program)
            gaitid.save()

            for fund_code in gid['fund_code']:
                if fund_code:
                    new_fund_code = FundCode(fund_code=fund_code, gaitid=gaitid)
                    new_fund_code.save()

        ProgramAdminAuditLog.created(
            program=program,
            created_by=self.context.get('request').user.tola_user,
            entry=program.admin_logged_fields,
        )

        return program

    @transaction.atomic
    def update(self, instance, validated_data):
        previous_state = instance.admin_logged_fields

        if '_using_results_framework' in validated_data and validated_data['_using_results_framework'] is None:
            validated_data['_using_results_framework'] = instance._using_results_framework

        # default for any unmigrated program is "auto" - so if someone sets their program to "not grouping" - reset it
        # to default ("auto")
        if ('_using_results_framework' in validated_data
                and validated_data['_using_results_framework'] == instance.NOT_MIGRATED):
            validated_data['auto_number_indicators'] = True

        original_countries = instance.country.all()
        incoming_countries = validated_data.pop('country')
        added_countries = [x for x in incoming_countries if x not in original_countries]
        removed_countries = [x for x in original_countries if x not in incoming_countries]

        original_sectors = instance.idaa_sector.all()
        incoming_sectors = validated_data.pop('idaa_sector')
        added_sectors = [x for x in incoming_sectors if x not in original_sectors]
        removed_sectors = [x for x in original_sectors if x not in incoming_sectors]

        original_outcome_theme = instance.idaa_outcome_theme.all()
        incoming_outcome_theme = validated_data.pop('idaa_outcome_theme')
        added_outcome_theme = [x for x in incoming_outcome_theme if x not in original_outcome_theme]
        removed_outcome_theme = [x for x in original_outcome_theme if x not in incoming_outcome_theme]

        gaitid_data = validated_data.pop('gaitid')
        for gid in gaitid_data:
            gaitid_obj, created = GaitID.objects.get_or_create(gaitid=gid['gaitid'], program_id=instance.id)
            gaitid_obj.donor = gid['donor']
            gaitid_obj.save()

            for fund_code in gid['fund_code']:
                _ = FundCode.objects.get_or_create(fund_code=fund_code, gaitid=gaitid_obj)

            if not created:
                for previous_fund_code in gaitid_obj.fundcode_set.all():
                    if previous_fund_code.fund_code not in gid['fund_code']:
                        previous_fund_code.delete()

        for previous_gaitid in instance.gaitid.all():
            if previous_gaitid.gaitid not in [gid['gaitid'] for gid in gaitid_data]:
                previous_gaitid.delete()


        instance.country.remove(*removed_countries)
        instance.country.add(*added_countries)
        instance.idaa_sector.remove(*removed_sectors)
        instance.idaa_sector.add(*added_sectors)
        instance.idaa_outcome_theme.remove(*removed_outcome_theme)
        instance.idaa_outcome_theme.add(*added_outcome_theme)

        ProgramAccess.objects.filter(program=instance, country__in=removed_countries).delete()

        updated_instance = super(ProgramAdminSerializer, self).update(instance, validated_data)
        ProgramAdminAuditLog.updated(
            program=instance,
            changed_by=self.context.get('request').user.tola_user,
            old=previous_state,
            new=instance.admin_logged_fields,
        )
        return updated_instance


class ProgramAuditLogIndicatorSerializer(ModelSerializer):
    class Meta:
        model = Indicator
        fields = (
            'name',
            'leveltier_name',
            'level_display_ontology',
            'results_aware_number',
        )


class ProgramAuditLogLevelSerializer(ModelSerializer):
    tier = SerializerMethodField()

    class Meta:
        model = Level
        fields = (
            'name',
            'display_ontology',
            'tier'
        )

    @staticmethod
    def get_tier(obj):
        return obj.leveltier.name

class ProgramAuditLogSerializer(ModelSerializer):
    id = IntegerField(allow_null=True, required=False)
    indicator = ProgramAuditLogIndicatorSerializer()
    level = ProgramAuditLogLevelSerializer()
    user = CharField(source='user.name', default='', read_only=True)
    organization = CharField(source='organization.name', default='', read_only=True)
    date = DateTimeField(format="%Y-%m-%d %H:%M:%S", default_timezone=pytz.timezone("UTC"))

    class Meta:
        model = ProgramAuditLog
        fields = (
            'id',
            'date',
            'user',
            'organization',
            'indicator',
            'result_info',
            'change_type',
            'rationale',
            'rationale_selected_options',
            'diff_list',
            'pretty_change_type',
            'level',
        )


class ProgramAdminAuditLogSerializer(ModelSerializer):
    id = IntegerField(allow_null=True, required=False)
    admin_user = SerializerMethodField()
    date = DateTimeField(format="%Y-%m-%d %H:%M:%S")

    class Meta:
        model = ProgramAdminAuditLog
        fields = (
            'id',
            'date',
            'admin_user',
            'change_type',
            'diff_list',
            'pretty_change_type',
        )

    def get_admin_user(self, obj):
        try:
            return obj.admin_user.name
        except AttributeError:
            return ''


class ProgramAdminViewSet(viewsets.ModelViewSet):
    serializer_class = ProgramAdminSerializer
    pagination_class = Paginator
    permission_classes = [permissions.IsAuthenticated, HasProgramAdminAccess]

    @classmethod
    def base_queryset(cls):
        """this adds annotations for the "users" and "organizations" links using annotations

            For performance reasons, looking up every user with permission to see the program individually for
            all 20+ in a paginated set was costly.  This annotates the information in one (admittedly spendy) query
            Note: prefetches here are to allow the serializer to rapidly loop through related objects without
            returning to the db.  Sector and country prefetches are for nested serializers, and program_users
            and country_with_users.country_users are for user and organization counts
        """
        superusers_count = TolaUser.objects.filter(user__is_superuser=True).count()
        queryset = Program.rf_aware_all_objects.all()
        queryset = queryset.annotate(
            su_count=models.Value(superusers_count, output_field=models.IntegerField())
        ).prefetch_related(
            models.Prefetch(
                'sector',
                queryset=Sector.objects.select_related(None).order_by().only('id')
            ),
            models.Prefetch(
                'idaa_outcome_theme',
                queryset=IDAAOutcomeTheme.objects.select_related(None).order_by().only('id', 'name')
            ),
            models.Prefetch(
                'idaa_sector',
                queryset=IDAASector.objects.select_related(None).order_by().only('id')
            ),
            models.Prefetch(
                'country',
                queryset=Country.objects.select_related(None).order_by().only('id')
            ),
            models.Prefetch(
                'programaccess_set',
                queryset=ProgramAccess.objects.filter(
                    tolauser__user__is_superuser=False
                ).select_related('tolauser'),
                to_attr='program_users'
            ),
            models.Prefetch(
                'country',
                queryset=Country.objects.select_related(None).order_by().prefetch_related(
                    models.Prefetch(
                        'countryaccess_set',
                        queryset=CountryAccess.objects.filter(
                            tolauser__user__is_superuser=False
                        ).select_related('tolauser'),
                        to_attr='country_users'
                    ),
                ),
                to_attr='country_with_users'
            ),
            models.Prefetch(
                'gaitid',
                queryset=GaitID.objects.select_related(None).order_by().prefetch_related(
                    models.Prefetch(
                        'fundcode_set',
                        queryset=FundCode.objects.order_by().only('fund_code')
                    )
                )
            )
        )
        return queryset

    def get_queryset(self):
        """Get programs user has access to, annotate for counts, and filter based on provided params from request"""
        auth_user = self.request.user
        params = self.request.query_params

        queryset = self.base_queryset()
        if not auth_user.is_superuser:
            queryset = queryset.filter(country__in=auth_user.tola_user.managed_countries)

        programStatus = params.get('programStatus')
        if programStatus == 'Active':
            queryset = queryset.filter(funding_status='Funded')
        elif programStatus == 'Inactive':
            queryset = queryset.exclude(funding_status='Funded')

        programParam = params.getlist('programs[]')
        if programParam:
            queryset = queryset.filter(pk__in=programParam)

        countryFilter = params.getlist('countries[]')
        if countryFilter:
            queryset = queryset.filter(country__in=countryFilter)

        sectorFilter = params.getlist('idaa_sectors[]')
        if sectorFilter:
            queryset = queryset.filter(idaa_sector__in=sectorFilter)

        usersFilter = params.getlist('users[]')
        if usersFilter:
            # some partner users have a base country assigned, so we need just MC users to filter on country:
            mc_users = list(TolaUser.objects.filter(
                organization_id=Organization.MERCY_CORPS_ID,
                pk__in=usersFilter
            ).values_list('pk', flat=True))
            queryset = queryset.filter(
                models.Q(user_access__id__in=usersFilter) |
                models.Q(country__in=Country.objects.filter(
                    models.Q(users__id__in=mc_users) |
                    models.Q(tolauser__in=mc_users)
                ))
            )
        organizationFilter = params.getlist('organizations[]')
        if organizationFilter:
            queryset = queryset.filter(
                models.Q(user_access__organization__in=organizationFilter) |
                models.Q(country__users__organization__in=organizationFilter)
            )
        return queryset.distinct()

    @action(detail=False, methods=["get"])
    def program_filter_options(self, request):
        """Provides a non paginated list of countries for the frontend filter"""
        auth_user = self.request.user
        params = self.request.query_params
        queryset = Program.rf_aware_all_objects

        if not auth_user.is_superuser:
            tola_user = auth_user.tola_user
            queryset = queryset.filter(
                models.Q(user_access=tola_user) | models.Q(country__users=tola_user)
            )

        countryFilter = params.getlist('countries[]')
        if countryFilter:
            queryset = queryset.filter(country__in=countryFilter)
        programs = [{
            'id': program.id,
            'name': program.name,
        } for program in queryset.distinct().all()]
        return Response(programs)


    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        program = Program.objects.get(pk=pk)
        history = (ProgramAdminAuditLog
            .objects
            .filter(program=program)
            .select_related('admin_user')
            .order_by('-date'))
        serializer = ProgramAdminAuditLogSerializer(history, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["post"])
    def bulk_update_status(self, request):
        ids = request.data.get("ids")
        new_funding_status = request.data.get("funding_status")
        new_funding_status = new_funding_status if new_funding_status in ["Completed", "Funded"] else None
        if new_funding_status:
            to_update = Program.objects.filter(pk__in=ids)
            to_update.update(funding_status=new_funding_status)
            updated = [{
                'id': p.pk,
                'funding_status': p.funding_status,
            } for p in to_update]
            return Response(updated)
        return Response({}, status=httpstatus.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["get"])
    def audit_log(self, request, pk=None):
        program = Program.objects.get(pk=pk)

        queryset = program.audit_logs.all().order_by('-date')
        page = self.paginate_queryset(list(queryset))
        if page is not None:
            serializer = ProgramAuditLogSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def export_audit_log(self, request, pk=None):
        program = Program.objects.get(pk=pk)
        workbook = Workbook()
        workbook.remove(workbook.active)
        # Translators: Sheet title for a log of all changes to indicators over a program's history
        ws = workbook.create_sheet(_('Program change log'))
        get_audit_log_workbook(ws, program)
        response = HttpResponse(content_type='application/ms-excel')
        filename = '{} Program Change Log {}.xlsx'.format(program.name, timezone.now().strftime('%b %d, %Y'))
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
        workbook.save(response)
        return response

    @action(detail=True, methods=['get'], url_path='gait/(?P<gaitid>[^/.]+)')
    def gait(self, request, pk=None, gaitid=None):
        response = {}
        if gaitid is None:
            response['unique'] = True
        else:
            programs = Program.objects.filter(gaitid=gaitid)
            if pk is not None:
                programs = programs.exclude(pk=pk)
            if programs.count() == 0:
                response['unique'] = True
            else:
                response['unique'] = False
                response['gait_link'] = ('https://gait.mercycorps.org/search.vm?Mode=edit&sort_by=g.GrantTitle'
                                         '&q=&kw=&GrantNumber=&CostCenter=&GrantID={0}&GrantMin=&SSD=&USD=&'
                                         'SED=&UED=&Emergency=').format(gaitid)
        return JsonResponse(response)
