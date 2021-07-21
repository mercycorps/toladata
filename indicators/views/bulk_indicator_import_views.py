import django.core.exceptions
import openpyxl
import re
import os
import json
import logging
import decimal
from datetime import datetime
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Alignment, Protection, Font, NamedStyle, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin, AccessMixin
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.utils.translation import gettext, gettext_noop, get_language, override
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.views.decorators.http import require_POST, require_GET

from indicators.models import Indicator, Level, LevelTier, Sector, BulkIndicatorImportFile
from indicators.views.view_utils import indicator_letter_generator
from workflow.models import Program
from workflow.serializers_new import BulkImportSerializer, BulkImportIndicatorSerializer
from tola.l10n_utils import str_without_diacritics
from tola.serializers import make_quantized_decimal
from tola_management.permissions import user_has_program_roles
from tola_management.models import ProgramAuditLog

VALIDATION_KEY_UOM_TYPE = 'uom_type_validation'
VALIDATION_KEY_DIR_CHANGE = 'dir_change_validation'
VALIDATION_KEY_TARGET_FREQ = 'target_frequency_validation'
VALIDATION_KEY_SECTOR = 'sector_validation'
VALIDATION_KEY_BASELINE = 'baseline_validation'

COLUMNS = [
    {'name': 'Level', 'required': True, 'field_name': 'level'},
    {'name': 'No.', 'required': False, 'field_name': 'number',
     'align': 'right'},
    {'name': 'Indicator', 'required': True, 'field_name': 'name'},
    {'name': 'Sector', 'required': False, 'field_name': 'sector',
     'validation': VALIDATION_KEY_SECTOR},
    {'name': 'Source', 'required': False, 'field_name': 'source'},
    {'name': 'Definition', 'required': False, 'field_name': 'definition'},
    {'name': 'Rationale or justification for indicator', 'required': False, 'field_name': 'justification'},
    {'name': 'Unit of measure', 'required': True, 'field_name': 'unit_of_measure'},
    # Note:  this lone string is being translated here because it's not the standard name for the field.
    # Translators:  Column header for the column that specifies whether the data in the row is expressed
    # as a number or a percent
    {'name': 'Number (#) or percentage (%)', 'required': True, 'field_name': 'unit_of_measure_type',
     'validation': VALIDATION_KEY_UOM_TYPE, 'default': 'Number (#)'},
    {'name': 'Rationale for target', 'required': False, 'field_name': 'rationale_for_target'},
    {'name': 'Baseline', 'required': True, 'field_name': 'baseline',
     'align': 'right'},
    {'name': 'Direction of change', 'required': True, 'field_name': 'direction_of_change',
     'validation': VALIDATION_KEY_DIR_CHANGE, 'default': 'Not applicable'},
    {'name': 'Target frequency', 'required': True, 'field_name': 'target_frequency',
     'validation': VALIDATION_KEY_TARGET_FREQ},
    {'name': 'Means of verification / data source', 'required': False, 'field_name': 'means_of_verification'},
    {'name': 'Data collection method', 'required': False, 'field_name': 'data_collection_method'},
    {'name': 'Data points', 'required': False, 'field_name': 'data_points'},
    {'name': 'Responsible person(s) and team', 'required': False, 'field_name': 'responsible_person'},
    {'name': 'Method of analysis', 'required': False, 'field_name': 'method_of_analysis'},
    {'name': 'Information use', 'required': False, 'field_name': 'information_use'},
    {'name': 'Data quality assurance details', 'required': False, 'field_name': 'quality_assurance'},
    {'name': 'Data issues', 'required': False, 'field_name': 'data_issues'},
    {'name': 'Comments', 'required': False, 'field_name': 'comments'}
]

COLUMNS_FIELD_INDEXES = {col['field_name']: i for i, col in enumerate(COLUMNS)}

BASE_TEMPLATE_NAME = 'BulkIndicatorImportTemplate.xlsx'
TEMPLATE_SHEET_NAME = 'Import indicators'
HIDDEN_SHEET_NAME = 'Hidden'

FIRST_USED_COLUMN = 2
PROGRAM_NAME_ROW = 2
COLUMN_HEADER_ROW = 6
DATA_START_ROW = 7

ERROR_MISMATCHED_PROGRAM = 100  # The program name taken from the Excel template doesn't match the current program being used by the user
ERROR_NO_NEW_INDICATORS = 101  # There aren't any new indicators in the template
ERROR_UNDETERMINED_LEVEL = 102  # When the first level header row not parsable into a level
ERROR_TEMPLATE_NOT_FOUND = 103  # Triggered when the system can't find the feedback template associated with the user/program
ERROR_MISMATCHED_TIERS = 104  # The tier names submitted when requesting a template don't match what is in the database
ERROR_INDICATOR_DATA_NOT_FOUND = 105  # When a template has been successfully uploaded but then the temp file can't be found when the save request is made
ERROR_MISMATCHED_LEVEL_COUNT = 106  # The number of level header rows doesn't match the number of levels in the program
ERROR_MISMATCHED_HEADERS = 107  # The column headers don't match the standard template headers
ERROR_SAVE_VALIDATION = 108  # This could happen if e.g. someone adds an indicator
ERROR_INVALID_LEVEL_HEADER = 109  # When the level header doesn't contain an identifiable level
ERROR_MULTIPLE_FEEDBACK_TEMPLATES = 110  # When there are multiple entries for the feedback template, usually happens if there's an error in a prior upload attempt
ERROR_UNEXPECTED_LEVEL = 111  # When the tier in the level column doesn't match what's in the level header
ERROR_UNEXPECTED_INDICATOR_NUMBER = 112  # The indicator number is not sequential in auto-numbered programs
ERROR_MALFORMED_INDICATOR = 201  # A catch-all for problems with an indicator that don't fall into other non-fatal categories

ERROR_INTERVENING_BLANK_ROW = 203  # Indicators are separated by an empty row would cause the indicator numbers to be wrong for auto-numbered programs


EMPTY_CHOICE = '---------'

# Translators: Not applicable. A value that can be entered into a form field when the field doesn't apply in a particular situation.
na = gettext_noop('N/A')
# Translators: An alternate form of N/A or Not applicable
na_alt = gettext_noop('NA')
# Translators: A value that can be entered into a form field when the field doesn't apply in a particular situation.
not_applicable = gettext_noop('Not applicable')
NULL_EQUIVALENTS = [na, na_alt, not_applicable]

GRAY_LIGHTEST = 'FFF5F5F5'
GRAY_LIGHTER = 'FFDBDCDE'
GRAY_DARK = 'FF54585A'
RED_ERROR = '66FFE3E7'

TITLE_STYLE = 'title_style'
REQUIRED_HEADER_STYLE = 'required_header_style'
OPTIONAL_HEADER_STYLE = 'optional_header_style'
LEVEL_HEADER_STYLE = 'level_header_style'
EXISTING_INDICATOR_STYLE = 'existing_indicator_style'
UNPROTECTED_NEW_INDICATOR_STYLE = 'unprotected_new_indicator_style'
PROTECTED_NEW_INDICATOR_STYLE = 'protected_new_indicator_style'

logger = logging.getLogger('__name__')

class BulkImportIndicatorsView(LoginRequiredMixin, UserPassesTestMixin, AccessMixin, View):
    """
    This view is used for the template download and upload process.

    For the download process, it checks the request to see how many rows each tier is configured for and then returns
    a template with the appropriate structure and fillable row numbers.

    For the upload process, it runs though lots of checks to validate the data.  If the validation passes, a json
    file is kept on the file system and a new instance of a BulkIndicatorImportFile model object is created
     that points to the json file.  When the user confirms that they want to save the indicators,
     a different view is used to retrieve the json file and create the Indicator model instances.

    If the validation checks fail, the errors are divided into fatal and non-fatal categories.  Non-fatal errors are
    generally problems with individual indicators that can be highlighted on a spreadsheet and returned to the user.
    In this case, a feedback template is saved to the file system and a new instance of the
    BulkIndicatorImportFile is created that points to the Excel file.  When the user requests the Excel file, it is
    provided to them by a different view.

    Fatal errors prevent meaningful processing of the template and are generally structural problems with the
    template (and shouldn't occur if the user hasn't messed with the spreadsheet protection).  If fatal errors
    occur, no feedback template is saved and a error response is sent with the codes that represent which kind
    of errors occurred.

    """
    redirect_field_name = None

    first_used_column = FIRST_USED_COLUMN
    data_start_row = DATA_START_ROW
    program_name_row = PROGRAM_NAME_ROW

    # Translators: Error message shown to a user when they have entered a value for a numeric field that isn't a number or is negative.
    VALIDATION_MSG_BASELINE_NA = gettext_noop('Enter a numeric value for the baseline. If a baseline is not yet known or not applicable, enter a zero or type "N/A". The baseline can always be updated at a later point in time.')

    default_border = Side(border_style='thin', color=GRAY_LIGHTER)

    title_style = NamedStyle(TITLE_STYLE)
    title_style.font = Font(name='Calibri', size=18)
    title_style.protection = Protection(locked=True)

    required_header_style = NamedStyle(REQUIRED_HEADER_STYLE)
    required_header_style.font = Font(name='Calibri', size=11, color='FFFFFFFF')
    required_header_style.fill = PatternFill('solid', fgColor='00000000')
    required_header_style.protection = Protection(locked=True)
    required_header_style.border = Border(
        left=default_border, right=default_border, top=default_border, bottom=default_border)

    optional_header_style = NamedStyle(OPTIONAL_HEADER_STYLE)
    optional_header_style.font = Font(name='Calibri', size=11, color='00000000')
    optional_header_style.fill = PatternFill('solid', fgColor=GRAY_LIGHTEST)
    optional_header_style.protection = Protection(locked=False)
    optional_header_style.border = Border(
        left=default_border, right=default_border, top=default_border, bottom=default_border)

    level_header_style = NamedStyle(LEVEL_HEADER_STYLE)
    level_header_style.fill = PatternFill('solid', fgColor=GRAY_LIGHTER)
    level_header_style.font = Font(name='Calibri', size=11, bold=True)
    level_header_style.protection = Protection(locked=True)

    existing_indicator_style = NamedStyle(EXISTING_INDICATOR_STYLE)
    existing_indicator_style.fill = PatternFill('solid', fgColor=GRAY_LIGHTEST)
    existing_indicator_style.font = Font(name='Calibri', size=11, color=GRAY_DARK)
    existing_indicator_style.border = Border(
        left=default_border, right=default_border, top=default_border, bottom=default_border)
    existing_indicator_style.alignment = Alignment(vertical='top', wrap_text=True)
    existing_indicator_style.protection = Protection(locked=True)

    unprotected_new_indicator_style = NamedStyle(UNPROTECTED_NEW_INDICATOR_STYLE)
    unprotected_new_indicator_style.font = Font(name='Calibri', size=11)
    unprotected_new_indicator_style.alignment = Alignment(vertical='top', wrap_text=True)
    unprotected_new_indicator_style.protection = Protection(locked=False)

    protected_new_indicator_style = NamedStyle(PROTECTED_NEW_INDICATOR_STYLE)
    protected_new_indicator_style.font = Font(name='Calibri', size=11)
    protected_new_indicator_style.alignment = Alignment(vertical='top', wrap_text=True)
    protected_new_indicator_style.protection = Protection(locked=True)

    def test_func(self):
        program_id = self.request.resolver_match.kwargs.get('program_id')
        return user_has_program_roles(self.request.user, [program_id], ['high']) \
            or self.request.user.is_superuser is True

    def _row_is_blank_or_spacer(self, ws, row_index, program):
        # Blank rows are rows that are are configured to contain data but don't have any (i.e. they contain
        # pre-filled level, number, etc....  In this case there will be four pre-filled cells.
        # Spacer rows are completely blank and should only precede a level header.  The number
        # field is being ignored on purpose.  If it's the only thing filled in for a row (in addition to the
        # other three fields), the row id considered blank.
        pre_filled_field_names = ['level', 'unit_of_measure_type', 'direction_of_change']
        pre_filled_indexes = [index for index, col in enumerate(COLUMNS)
            if col['field_name'] in pre_filled_field_names]
        row_filled_indexes = [index for index, col in enumerate(COLUMNS)
            if ws.cell(row_index, self.first_used_column + index).value and col['field_name'] != 'number']
        if pre_filled_indexes == row_filled_indexes:
            return 'blank'
        if len(row_filled_indexes) == 0:
            return 'spacer'
        return False

    @staticmethod
    def setup_validations(hidden_ws):
        validation_map = {}
        uom_type_options = [str(option[1]) for option in Indicator.UNIT_OF_MEASURE_TYPES]
        uom_type_validation = DataValidation(
            type="list", formula1=f'"{",".join(uom_type_options)}"', allow_blank=False)
        validation_map[VALIDATION_KEY_UOM_TYPE] = uom_type_validation

        dir_change_options = [str(option[1]) for option in Indicator.DIRECTION_OF_CHANGE]
        dir_change_validation = DataValidation(
            type="list", formula1=f'"{",".join(dir_change_options)}"', allow_blank=True)
        validation_map[VALIDATION_KEY_DIR_CHANGE] = dir_change_validation

        target_freq_options = [str(option[1]) for option in Indicator.TARGET_FREQUENCIES]
        target_freq_validation = DataValidation(
            type="list", formula1=f'"{",".join(target_freq_options)}"', allow_blank=False)
        validation_map[VALIDATION_KEY_TARGET_FREQ] = target_freq_validation

        for dv in validation_map.values():
            # Translators: Error message shown to a user when they have entered a value that is not in a list of approved values
            dv.error = gettext('Your entry is not in the list')
            # Translators:  Title of a popup box that informs the user they have entered an invalid value
            dv.errorTitle = gettext('Invalid Entry')

        sector_options = [EMPTY_CHOICE] + sorted(
            [gettext(sector.sector) for sector in Sector.objects.all()],
            key=lambda choice: str_without_diacritics(choice))
        sectors_col = 'A'
        sectors_start_row = 2
        sectors_end_row = sectors_start_row + len(sector_options) - 1
        for i, sector in enumerate(sector_options):
            hidden_ws[f'{sectors_col}{i + 2}'].value = sector
        sector_range = f'${sectors_col}${sectors_start_row}:${sectors_col}${sectors_end_row}'
        sector_validation = DataValidation(
            type='list', formula1=f'{HIDDEN_SHEET_NAME}!{sector_range}', allow_blank=True)
        validation_map[VALIDATION_KEY_SECTOR] = sector_validation
        return validation_map, sector_options

    def apply_validations_to_row(self, ws, row_index, validation_map):
        # TODO: limit character count of cells with max_length
        for col_num, col in enumerate(COLUMNS):
            active_cell = ws.cell(row_index, FIRST_USED_COLUMN + col_num)
            if 'validation' in col.keys():
                validation_map[col['validation']].add(active_cell)

            if col['field_name'] == 'baseline':
                cell_coord = active_cell.coordinate
                null_check_strings = [f'lower({cell_coord})=lower("{ne}")' for ne in NULL_EQUIVALENTS]
                active_language = get_language()
                if active_language != 'en':
                    null_check_strings.extend(
                        [f'lower({cell_coord})=lower("{gettext(ne)}")' for ne in NULL_EQUIVALENTS])
                formula = f'OR(AND(ISNUMBER({cell_coord}), {cell_coord}>=0), {",".join(null_check_strings)})'
                validator = DataValidation(type='custom', formula1=formula)
                validator.error = gettext(self.VALIDATION_MSG_BASELINE_NA)
                # Translators:  Title of a popup box that informs the user they have entered an invalid value
                validator.errorTitle = gettext('Invalid Entry')
                ws.add_data_validation(validator)
                validator.add(active_cell)

    @staticmethod
    def get_comment_obj(help_text):
        comment = Comment(help_text, '')
        comment.width = 300
        comment.height = 15 + len(help_text) * .5
        return comment

    def get(self, request, *args, **kwargs):
        program_id = kwargs['program_id']
        try:
            program = Program.objects.get(pk=program_id)
        except Program.DoesNotExist:
            return JsonResponse({'error_codes': [ERROR_MISMATCHED_PROGRAM]}, status=404)

        levels = Level.objects.filter(program=program).select_related().prefetch_related(
            'indicator_set', 'program', 'parent__program__level_tiers')
        serialized_levels = sorted(BulkImportSerializer(levels, many=True).data, key=lambda level: level['ontology'])

        leveltiers_in_db = sorted([tier.name for tier in LevelTier.objects.filter(program=program)])
        request_leveltier_names = sorted(request.GET.keys())
        if leveltiers_in_db != request_leveltier_names:
            return JsonResponse({'error_codes': [ERROR_MISMATCHED_TIERS]}, status=400)

        wb = openpyxl.load_workbook(filename=BulkIndicatorImportFile.get_file_path(BASE_TEMPLATE_NAME))
        ws = wb.worksheets[0]
        hidden_ws = wb.get_sheet_by_name('Hidden')

        # Create data validations
        reverse_field_name_map = {
            'unit_of_measure_type': {str(name): value for value, name in Indicator.UNIT_OF_MEASURE_TYPES},
            'direction_of_change':  {str(name): value for value, name in Indicator.DIRECTION_OF_CHANGE},
            'target_frequency': {str(name): value for value, name in Indicator.TARGET_FREQUENCIES}
        }


        validation_map, sector_options = self.setup_validations(hidden_ws)

        # Spreadsheet loads as corrupted when validation is added to a sheet but no cells are assigned to the
        # validation. So we'll check that there is at least one blank row that will be created.
        blank_row_count = sum([int(request.GET[level['tier_name']]) for level in serialized_levels])
        if blank_row_count > 0:
            for  validation in validation_map.values():
                ws.add_data_validation(validation)

        wb.add_named_style(self.title_style)
        wb.add_named_style(self.level_header_style)
        wb.add_named_style(self.existing_indicator_style)
        wb.add_named_style(self.unprotected_new_indicator_style)
        wb.add_named_style(self.protected_new_indicator_style)
        wb.add_named_style(self.optional_header_style)
        wb.add_named_style(self.required_header_style)

        last_used_column = self.first_used_column + len(COLUMNS) - 1
        # Translators: Heading placed in the cell of a spreadsheet that allows users to upload Indicators in bulk
        ws.cell(1, self.first_used_column).value = gettext("Import indicators")
        ws.cell(self.program_name_row, self.first_used_column).value = program.name
        ws.cell(2, self.first_used_column).style = TITLE_STYLE
        # Translators: Instructions provided as part of an Excel template that allows users to upload Indicators
        instructions = gettext("INSTRUCTIONS\n"
            "1. Indicator rows are provided for each result level. You can delete indicator rows you do not need. You can also leave them empty and they will be ignored.\n"
            "2. Required columns are highlighted with a dark background and an asterisk (*) in the header row. Unrequired columns can be left empty but cannot be deleted.\n"
            "3. When you are done, upload the template to the results framework or program page."
        )

        ws.cell(4, self.first_used_column).value = instructions
        ws.cell(4, self.first_used_column).fill = PatternFill('solid', fgColor=GRAY_LIGHTER)
        # Translators: Section header of an Excel template that allows users to upload Indicators.  This section is where users will add their own information.
        ws.cell(5, self.first_used_column).value = gettext("Enter indicators")

        # Output column header row
        for i, col in enumerate(COLUMNS):
            column_index = i+2
            header_cell = ws.cell(6, column_index)
            if col['required']:
                header_cell.value = gettext(col['name']) + "*"
                header_cell.style = REQUIRED_HEADER_STYLE
            else:
                header_cell.value = gettext(col['name'])
                header_cell.style = OPTIONAL_HEADER_STYLE

            # Add notes to header row cells
            if col['field_name'] != 'comments':
                if col['field_name'] == 'number' and program.auto_number_indicators:
                    # Translators: This is help text for a form field that gets filled in automatically
                    help_text = gettext('This number is automatically generated through the results framework.')
                elif col['field_name'] == 'baseline':
                    # Translators: This is help text for a form field
                    help_text = gettext('Enter a numeric value for the baseline. If a baseline is not yet known '
                                        'or not applicable, enter a zero or type "N/A". The baseline can always '
                                        'be updated at a later point in time.')
                else:
                    help_text = Indicator._meta.get_field(col['field_name']).help_text
                header_cell.comment = self.get_comment_obj(help_text)

            # Apply specified alignment
            if 'align' in col.keys():
                header_cell.alignment = Alignment(horizontal='right')

            # Add other column-specific styling
            if col['field_name'] == 'sector':
                col_width = max(len(option) for option in sector_options)
                ws.column_dimensions[get_column_letter(column_index)].width = col_width
            elif col['field_name'] == 'definition':
                ws.column_dimensions[get_column_letter(column_index)].width = 50
            elif col['field_name'] in ['source', 'justification', 'unit_of_measure', 'rationale_for_target']:
                ws.column_dimensions[get_column_letter(column_index)].width = 40
            elif col['field_name'] == 'baseline':
                ws.column_dimensions[get_column_letter(column_index)].width = 20
            elif 7 < column_index < 15:
                col_width = len(header_cell.value)
                ws.column_dimensions[get_column_letter(column_index)].width = col_width
            elif column_index >= 15:
                ws.column_dimensions[get_column_letter(column_index)].width = 40

        # Need to make sure the instructions are covered by merged cells and that cell widths are calculated
        # after the individual header cells have been styled
        max_line_width = max([len(line) for line in instructions.split('\n')])
        required_width = .8 * max_line_width
        merged_width = 0
        max_merge_col_index = 2
        while merged_width < required_width:
            merged_width += ws.column_dimensions[get_column_letter(max_merge_col_index)].width
            max_merge_col_index += 1
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=max_merge_col_index - 1)

        current_row_index = self.data_start_row
        for level in serialized_levels:
            displayed_ontology = ' ' + level['display_ontology'] if len(level['display_ontology']) > 0 else ''
            ws.cell(current_row_index, self.first_used_column)\
                .value = f"{gettext(level['tier_name'])}{displayed_ontology}: {level['level_name']}"
            ws.merge_cells(
                start_row=current_row_index,
                start_column=self.first_used_column,
                end_row=current_row_index,
                end_column=last_used_column)
            ws.cell(current_row_index, self.first_used_column).style = LEVEL_HEADER_STYLE
            current_row_index += 1

            for indicator in level['indicator_set']:
                current_column_index = self.first_used_column
                first_indicator_cell = ws.cell(current_row_index, current_column_index)
                first_indicator_cell.value = gettext(level['tier_name'])
                first_indicator_cell.style = EXISTING_INDICATOR_STYLE
                first_indicator_cell.border = Border(
                    left=None, right=self.default_border, top=self.default_border, bottom=self.default_border)

                current_column_index += 1
                ws.row_dimensions[current_row_index].height = 16
                for column in COLUMNS[1:]:
                    active_cell = ws.cell(current_row_index, current_column_index)
                    if column['field_name'] == 'number':
                        if program.auto_number_indicators:
                            active_cell.value = \
                                level['display_ontology'] + indicator['display_ontology_letter']
                        else:
                            active_cell.value = indicator['number']
                    else:
                        raw_value = indicator.get(column['field_name'], None)
                        if column['field_name'] == 'baseline':
                            raw_value = make_quantized_decimal(raw_value)
                        # These are potentially translated objects that need to be resolved, but converting None
                        # to a string results in a cell value of "None" rather than an empty cell.
                        active_cell.value = raw_value if raw_value is None else str(raw_value)
                    active_cell.style = EXISTING_INDICATOR_STYLE
                    if 'align' in column.keys():
                        active_cell.alignment = Alignment(horizontal='right')
                    if column['field_name'] == 'baseline':
                        active_cell.number_format = '0.00'
                    current_column_index += 1
                current_row_index += 1

            letter_gen = indicator_letter_generator(len(level['indicator_set']) + 1)
            empty_row_count = int(request.GET[level['tier_name']])
            for i in range(empty_row_count):
                self.apply_validations_to_row(ws, current_row_index, validation_map)
                ws.row_dimensions[current_row_index].height = 16
                ws.cell(current_row_index, self.first_used_column).value = gettext(level['tier_name'])
                if program.auto_number_indicators:
                    ws.cell(current_row_index, FIRST_USED_COLUMN + COLUMNS_FIELD_INDEXES['number']).value = \
                        level['display_ontology'] + next(letter_gen)

                for col_n, column in enumerate(COLUMNS):
                    active_cell = ws.cell(current_row_index, self.first_used_column + col_n)
                    if column['field_name'] == 'level' or \
                            (column['field_name'] == 'number' and program.auto_number_indicators):
                        active_cell.style = PROTECTED_NEW_INDICATOR_STYLE
                    else:
                        active_cell.style = UNPROTECTED_NEW_INDICATOR_STYLE

                    if 'default' in column:
                        active_cell.value = gettext(column['default'])
                    if column['field_name'] in ['number', 'baseline']:
                        active_cell.alignment = Alignment(horizontal='right')
                    if column['field_name'] == 'baseline':
                        active_cell.number_format = '0.00'

                current_row_index += 1
            current_row_index += 1  # Leave one row blank at the end of each level section

        ws.protection.enable()
        ws.title = gettext(TEMPLATE_SHEET_NAME)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(BASE_TEMPLATE_NAME)
        wb.save(response)
        return response

    def post(self, request, *args, **kwargs):
        # Translators: Message provided to user when they have failed to enter a required field on a form.
        VALIDATION_MSG_REQUIRED = gettext('This information is required.')
        # Translators: Message provided to user when they have not chosen from a pre-selected list of options.
        VALIDATION_MSG_CHOICE = gettext('The {field_name} you selected is unavailable. Please select a different {field_name}.')
        FIRST_CELL_ERROR_VALUE = r"⚠️"
        PATTERN_FILL_ERROR = PatternFill('solid', fgColor=RED_ERROR)

        program_id = kwargs['program_id']
        program = Program.objects.get(pk=program_id)
        ProgramAuditLog.log_template_uploaded(request.user, program)

        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.worksheets[0]

        if ws.cell(self.program_name_row, self.first_used_column).value != program.name:
            return JsonResponse({'error_codes': [ERROR_MISMATCHED_PROGRAM]}, status=406)

        # The template could be any language so we're going to kill two birds with one stone by using the
        # column headers to figure out what the template language is.  If the headers don't match any of the
        # translated header sets, we know we've got a problem with the headers.  If they do match one of the
        # language sets, we'll know which language the template is in.
        headers_by_lang = {}
        language_codes = ['en', 'fr', 'es']
        template_headers = [ws.cell(DATA_START_ROW - 1, FIRST_USED_COLUMN + i).value.strip('*')
                            for i in range(len(COLUMNS))]
        template_language = None
        for lang_code in language_codes:
            with override(lang_code):
                for col_index, col in enumerate(COLUMNS):
                    try:
                        headers_by_lang[lang_code].append(gettext(col['name']))
                    except KeyError:
                        headers_by_lang[lang_code] = [gettext(col['name'])]
        for lang_code in language_codes:
            if template_headers == headers_by_lang[lang_code]:
                template_language = lang_code
                break
        if not template_language:
            return JsonResponse({'error_codes': [ERROR_MISMATCHED_HEADERS]}, status=406)

        # Set the language for template processing so the in-Excel error messages will match the original language.
        with override(template_language):
            level_refs = {}
            for level in program.levels.all():
                ontology = f' {level.display_ontology}' if len(level.display_ontology) > 0 else ''
                level_refs[f'{gettext(str(level.leveltier))}{ontology}'] = level

            non_fatal_errors = []  # Errors that can be highlighted on a spreadsheet and get sent back to the user
            fatal_errors = []  # So bad that it's not possible to highlight a spreadsheet and send it back to the suer
            current_level = None
            current_tier = None
            new_indicators_data = []
            indicator_status = {'valid': 0, 'invalid': 0}
            reverse_field_name_map = {
                'unit_of_measure_type': {str(name): value for value, name in Indicator.UNIT_OF_MEASURE_TYPES},
                'direction_of_change': {str(name): value for value, name in Indicator.DIRECTION_OF_CHANGE},
                'target_frequency': {str(name): value for value, name in Indicator.TARGET_FREQUENCIES}
            }
            letter_gen = None
            pattern = r'^([^:]+)'
            level_finder = re.compile(pattern)
            ws_level_count = 0
            skipped_row_indexes = []

            # Need to wipe all validations and re-add them just in case user has modified or deleted them.
            # openpyxl doesn't make it easy to identify which validation is which, so it's easier just to wipe and
            # recreate
            ws.data_validations = openpyxl.worksheet.datavalidation.DataValidationList()  # wipes existing validations
            hidden_ws = wb.get_sheet_by_name('Hidden')
            validation_map, sector_options = self.setup_validations(hidden_ws)
            for validation in validation_map.values():
                ws.add_data_validation(validation)

            for current_row_index in range(self.data_start_row, ws.max_row):
                first_cell = ws.cell(current_row_index, self.first_used_column)
                # If this is a level header row, parse the level name out of the header string
                if first_cell.style == LEVEL_HEADER_STYLE:
                    matches = level_finder.match(first_cell.value)
                    if not matches or matches.group(1).strip() not in level_refs.keys():
                        fatal_errors.append(ERROR_INVALID_LEVEL_HEADER)
                        continue
                    else:
                        tier_label = matches.group(1).strip()
                        current_level = level_refs[tier_label]
                        current_tier = tier_label
                        if tier_matches := re.search(r'(.*)(\s[\d\.]+$)', tier_label):  # Separate tier name from ontology
                            current_tier = tier_matches.group(1).strip()
                        letter_gen = indicator_letter_generator()
                        ws_level_count += 1
                        skipped_row_indexes = []
                        continue

                # Getting to this point without a parsed level header means something has gone wrong
                if not current_level:
                    fatal_errors.append(ERROR_UNDETERMINED_LEVEL)
                    return JsonResponse({'error_codes': fatal_errors}, status=406)

                # skip indicators that are already in the system
                if ws.cell(current_row_index, self.first_used_column).style == EXISTING_INDICATOR_STYLE:
                    next(letter_gen)
                    continue

                # Remove existing error highlighting and comments, in case user has upload a feedback template
                # Also add back dropdowns if they are missing.
                ws.cell(current_row_index, 1).value = None
                for c in range(1, len(COLUMNS) + 1):
                    cell_to_clean = ws.cell(current_row_index, c)
                    cell_to_clean.fill = PatternFill(fill_type=None)
                    cell_to_clean.comment = None

                # Skip blank rows but if there was an intervening blank row, give it error highlighting
                if blank_type := self._row_is_blank_or_spacer(ws, current_row_index, program):
                    if blank_type == 'blank':
                        self.apply_validations_to_row(ws, current_row_index, validation_map)
                        next(letter_gen)
                    skipped_row_indexes.append(current_row_index)
                    continue
                elif len(skipped_row_indexes) > 0:
                    non_fatal_errors.append(ERROR_INTERVENING_BLANK_ROW)
                    for skipped_row_index in skipped_row_indexes:
                        first_cell = ws.cell(skipped_row_index, 1)
                        # Translators: Error message provided to users when they are entering data into Excel and they skip a row
                        first_cell.comment = self.get_comment_obj(gettext('Indicator rows cannot be skipped.'))
                        first_cell.value = FIRST_CELL_ERROR_VALUE
                        for col_index in range(1, self.first_used_column + len(COLUMNS)):
                            ws.cell(skipped_row_index, col_index).fill = PATTERN_FILL_ERROR
                    skipped_row_indexes = []

                self.apply_validations_to_row(ws, current_row_index, validation_map)

                # Check if auto-numbered indicator numbers match what's expected
                number_cell = ws.cell(current_row_index, self.first_used_column + 1)
                if program.auto_number_indicators and number_cell.value is not None:
                    expected_ind_number = f'{current_level.display_ontology}{next(letter_gen)}'
                    if number_cell.value != expected_ind_number:
                        fatal_errors.append(ERROR_UNEXPECTED_INDICATOR_NUMBER)

                # Check if the value of the Level column matches the one used in the level header
                level_cell = ws.cell(current_row_index, FIRST_USED_COLUMN)
                if level_cell.value != current_tier:
                    level_cell.fill = PatternFill('solid', fgColor=RED_ERROR)
                    fatal_errors.append(ERROR_UNEXPECTED_LEVEL)

                indicator_data = {}
                for i, column in enumerate(COLUMNS[1:]):
                    indicator_data[column['field_name']] = \
                        ws.cell(current_row_index, self.first_used_column + i + 1).value

                validation_errors = {}
                # Use the serializer to check for character length and any other field issues that it handles
                deserialized_data = BulkImportIndicatorSerializer(data=indicator_data)
                deserialized_data.is_valid()
                indicator_data['baseline'] = deserialized_data.data['baseline']

                # Capture validation problems from the deserialization process
                for field_name, error_list in deserialized_data.errors.items():
                    for error in error_list:
                        if error.code == 'null':
                            # We're checking for all required fields later, so counting them here would be duplicative
                            continue
                        else:
                            validation_errors[field_name] = []
                        if error.code == 'max_length':
                            # System generated error message is 'Ensure this field has no more than 500 characters.'
                            matches = re.search(r'(\d+)', str(error))
                            if matches:
                                # Translators: Error message provided when user has exceeded the character limit on a form
                                msg = gettext(f'Please enter {matches.group(1)} or fewer characters.')
                                validation_errors[field_name].append(msg)
                            else:
                                validation_errors[field_name].append(
                                    # Translators: Error message provided when user has exceeded the character limit on a form
                                    gettext("You have exceeded the character limit of this field"))
                                logger.error(f'New validation string of code "{error.code}" found.\n{str(error)}')
                        elif str(error) == 'not_a_number' and error.code == 'invalid':
                            # This means the baseline value may not be valid.  The baseline value is handled further along in the process.
                            pass
                        else:
                            logger.error(f'New validation string of code {error.code} found.\n{str(error)}')
                            validation_errors[field_name].append(str(error))

                # Check for missing required fields and tinker with baseline field
                for i, column in enumerate(COLUMNS):
                    active_cell = ws.cell(current_row_index, FIRST_USED_COLUMN + i)
                    if column['field_name'] == 'baseline':
                        if 'baseline' in validation_errors:  # Means it's either None or a text string
                            null_equivalents_with_translations = [ne.lower() for ne in NULL_EQUIVALENTS] + \
                                [gettext(ne).lower() for ne in NULL_EQUIVALENTS]
                            if active_cell.value is None:  # This means no value in cell, which is not ok
                                validation_errors['baseline'] = [VALIDATION_MSG_REQUIRED]
                                non_fatal_errors.append(ERROR_MALFORMED_INDICATOR)
                            elif active_cell.value.lower() in null_equivalents_with_translations:
                                validation_errors.pop('baseline')
                                active_cell.value = None
                            else:  # cell value is not None or one of the acceptable null equivalent values
                                validation_errors['baseline'] = [gettext(self.VALIDATION_MSG_BASELINE_NA)]
                                non_fatal_errors.append(ERROR_MALFORMED_INDICATOR)
                        else:
                            try:
                                indicator_data['baseline'] = make_quantized_decimal(active_cell.value)
                            except (ValueError, decimal.InvalidOperation):
                                # User may have copied a bad value in to the cell
                                validation_errors['baseline'] = [VALIDATION_MSG_REQUIRED]
                                non_fatal_errors.append(ERROR_MALFORMED_INDICATOR)
                    elif column['required'] and active_cell.value is None:
                        try:
                            validation_errors[column['field_name']].append(VALIDATION_MSG_REQUIRED)
                        except KeyError:
                            validation_errors[column['field_name']] = [VALIDATION_MSG_REQUIRED]
                        non_fatal_errors.append(ERROR_MALFORMED_INDICATOR)

                # Convert enumerated fields into numerical ids
                for field in reverse_field_name_map.keys():
                    col_number = self.first_used_column + COLUMNS_FIELD_INDEXES[field]
                    cell_value = ws.cell(current_row_index, col_number).value
                    if cell_value is None:
                        pass
                    else:
                        try:
                            indicator_data[field] = reverse_field_name_map[field][cell_value]
                        except KeyError:
                            display_field_name = COLUMNS[COLUMNS_FIELD_INDEXES[field]]['name']
                            message = VALIDATION_MSG_CHOICE.format(field_name=display_field_name.lower())
                            if field in validation_errors:
                                validation_errors[field].append(message)
                            else:
                                validation_errors[field] = [message]
                            non_fatal_errors.append(ERROR_MALFORMED_INDICATOR)

                # Convert text in sector dropdown to sector pk
                col_number = self.first_used_column + COLUMNS_FIELD_INDEXES['sector']
                cell_value = ws.cell(current_row_index, col_number).value
                if cell_value is None or cell_value in [EMPTY_CHOICE, '', 'None']:
                    pass
                else:
                    try:
                        sector_translation_map = {str(gettext(sector.sector)): sector.sector
                            for sector in Sector.objects.all()}
                        sector = Sector.objects.get(sector=sector_translation_map[cell_value])
                        indicator_data['sector_id'] = sector.pk
                        indicator_data.pop('sector')
                    except (Sector.DoesNotExist, KeyError):
                        error_string = VALIDATION_MSG_CHOICE.format(field_name=gettext('Sector'))
                        try:
                            validation_errors['sector'].append(error_string)
                        except KeyError:
                            validation_errors['sector'] = [error_string]
                        non_fatal_errors.append(ERROR_MALFORMED_INDICATOR)

                # Final data manipulation and updates
                indicator_data['level_id'] = current_level.pk
                indicator_data['program_id'] = program.id
                # The number order has already been checked if program is autonumbered, so we don't need the numbers
                # any more
                if program.auto_number_indicators:
                    indicator_data.pop('number')

                if len(validation_errors) == 0:
                    new_indicators_data.append(indicator_data)
                    indicator_status['valid'] += 1
                else:
                    for field, error_strings in validation_errors.items():
                        column_index = self.first_used_column + COLUMNS_FIELD_INDEXES[field]
                        error_cell = ws.cell(current_row_index, column_index)
                        error_cell.fill = PATTERN_FILL_ERROR
                        error_cell.comment = self.get_comment_obj('\n'.join(error_strings))
                        ws.cell(current_row_index, 1).value = FIRST_CELL_ERROR_VALUE
                        ws.cell(current_row_index, 1).fill = PATTERN_FILL_ERROR
                    indicator_status['invalid'] += 1

            if len(level_refs) != ws_level_count:
                return JsonResponse({'error_codes': [ERROR_MISMATCHED_LEVEL_COUNT]}, status=406)

            if len(fatal_errors) > 0:
                pruned_fatal = sorted(list(set(fatal_errors)))
                return JsonResponse({'error_codes': pruned_fatal}, status=406)

        if indicator_status['invalid'] > 0 or len(non_fatal_errors) > 0:
            # Clean out any existing temp Excel file reference, since we're about to replace it
            try:
                old_file_objs = BulkIndicatorImportFile.objects.filter(
                    user=request.user.tola_user,
                    program=program,
                    file_type=BulkIndicatorImportFile.INDICATOR_TEMPLATE_TYPE)
                for old_file_obj in old_file_objs:
                    os.remove(old_file_obj.file_path)
                    old_file_obj.delete()
            except BulkIndicatorImportFile.DoesNotExist:
                pass

            file_name = f'{datetime.strftime(datetime.now(), "%Y%m%d-%H%M%S")}-{request.user.id}-{program.pk}.xlsx'
            pruned_non_fatal = sorted(list(set(non_fatal_errors)))
            file_obj = BulkIndicatorImportFile.objects.create(
                user=request.user.tola_user,
                program=program,
                file_name=file_name,
                file_type=BulkIndicatorImportFile.INDICATOR_TEMPLATE_TYPE)
            with open(file_obj.file_path, 'w') as fh:
                wb.save(file_obj.file_path)
            return JsonResponse({
                'invalid': indicator_status['invalid'],
                'valid': indicator_status['valid'],
                'error_codes': pruned_non_fatal
            }, status=400)

        if len(new_indicators_data) == 0:
            return JsonResponse({'error_codes': [ERROR_NO_NEW_INDICATORS]}, status=406)

        try:
            old_file_obj = BulkIndicatorImportFile.objects.get(
                user=request.user.tola_user,
                program=program,
                file_type=BulkIndicatorImportFile.INDICATOR_DATA_TYPE)
            os.remove(old_file_obj.file_path)
            old_file_obj.delete()
        except (BulkIndicatorImportFile.DoesNotExist, FileNotFoundError):
            pass

        file_name = f'{datetime.strftime(datetime.now(), "%Y%m%d-%H%M%S")}-{request.user.id}-{program.pk}.json'
        file_obj = BulkIndicatorImportFile.objects.create(
            user=request.user.tola_user,
            program=program,
            file_name=file_name,
            file_type=BulkIndicatorImportFile.INDICATOR_DATA_TYPE)
        with open(file_obj.file_path, 'w') as fh:
            fh.write(json.dumps(new_indicators_data, cls=DjangoJSONEncoder))
        return JsonResponse({'message': 'success', 'valid': len(new_indicators_data), 'invalid': 0}, status=200)


@login_required()
@require_POST
def save_bulk_import_data(request, *args, **kwargs):
    try:
        file_entry = BulkIndicatorImportFile.objects.get(
            user=request.user.tola_user,
            program_id=kwargs['program_id'],
            file_type=BulkIndicatorImportFile.INDICATOR_DATA_TYPE)
    except BulkIndicatorImportFile.DoesNotExist:
        return JsonResponse({'error_codes': [ERROR_INDICATOR_DATA_NOT_FOUND]})

    with open(file_entry.file_path, 'r') as fh:
        stored_indicators = json.loads(fh.read())
    os.remove(file_entry.file_path)
    file_entry.delete()

    try:
        with transaction.atomic():
            for indicator_data in stored_indicators:
                indicator = Indicator.objects.create(**indicator_data)
                ProgramAuditLog.log_indicator_imported(request.user, indicator, 'N/A')
    except django.core.exceptions.ValidationError:
        return JsonResponse({'error_codes': [ERROR_SAVE_VALIDATION]}, status=400)
    return JsonResponse({'message': 'success'}, status=200)


@login_required()
@require_GET
def get_feedback_bulk_import_template(request, *args, **kwargs):
    try:
        file_entry = BulkIndicatorImportFile.objects.get(
            user=request.user.tola_user,
            program_id=kwargs['program_id'],
            file_type=BulkIndicatorImportFile.INDICATOR_TEMPLATE_TYPE)
    except BulkIndicatorImportFile.DoesNotExist:
        return JsonResponse({'error_codes': [ERROR_TEMPLATE_NOT_FOUND]}, status=400)
    except BulkIndicatorImportFile.MultipleObjectsReturned:
        file_entries = BulkIndicatorImportFile.objects.filter(
            user=request.user.tola_user,
            program_id=kwargs['program_id'],
            file_type=BulkIndicatorImportFile.INDICATOR_TEMPLATE_TYPE)
        file_entries.delete()
        return JsonResponse({'error_codes': [ERROR_MULTIPLE_FEEDBACK_TEMPLATES]}, status=400)


    with open(file_entry.file_path, 'rb') as fh:
        template_file = fh.read()
    os.remove(file_entry.file_path)
    file_entry.delete()
    response = HttpResponse(template_file, content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format('Marked-up-template.xlsx')
    return response
