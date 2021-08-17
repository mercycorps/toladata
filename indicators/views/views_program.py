# -*- coding: utf-8 -*-
"""
    Program views: logframe, program page api, etc.
"""

import math
from operator import itemgetter
import csv
import datetime
import logging
import openpyxl
from openpyxl import styles, utils
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.utils.translation import gettext
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, reverse, redirect, get_object_or_404
from django.db.models import Max
from indicators.queries import ProgramWithMetrics
from indicators.xls_export_utils import TAN, apply_title_styling, apply_label_styling
from indicators.models import Indicator, PinnedReport, Level
from workflow.models import Program, Country
from workflow.serializers import LogframeProgramSerializer
from workflow.serializers_new import (
    ProgramPageProgramSerializer,
    ProgramPageIndicatorUpdateSerializer,
)

from tola_management.permissions import (
    has_program_read_access,
    indicator_pk_adapter,
    has_indicator_read_access
)
from tola_management.models import ProgramAuditLog


logger = logging.getLogger(__name__)

TITLE_FONT = openpyxl.styles.Font(size=18)
HEADER_FONT = openpyxl.styles.Font(bold=True)
HEADER_FILL = openpyxl.styles.PatternFill('solid', TAN)
TOP_LEFT_ALIGN_WRAP = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)
LEVEL_ROW_FILL = openpyxl.styles.PatternFill('solid', TAN)
BLACK_BORDER = openpyxl.styles.Side(border_style="thin", color="000000")
BORDER_TOP = openpyxl.styles.Border(top=BLACK_BORDER)

def add_title_cell(ws, row, column, value):
    cell = ws.cell(row=row, column=column)
    cell.value = value
    cell.font = TITLE_FONT
    apply_title_styling(cell)
    return cell

def add_header_cell(ws, row, column, value):
    cell = ws.cell(row=row, column=column)
    cell.value = value.upper()
    apply_label_styling(cell)
    ws.row_dimensions[row].height = 30
    return cell

def get_child_levels(level, levels_by_pk):
    levels = [level]
    for child_pk in level['child_levels']:
        levels += get_child_levels(levels_by_pk[child_pk], levels_by_pk)
    return levels


@login_required
@has_program_read_access
def logframe_view(request, program):
    """
    Logframe view
    """
    serialized_program = LogframeProgramSerializer.load(program)
    context = {
        'js_context': serialized_program.data
    }
    return render(request, 'indicators/logframe/logframe.html', context)


@login_required
@has_program_read_access
def logframe_excel_view(request, program):
    """
    Logframe Excel Export view
    """
    program = LogframeProgramSerializer.load(program).data
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(gettext('Logframe'))
    add_title_cell(ws, 1, 1, program['name'])
    ws.merge_cells(
        start_row=1, end_row=1,
        start_column=1, end_column=4
    )
    add_title_cell(ws, 2, 1, gettext('Logframe'))
    ws.merge_cells(
        start_row=2, end_row=2,
        start_column=1, end_column=4
    )
    for col, name in enumerate([
            gettext('Result level'),
            gettext('Indicators'),
            gettext('Means of verification'),
            gettext('Assumptions')
        ]):
        add_header_cell(ws, 3, col+1, name)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col + 1)].width = 50
    levels = program['levels']
    if request.GET.get('groupby') == "2":
        sorted_levels = sorted(levels, key=itemgetter('level_depth', 'ontology'))
    else:
        levels_by_pk = {l['pk']: l for l in levels}
        sorted_levels = []
        for level in [l for l in sorted(levels, key=itemgetter('ontology')) if l['level_depth'] == 1]:
            sorted_levels += get_child_levels(level, levels_by_pk)
        levels = sorted_levels
    row = 4
    for level in sorted_levels:
        merge_start = row
        cell = ws.cell(row=row, column=1)
        cell.value = level['display_name']
        cell.alignment = TOP_LEFT_ALIGN_WRAP
        cell.fill = LEVEL_ROW_FILL
        cell = ws.cell(row=row, column=4)
        cell.value = level['assumptions']
        cell.alignment = TOP_LEFT_ALIGN_WRAP
        for indicator in sorted(level['indicators'], key=itemgetter('level_order')):
            cell = ws.cell(row=row, column=2)
            value = gettext('Indicator')
            if program['manual_numbering']:
                value += ' {}'.format(indicator['number']) if indicator['number'] else ''
            elif indicator['level_order_display'] or level['display_ontology']:
                value += ' {}{}'.format(level['display_ontology'], indicator['level_order_display'])
            value += ': {}'.format(indicator['name'])
            cell.value = value
            cell.alignment = TOP_LEFT_ALIGN_WRAP
            cell = ws.cell(row=row, column=3)
            cell.value = indicator['means_of_verification']
            cell.alignment = TOP_LEFT_ALIGN_WRAP
            row += 1
        if merge_start == row:
            row += 1
        else:
            ws.merge_cells(
                start_row=merge_start, end_row=row-1,
                start_column=1, end_column=1
            )
            ws.merge_cells(
                start_row=merge_start, end_row=row-1,
                start_column=4, end_column=4
            )
        cell = ws.cell(row=merge_start, column=1)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=2)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=3)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=4)
        cell.border = BORDER_TOP
    if program['unassigned_indicators']:
        merge_start = row
        cell = ws.cell(row=row, column=1)
        cell.value = gettext('Indicators unassigned to a results framework level')
        cell.alignment = TOP_LEFT_ALIGN_WRAP
        cell.fill = LEVEL_ROW_FILL
        for indicator in program['unassigned_indicators']:
            cell = ws.cell(row=row, column=2)
            value = gettext('Indicator')
            if program['manual_numbering']:
                value += ' {}'.format(indicator['number']) if indicator['number'] else ''
            cell.value = '{}: {}'.format(value, indicator['name'])
            cell.alignment = TOP_LEFT_ALIGN_WRAP
            cell = ws.cell(row=row, column=3)
            cell.value = indicator['means_of_verification']
            cell.alignment = TOP_LEFT_ALIGN_WRAP
            row += 1
        if merge_start == row:
            row += 1
        else:
            ws.merge_cells(
                start_row=merge_start, end_row=row-1,
                start_column=1, end_column=1
            )
            ws.merge_cells(
                start_row=merge_start, end_row=row-1,
                start_column=4, end_column=4
            )
        cell = ws.cell(row=merge_start, column=1)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=2)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=3)
        cell.border = BORDER_TOP
        cell = ws.cell(row=merge_start, column=4)
        cell.border = BORDER_TOP
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(
        '{} - {}.xlsx'.format(program['name'], gettext('Logframe'))
    )
    wb.save(response)
    return response


@login_required
def programs_rollup_export(request):
    program_pks = [p.pk for p in request.user.tola_user.available_programs]
    annotated_programs = ProgramWithMetrics.home_page.filter(pk__in=program_pks).with_annotations()
    data = {
        p.gaitid if p.gaitid else "no gait id {}".format(count): {
            'gaitid': p.gaitid,
            'name': p.name,
            'tola_creation_date': p.create_date.date().isoformat(),
            'is_active': p.funding_status.lower().strip() == 'funded',
            'indicator_count': p.metrics['indicator_count'],
            'indicators_with_targets': p.metrics['targets_defined'],
            'indicators_with_results': p.metrics['reported_results'],
            'results_count': p.metrics['results_count'],
            'results_with_evidence': p.metrics['results_evidence'],
            'indicators_reporting_scope': p.scope_counts['reporting_count'],
            'indicators_reporting_on_target': p.scope_counts['on_scope'],
            'indicators_reporting_below_target': p.scope_counts['low'],
            'indicators_reporting_above_target': p.scope_counts['high'],
            } for count, p in enumerate(annotated_programs)
        }
    return JsonResponse(data)


@login_required
def programs_rollup_export_csv(request):
    # TODO: after LevelUp please remove unicode calls:
    CSV_HEADERS = [
        'program_name', 'gait_id', 'countries', 'sectors', 'status', 'funding_status', 'start_date', 'end_date',
        'tola_creation_date', 'most_recent_change_log_entry', 'program_period', 'indicator_count', 'indicators_reporting_above_target',
        'indicators_reporting_on_target', 'indicators_reporting_below_target', 'indicators_with_targets',
        'indicators_with_results', 'results_count', 'results_with_evidence'
    ]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="program_rollup_{}.csv"'.format(
        datetime.date.today().isoformat()
    )
    writer = csv.writer(response)
    writer.writerow(CSV_HEADERS)
    program_pks = [p.pk for p in request.user.tola_user.available_programs]
    annotated_programs = ProgramWithMetrics.home_page.filter(pk__in=program_pks).with_annotations()
    recent_change_log = {k:v.date() for (k, v) in ProgramAuditLog.objects.filter(program__in=annotated_programs) \
        .values_list('program').annotate(latest_date=Max('date'))}
    for program in sorted([p for p in annotated_programs], key=lambda p: p.name):
        row = [
            program.name,
            program.gaitid if program.gaitid else "no gait_id, program id {}".format(program.id),
            " / ".join([c.country for c in program.country.all()]),
            " / ".join(set([i.sector.sector for i in program.indicator_set.all() if i.sector and i.sector.sector])),
            "active" if program.funding_status.lower().strip() == 'funded' else "inactive",
            program.funding_status,
            program.reporting_period_start.isoformat() if program.reporting_period_start else '',
            program.reporting_period_end.isoformat() if program.reporting_period_end else '',
            program.create_date.date().isoformat() if program.create_date else '',
            recent_change_log.get(program.id, "None"),
            '{}%'.format(program.percent_complete) if program.percent_complete >= 0 else '',
            program.metrics['indicator_count'],
            program.scope_counts['high'],
            program.scope_counts['on_scope'],
            program.scope_counts['low'],
            program.metrics['targets_defined'],
            program.metrics['reported_results'],
            program.metrics['results_count'],
            program.metrics['results_evidence']
        ]
        row = [str(s) for s in row]
        writer.writerow(row)
    return response


@login_required
def indicator_detail_export_csv(request):
    CSV_HEADERS = [
        'indicator_id', 'indicator_name', 'sector', 'country_strategic_objective', 'level', 'target_frequency',
        'baseline_value', 'number_or_percent', 'direction_of_change', 'lop_target_value',
        'program_name', 'gait_id', 'countries', 'regions','program_status', 'start_date', 'end_date',
    ]
    output_file_template = 'attachment; filename="indicator_detail_{}.csv"'
    if 'funded' in request.path:
        output_file_template = 'attachment; filename="indicator_detail_active_{}.csv"'
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = output_file_template.format(
        datetime.date.today().isoformat()
    )
    writer = csv.writer(response)
    writer.writerow(CSV_HEADERS)

    target_frequency_map = {id: label for id, label in Indicator.TARGET_FREQUENCIES}
    uom_type_map = {id: label for id, label in Indicator.UNIT_OF_MEASURE_TYPES}
    direction_of_change_map = {id: label for id, label in Indicator.DIRECTION_OF_CHANGE}

    programs = request.user.tola_user.available_programs.prefetch_related('country', 'country__region')
    if 'funded' in request.path:
        programs = programs.exclude(~Q(funding_status='funded'))
    program_map = {p.id: p for p in programs}
    program_tiers = {}
    for program in programs:
        program_tiers[program.pk] = list(program.level_tiers.order_by('tier_depth').values_list('name', flat=True))
    indicator_data = Indicator.program_page_objects.\
        filter(program_id__in=program_map.keys())\
        .select_related('sector', 'level', 'program')\
        .prefetch_related('program__level_tiers', 'strategic_objectives')
    country_data = Country.objects.select_related('region')
    country_map = {c.country: c.region.name for c in country_data}
    objective_map = {i.id: [so.name for so in i.strategic_objectives.all()] for i in indicator_data}

    # Build the level names in memory rather than doing db queries, since there are so many to do
    level_map = {level.pk: level for level in Level.objects.filter(program_id__in=program_map.keys())}
    level_name_map = {}
    for level in level_map.values():
        def get_parent_segment(ontology_segments, level):
            if level.parent_id:
                parent = level_map[level.parent_id]
                parent_segment = parent.customsort
                return get_parent_segment([parent_segment] + ontology_segments, parent)
            else:
                return ontology_segments
        ontology_segments = get_parent_segment([level.customsort], level)[1:]  # Get rid of the top level segment

        tier_name = program_tiers[level.program.pk][len(ontology_segments)]
        if len(ontology_segments) > 0:
            level_name_map[level.pk] = f'{tier_name} {".".join([str(o) for o in ontology_segments])}: {level.name}'
        else:
            level_name_map[level.pk] = f'{tier_name}: {level.name}'

    for indicator in sorted([i for i in indicator_data], key=lambda i: i.program_id):
        program = program_map[indicator.program_id]
        regions = [country_map[c.strip()] for c in program.countries.split(',')]
        row = [
            indicator.id,
            indicator.name,
            indicator.sector.sector if indicator.sector else 'None',
            '/'.join(objective_map[indicator.id]) if len(objective_map[indicator.id]) > 0 else 'None',
            level_name_map[indicator.level.pk] if indicator.level else 'None',
            target_frequency_map[indicator.target_frequency] if indicator.target_frequency else 'None',
            indicator.baseline,
            uom_type_map[indicator.unit_of_measure_type] if indicator.unit_of_measure_type else 'None',
            direction_of_change_map[indicator.direction_of_change] if indicator.direction_of_change else 'None',
            indicator.lop_target_calculated if indicator.lop_target_calculated else indicator.lop_target,
            program.name,
            program.gaitid if program.gaitid else "no gait_id, program id {}".format(program.id),
            '/'.join([c.strip() for c in program.countries.split(',')]),
            '/'.join(set(regions)),
            'Active' if program.funding_status == 'Funded' else 'Inactive',
            program.reporting_period_start,
            program.reporting_period_end
        ]
        row = [str(s) for s in row]
        writer.writerow(row)
    return response


# API views:

@login_required
@has_program_read_access
def program_page(request, program):
    """Program Page Template view - returns the program page template with JSON populated for React logic"""
    # redirect to home if program isn't active or doesn't exist:
    try:
        program = Program.rf_aware_objects.only(
            'reporting_period_start', 'reporting_period_end'
        ).get(pk=int(program))
    except (Program.DoesNotExist, ValueError):
        return redirect('/')
    # redirect to setup page if reporting period isn't complete:
    if any([program.reporting_period_start is None, program.reporting_period_end is None]):
        return render(
            request,
            'indicators/program_setup_incomplete.html',
            {'program': program, 'redirect_url': request.path}
        )
    context = {
        'program': ProgramPageProgramSerializer.load_for_pk(program.pk).data,
        'pinned_reports': list(PinnedReport.objects.filter(
            program_id=program.pk, tola_user=request.user.tola_user
            )) + [PinnedReport.default_report(program.pk)],
        'delete_pinned_report_url': reverse('delete_pinned_report'),
        'indicator_on_scope_margin': Indicator.ONSCOPE_MARGIN,
        'readonly': not request.has_write_access,
        'result_readonly': not request.has_medium_access
    }
    return render(request, 'indicators/program_page.html', context)


@login_required
def old_program_page(request, program_id, indicator_id, indicator_type_id):
    """Redirects old /program/<program_id>/<indicator_id>/<indicator_type_id>/ urls to new program page url"""
    program = get_object_or_404(Program, pk=program_id)
    if indicator_id != 0 or indicator_type_id != 0:
        logger.warning('attempt to access program page with filters indicator id {0} and indicator type id {1}'.format(
            indicator_id, indicator_type_id))
    return redirect(program.program_page_url, permanent=True)


@login_required
@has_program_read_access
def api_program_ordering(request, program):
    """Returns program-wide RF-aware ordering (used after indicator deletion on program page)"""
    try:
        data = ProgramPageIndicatorUpdateSerializer.load_for_pk(program).data
    except Program.DoesNotExist:
        logger.warning('attempt to access program page ordering for bad pk {}'.format(program))
        return JsonResponse({'success': False, 'msg': 'bad Program PK'})
    return JsonResponse(data)



@login_required
@indicator_pk_adapter(has_indicator_read_access)
def api_program_page_indicator(request, pk, program):
    """Returns single indicator updated JSON and ordering information for program page)"""
    try:
        data = ProgramPageIndicatorUpdateSerializer.load_for_indicator_and_program(pk, program).data
    except (Program.DoesNotExist, Indicator.DoesNotExist):
        logger.warning('attempt to access indicator update for bad pk {}'.format(pk))
        return JsonResponse({'success': False, 'msg': 'bad Indicator PK'})
    return JsonResponse(data)


@login_required
@has_program_read_access
def api_program_page(request, program):
    """Returns program JSON to hydrate Program Page react models"""
    try:
        data = ProgramPageProgramSerializer.load_for_pk(program).data
    except Program.DoesNotExist:
        logger.warning('attempt to access program page ordering for bad pk {}'.format(program))
        return JsonResponse({'success': False, 'msg': 'bad Program PK'})
    return JsonResponse(data)


@login_required
@has_program_read_access
def results_framework_export(request, program):
    """Returns .XLSX containing program's results framework"""
    program = Program.rf_aware_objects.get(pk=program)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(gettext("Results Framework"))
    get_font = lambda attrs: styles.Font(**{**{'name': 'Calibri', 'size': 12}, **attrs})
    ws.cell(row=2, column=2).value = gettext("Results Framework")
    ws.cell(row=2, column=2).font = get_font({'size': 18, 'bold': True})
    ws.cell(row=3, column=2).value = program.name
    ws.cell(row=3, column=2).font = get_font({'size': 18})
    level_span_style = styles.NamedStyle(name='level_span')
    level_span_style.font = get_font({})
    level_span_style.alignment = styles.Alignment(wrap_text=True, vertical='center', horizontal='center')
    level_span_style.fill = styles.PatternFill('solid', 'E5E5E5')
    wb.add_named_style(level_span_style)
    level_single_style = styles.NamedStyle(name='level_no_span')
    level_single_style.font = get_font({})
    level_single_style.alignment = styles.Alignment(wrap_text=True, vertical='top', horizontal='left')
    level_single_style.fill = styles.PatternFill('solid', 'E5E5E5')
    wb.add_named_style(level_single_style)
    bottom_tier = program.level_tiers.count()
    def row_height_getter(cell):
        lines_of_text = str(cell.value).splitlines()
        row = cell.row
        def get_row_height_decorated(w):
            lines = sum([math.ceil(len(s)/w) or 1 for s in lines_of_text])
            height = 26 + lines * 15
            if lines == 1:
                height = 30
            return max(height, ws.row_dimensions[row].height or 0, 30)
        return get_row_height_decorated
    def write_level(parent, start_row, start_column):
        levels = program.levels.filter(parent=parent).order_by('customsort')
        column = start_column
        row = start_row
        if not levels:
            return column + 2
        for level in levels:
            current_column = column
            cell = ws.cell(row=row, column=column)
            cell.value = level.display_name
            get_row_height = row_height_getter(cell)
            if level.level_depth == bottom_tier:
                cell.style = 'level_no_span'
                row = row + 2
                ws.row_dimensions[cell.row].height = get_row_height(24)
            else:
                column = write_level(level, row+2, column)
                if column - 2 <= current_column:
                    cell.style = 'level_no_span'
                    ws.row_dimensions[cell.row].height = get_row_height(24)
                else:
                    cell.style = 'level_span'
                    ws.merge_cells(start_row=row, end_row=row, start_column=current_column, end_column=column-2)
                    width = 24 + 29 * ((column - 2 - current_column) / 2)
                    ws.row_dimensions[cell.row].height = get_row_height(width)
        if parent and parent.level_depth == bottom_tier-1:
            column = column + 2
        if parent is None:
            for column in range(column):
                width = 24.5 if (column + 1) % 2 == 0 else 3
                ws.column_dimensions[utils.get_column_letter(column + 1)].width = width
            for r in range(3, ws.max_row+2):
                if r % 2 == 0:
                    ws.row_dimensions[r].height = 10
        return column
    write_level(None, 5, 2)
    filename = "Results Framework.xlsx"
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    wb.save(response)
    return response
