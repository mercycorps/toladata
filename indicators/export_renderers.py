import openpyxl
from django.utils.translation import ugettext
from django.http import HttpResponse

from tola_management.programadmin import get_audit_log_workbook
from workflow.models import Program


EM_DASH = '–'


CENTER_ALIGN = openpyxl.styles.Alignment(horizontal='center', vertical='bottom')
RIGHT_ALIGN = openpyxl.styles.Alignment(horizontal='right', vertical='bottom')
LEFT_ALIGN_WRAP = openpyxl.styles.Alignment(wrap_text=True)

class ExcelRendererBase:
    """Set of utility functions for rendering a serialized IPTT into an Excel export"""
    TITLE_FONT = openpyxl.styles.Font(size=18)
    HEADER_FONT = openpyxl.styles.Font(bold=True)
    HEADER_FILL = openpyxl.styles.PatternFill('solid', 'EEEEEE')
    LEVEL_ROW_FILL = openpyxl.styles.PatternFill('solid', 'CCCCCC')

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def initialize_workbook(self):
        title_style = openpyxl.styles.NamedStyle(name='title')
        title_style.font = openpyxl.styles.Font(size=18)
        self.wb.add_named_style(title_style)
        header_style = openpyxl.styles.NamedStyle(name='header')
        header_style.font = openpyxl.styles.Font(bold=True)
        header_style.alignment = CENTER_ALIGN
        header_style.fill = openpyxl.styles.PatternFill('solid', 'EEEEEE')
        self.wb.add_named_style(header_style)
        sub_header_style = openpyxl.styles.NamedStyle(name='sub_header')
        sub_header_style.font = openpyxl.styles.Font(bold=True)
        sub_header_style.alignment = CENTER_ALIGN
        self.wb.add_named_style(sub_header_style)
        right_align_header_style = openpyxl.styles.NamedStyle(name='header_right')
        right_align_header_style.font = openpyxl.styles.Font(bold=True)
        right_align_header_style.alignment = RIGHT_ALIGN
        right_align_header_style.fill = openpyxl.styles.PatternFill('solid', 'EEEEEE')
        self.wb.add_named_style(right_align_header_style)
        level_row_style = openpyxl.styles.NamedStyle(name='level_row')
        level_row_style.font = openpyxl.styles.Font(bold=True)
        level_row_style.fill = openpyxl.styles.PatternFill('solid', 'CCCCCC')
        self.wb.add_named_style(level_row_style)

    @property
    def header_columns(self):
        """accurate list of headers for non-report (non period) data (name and description)"""
        headers = [
            ugettext('Program ID'),
            ugettext('Indicator ID'),
            # Translators: "No." as in abbreviation for Number
            ugettext('No.'),
            ugettext('Indicator')
        ]
        if self.level_column:
            headers += [ugettext('Level')]
        if self.uom_column:
            headers += [ugettext('Unit of measure')]
        if self.change_column:
            headers += [
                # Translators: this is short for "Direction of Change" as in + or -
                ugettext('Change')
            ]
        if self.cnc_column:
            headers += [
                # Translators: 'C' as in Cumulative and 'NC' as in Non Cumulative
                ugettext('C / NC'),
            ]
        if self.uom_type_column:
            headers += ['# / %']
        if self.baseline_column:
            headers += [ugettext('Baseline')]
        return headers

    def _get_name(self):
        name = {
            1: ugettext('Life of Program (LoP) only'),
            2: ugettext('Midline and endline'),
            3: ugettext('Annual'),
            4: ugettext('Semi-annual'),
            5: ugettext('Tri-annual'),
            # Translators: this is the measure of time (3 months)
            6: ugettext('Quarterly'),
            7: ugettext('Monthly')
        }[self.frequency]
        if name and len(name) > 30:
            name = name[:26] + '...'
        return name

    def add_sheet(self):
        level_rows = self.serializer['level_rows'][self.frequency]
        try:
            level_row = next(level_rows)
        except StopIteration:
            return None
        sheet = self.wb.create_sheet(self._get_name())
        self.add_headers(sheet)
        current_row = 5
        while True:
            current_row = self.add_level_row(level_row, sheet, current_row)
            try:
                level_row = next(level_rows)
            except StopIteration:
                self.set_column_widths(sheet)
                self.add_explainer_row(sheet)
                return sheet

    def add_explainer_row(self, sheet):
        sheet.append([None, None,
                      # Translators: Explanation at the bottom of a report (as a footnote) about value rounding
                      ugettext("*All actual values in this report are rounded to two decimal places.")])

    def add_headers(self, sheet):
        current_row = 1
        for title in [self.serializer['report_title'], self.serializer['report_date_range'],
                      self.serializer['program_name']]:
            sheet.merge_cells(
                start_row=current_row, start_column=self.TITLE_START_COLUMN,
                end_row=current_row, end_column=len(self.header_columns)
                )
            cell = sheet.cell(row=current_row, column=self.TITLE_START_COLUMN)
            cell.value = str(title)
            cell.style = 'title'
            current_row += 1
        current_column = 1
        for header in self.header_columns:
            cell = sheet.cell(row=current_row, column=current_column)
            cell.value = str(header)
            cell.style = 'header'
            current_column += 1
        current_column = self.add_period_header(sheet, current_column, self.serializer['lop_period'], lop=True)
        for period in self.serializer['periods'][self.frequency]:
            current_column = self.add_period_header(
                sheet, current_column, period
            )

    @staticmethod
    def add_period_header(sheet, col, period, lop=False):
        for header, row in [
                (period.header, 2),
                (period.subheader, 3)
            ]:
            if header:
                if period.tva:
                    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
                cell = sheet.cell(row=row, column=col)
                cell.value = str(header)
                cell.style = 'sub_header'
        actual_header = ugettext('Actual')
        if lop:
            actual_header += " *"
        columns = [
            ugettext('Target'), actual_header, str(ugettext('% Met')).title()
        ] if period.tva else [actual_header,]
        for col_no, col_header in enumerate(columns):
            cell = sheet.cell(row=4, column=col+col_no)
            cell.value = str(col_header)
            cell.style = 'header_right'
        return col + len(columns)

    def add_level_row(self, level_row, sheet, current_row):
        if level_row['level']:
            sheet.cell(row=current_row, column=1).style = 'level_row'
            sheet.cell(row=current_row, column=2).style = 'level_row'
            sheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=self.column_count)
            cell = sheet.cell(row=current_row, column=3)
            cell.value = str(level_row['level']['full_name'])
            cell.style = 'level_row'
            current_row += 1
        for indicator in level_row['indicators']:
            current_row = self.add_indicator_data(indicator, sheet, current_row)
        return current_row

    @staticmethod
    def int_cell(value):
        if not value and value != 0:
            return None, 'General'
        return int(value), '0'

    @staticmethod
    def float_cell(value):
        if not value and value != 0:
            return None, 'General'
        value = round(float(value), 2)
        if value == int(value):
            return int(value), '0'
        if value == round(value, 1):
            return round(value, 1), '0.0'
        return value, '0.00'

    @staticmethod
    def percent_cell(value):
        if not value and value != 0:
            return None, 'General'
        value = round(float(value), 4)
        if value == round(value, 2):
            return round(value, 2), '0%'
        if value == round(value, 3):
            return round(value, 3), '0.0%'
        return value, '0.00%'

    @staticmethod
    def percent_value_cell(value):
        if not value and value != 0:
            return None, 'General'
        value = round(float(value), 2)
        if value == int(value):
            return value/100, '0%'
        return value/100, '0.00%'

    @staticmethod
    def str_cell(value):
        if not value:
            return None, 'General'
        value = str(value)
        return value, 'General'

    def get_period_report_data_columns(self, indicator_periods, values_func):
        for period_header, period_data in zip(self.serializer['periods'][self.frequency], indicator_periods):
            assert period_header.count == period_data['count']
            if period_header.tva:
                yield (period_data['target'], values_func, None, None)
            yield (period_data['actual'], values_func, None, None)
            if period_header.tva:
                yield (period_data['met'], self.percent_cell, None, None)

    @staticmethod
    def write_indicator_row(sheet, current_row, indicator_columns):
        indicator_pk = indicator_columns[1][0]
        for column, (value, format_func, alignment, style) in enumerate(indicator_columns):
            number_format = None
            cell = sheet.cell(row=current_row, column=column+1)
            empty_cell = EM_DASH
            if style == 'empty_blank':
                empty_cell = ''
                style = None
            elif style == 'empty_na':
                empty_cell = ugettext('N/A')
                style = None
            try:
                if value is None:
                    value, number_format = None, None
                else:
                    value, number_format = format_func(value)
            except (AttributeError, TypeError, ValueError) as e:
                cell.value = None
                cell.comment = openpyxl.comments.Comment(
                    'error {} with attribute {} on indicator pk {}'.format(
                        e, value, indicator_pk
                        ), 'Tola System')
            else:
                if value is None:
                    value = empty_cell
                    alignment = RIGHT_ALIGN
                cell.value = value
                if alignment is not None:
                    cell.alignment = alignment
                if style is not None:
                    cell.style = style
                if number_format is not None:
                    cell.number_format = number_format

    def get_label_values(self, label_pk, report_data):
        label_values = []
        if self.baseline_column:
            label_values.append(None)
        label_values += [
            None, report_data['lop_period'].get('disaggregations', {}).get(label_pk, {}).get('actual'), None]
        for period_header, period_data in zip(self.serializer['periods'][self.frequency], report_data['periods']):
            if period_header.tva:
                label_values.append(None)
            label_values.append(period_data.get('disaggregations', {}).get(label_pk, {}).get('actual'))
            if period_header.tva:
                label_values.append(None)
        return label_values

    def add_indicator_data(self, indicator, sheet, current_row):
        if indicator['unit_of_measure_type'] == '%':
            values_func = self.percent_value_cell
        else:
            values_func = self.float_cell
        indicator_columns = [
            (indicator['program_pk'], self.int_cell, RIGHT_ALIGN, None),
            (indicator['pk'], self.int_cell, RIGHT_ALIGN, None),
            (indicator['number'], self.str_cell, LEFT_ALIGN_WRAP, None),
            (indicator['name'], self.str_cell, None, self.INDICATOR_NAME_CELL)
        ]
        if self.level_column:
            indicator_columns.append(
                (indicator['old_level_name'], self.str_cell, None, None)
            )
        if self.uom_column:
            indicator_columns += [
                (indicator['unit_of_measure'], self.str_cell, None, None),
            ]
        if self.change_column:
            indicator_columns += [
                (indicator['direction_of_change'], self.str_cell, CENTER_ALIGN, 'empty_na'),
            ]
        if self.cnc_column:
            indicator_columns += [
                (self.CUMULATIVE if indicator['is_cumulative'] else self.NON_CUMULATIVE,
                 self.str_cell, None, None),
            ]
        if self.uom_type_column:
            indicator_columns += [
                (indicator['unit_of_measure_type'], self.str_cell, CENTER_ALIGN, 'empty_blank'),
            ]
        if self.baseline_column:
            indicator_columns += [
                (indicator['baseline'], values_func, None, 'empty_na'),
                ]
        indicator_columns += [
            (indicator['report_data']['lop_period']['target'], values_func, None, None),
            (indicator['report_data']['lop_period']['actual'], values_func, None, None),
            (indicator['report_data']['lop_period']['met'], self.percent_cell, None, None),
        ]
        for period_column in self.get_period_report_data_columns(indicator['report_data']['periods'], values_func):
            indicator_columns.append(period_column)
        self.write_indicator_row(sheet, current_row, indicator_columns)
        current_row += 1
        label_merge_column = len(self.header_columns) - (1 if self.baseline_column else 0)
        for disaggregation in sorted(indicator.get('disaggregations', []),
                                     key=lambda disagg: ugettext(disagg['name'])):
            labels = disaggregation['labels']
            if self.hide_empty_disagg_categories:
                # only include row if we are _not_ hiding empty categories or this category isn't empty:
                labels = list(filter(
                    lambda label: indicator['report_data']['lop_period'].get('disaggregations', {}).get(
                        label['pk'], {}).get('actual', None) is not None,
                    labels))
            if labels:
                top_row = current_row
                for label in labels:
                    sheet.merge_cells(
                        start_row=current_row, start_column=4,
                        end_row=current_row, end_column=label_merge_column
                    )
                    cell = sheet.cell(row=current_row, column=4)
                    cell.value = label['name']
                    cell.style = self.DISAGGREGATION_CATEGORY_CELL
                    for column_count, value in enumerate(self.get_label_values(label['pk'], indicator['report_data'])):
                        cell = sheet.cell(row=current_row, column=label_merge_column+column_count+1)
                        value, number_format = values_func(value)
                        if value is None:
                            cell.value = EM_DASH
                            cell.alignment = RIGHT_ALIGN
                        else:
                            cell.value = value
                        if number_format is not None:
                            cell.number_format = number_format
                    current_row += 1
                sheet.merge_cells(start_row=top_row, start_column=3, end_row=current_row-1, end_column=3)
                cell = sheet.cell(row=top_row, column=3)
                cell.value = ugettext(disaggregation['name'])
                cell.style = self.DISAGGREGATION_CELL
                for row in range(top_row, current_row):
                    sheet.row_dimensions[row].hidden = True
        return current_row

    def set_column_widths(self, sheet):
        widths = [10, 10, 17, 100]
        if self.level_column:
            widths.append(12)
        if self.uom_column:
            widths.append(30)
        if self.change_column:
            widths.append(12)
        if self.cnc_column:
            widths.append(15)
        if self.uom_type_column:
            widths.append(8)
        if self.baseline_column:
            widths.append(20)
        widths += [12]*3
        for period in self.serializer['periods'][self.frequency]:
            widths += [12]*3 if period.tva else [12,]
        for col_no, width in enumerate(widths):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_no + 1)].width = width
        sheet.column_dimensions['A'].hidden = True
        sheet.column_dimensions['B'].hidden = True


    def render_to_response(self):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(self.filename)
        self.wb.save(response)
        return response

    @property
    def column_count(self):
        return len(self.header_columns) + 3 + sum(
            [3 if period.tva else 1 for period in self.serializer['periods'][self.frequency]]
            )

    def add_change_log(self, program_pk):
        sheet = self.wb.create_sheet(ugettext('Change log'))
        program = Program.rf_aware_objects.get(pk=program_pk)
        get_audit_log_workbook(sheet, program)


class IPTTExcelRenderer(ExcelRendererBase):
    TITLE_START_COLUMN = 3 # 2 rows currently hidden at start, title starts at column C
    INDICATOR_NAME_CELL = openpyxl.styles.NamedStyle(
        name="indicator_name",
        font=openpyxl.styles.Font(underline='single'),
        alignment=openpyxl.styles.Alignment(wrap_text=True)
    )
    DISAGGREGATION_CELL = openpyxl.styles.NamedStyle(
        name='disaggregation_cell',
        font=openpyxl.styles.Font(bold=True),
        alignment=openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)
    )
    DISAGGREGATION_CATEGORY_CELL = openpyxl.styles.NamedStyle(
        name='category_cell',
        alignment=openpyxl.styles.Alignment(horizontal='right', wrap_text=True)
    )

    def __init__(self, serializer, **kwargs):
        super().__init__()
         # Translators: referring to an indicator whose results accumulate over time
        self.CUMULATIVE = ugettext("Cumulative")
        # Translators: referring to an indicator whose results do not accumulate over time
        self.NON_CUMULATIVE = ugettext("Non-cumulative")
        params = kwargs.get('params', {})
        self.filename = serializer.filename
        self.serializer = serializer.data
        self.initialize_workbook()
        self.columns = params.get('columns', [])
        self.hide_empty_disagg_categories = params.get('hide_empty_disagg_categories', False)
        has_sheets = False
        for frequency in self.serializer['frequencies']:
            self.frequency = frequency
            if self.add_sheet() is not None:
                has_sheets = True
        if not has_sheets:
            self.add_blank_sheet()

    def add_blank_sheet(self):
        # Translators: This is the title of a spreadsheet tab that shows do data has been found for the user's query.
        sheet = self.wb.create_sheet(ugettext('No data'))
        sheet['A1'].value = "No data matched the criteria"

    @property
    def level_column(self):
        return not self.serializer['results_framework']

    @property
    def uom_column(self):
        return 0 not in self.columns

    @property
    def change_column(self):
        return 1 not in self.columns

    @property
    def cnc_column(self):
        return 2 not in self.columns

    @property
    def uom_type_column(self):
        return 3 not in self.columns

    @property
    def baseline_column(self):
        return 4 not in self.columns

    @property
    def optional_columns_length(self):
        return sum([
            (1 if self.uom_column else 0),
            (1 if self.change_column else 0),
            (1 if self.cnc_column else 0),
            (1 if self.uom_type_column else 0),
            (1 if self.baseline_column else 0),
        ])
