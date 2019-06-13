"""
Data definitions and code related to Indicator Plan table view and XLS export
"""
from itertools import groupby

from django.utils.translation import (
    ugettext,
    ugettext_lazy as _
)

# Column groupings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

from indicators import models

PERFORMANCE_INDICATOR = _('Performance Indicator')
TARGETS = _('Targets')
DATA_ACQUISITION = _('Data Acquisition')
ANALYSES_AND_REPORTING = _('Analyses and Reporting')

# XLS cell widths
LARGE_CELL = 40
MEDIUM_CELL = 20

# fields are either an Indicator attribute name (as str), or a callable
COLUMNS = [
    {
        'name': _('Level'),
        'category': PERFORMANCE_INDICATOR,
        'field': lambda i: i.leveltier_name if i.results_framework else i.old_level,
        'cell_width': 10,
    },
    {
        'name': _('No.'),
        'category': PERFORMANCE_INDICATOR,
        'field': lambda i: i.results_aware_number,
    },
    {
        'name': _('Performance indicator'),
        'category': PERFORMANCE_INDICATOR,
        'field': 'name',
        'cell_width': LARGE_CELL,
    },
    {
        'name': _('Indicator source'),
        'category': PERFORMANCE_INDICATOR,
        'field': 'source',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Indicator definition'),
        'category': PERFORMANCE_INDICATOR,
        'field': 'definition',
        'cell_width': LARGE_CELL,
    },
    {
        'name': _('Disaggregation'),
        'category': PERFORMANCE_INDICATOR,
        'field': 'disaggregations',
        'cell_width': MEDIUM_CELL,
    },

    {
        'name': _('Unit of measure'),
        'category': TARGETS,
        'field': 'unit_of_measure',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Direction of change'),
        'category': TARGETS,
        'field': lambda i: i.get_direction_of_change_display(),
    },
    {
        'name': _('# / %'),
        'category': TARGETS,
        'field': lambda i: i.get_unit_of_measure_type_display(),
    },
    {
        'name': _('Calculation'),
        'category': TARGETS,
        'field': lambda i: 'C' if i.is_cumulative else 'NC',
    },
    {
        'name': _('Baseline value'),
        'category': TARGETS,
        'field': 'baseline',
    },
    {
        'name': _('LOP target'),
        'category': TARGETS,
        'field': 'lop_target',
    },
    {
        'name': _('Rationale for target'),
        'category': TARGETS,
        'field': 'rationale_for_target',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Target frequency'),
        'category': TARGETS,
        'field': lambda i: i.get_target_frequency_display(),
        'cell_width': MEDIUM_CELL,
    },

    {
        'name': _('Means of verification'),
        'category': DATA_ACQUISITION,
        'field': 'means_of_verification',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Data collection method'),
        'category': DATA_ACQUISITION,
        'field': 'data_collection_method',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Frequency of data collection'),
        'category': DATA_ACQUISITION,
        'field': lambda i: i.data_collection_frequency.frequency if i.data_collection_frequency else None,
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Data points'),
        'category': DATA_ACQUISITION,
        'field': 'data_points',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Responsible person(s) & team'),
        'category': DATA_ACQUISITION,
        'field': 'responsible_person',
        'cell_width': MEDIUM_CELL,
    },

    {
        'name': _('Method of analysis'),
        'category': ANALYSES_AND_REPORTING,
        'field': 'method_of_analysis',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Information use'),
        'category': ANALYSES_AND_REPORTING,
        'field': 'information_use',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Frequency of reporting'),
        'category': ANALYSES_AND_REPORTING,
        'field': lambda i: i.reporting_frequency.frequency if i.reporting_frequency else None,
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Quality assurance measures'),
        'category': ANALYSES_AND_REPORTING,
        'field': 'quality_assurance',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Data issues'),
        'category': ANALYSES_AND_REPORTING,
        'field': 'data_issues',
        'cell_width': MEDIUM_CELL,
    },
    {
        'name': _('Comments'),
        'category': ANALYSES_AND_REPORTING,
        'field': 'comments',
        'cell_width': MEDIUM_CELL,
    },
]


def row(indicator):
    """
    Return a row of the plan for an indicator with display values
    """
    r = []
    for col in COLUMNS:
        field = col['field']
        r.append(field(indicator) if callable(field) else getattr(indicator, field))
    return r


def column_names():
    """
    A list of column names
    """
    return [c['name'] for c in COLUMNS]


def non_rf_indicator_queryset(program_id):
    """
    A QS of indicators to create the indicator plan from
    """
    return models.Indicator.objects.filter(program_id=program_id).select_related().with_logframe_sorting()

def tier_sorted_indicator_queryset(program_id):
    """RF program with the "sort by level" option picked queryset"""
    return sorted(
        models.Level.objects.filter(program_id=program_id),
        key=lambda l: l.get_level_depth()
        )

def chain_sorted_indicator_queryset(program_id):
    """RF program with the "sort by outcome chain" option (default) picked queryset"""
    levels = []
    top_tier = models.Level.objects.filter(program_id=program_id, parent__isnull=True)
    for level in top_tier:
        levels += level.get_children()
    return levels

def get_rf_rows(level_qs, program_id):
    rows = []
    for level in level_qs:
        indicators = level.indicator_set.all()
        if indicators:
            rows.append(
                {
                    'row_type': 'level',
                    'row_data': level.display_name
                }
            )
            rows += [{'row_type': 'indicator', 'row_data': row(i)} for i in indicators]
    unassigned_indicators = models.Indicator.objects.filter(
        program_id=program_id, level_id__isnull=True
        ).with_logframe_sorting()
    if unassigned_indicators:
        rows.append(
            {
                'row_type': 'level',
                'row_data': _('Indicators unassigned to a results framework level')
            }
        )
        rows += [{'row_type': 'indicator', 'row_data': row(i)} for i in unassigned_indicators]
    return rows

def get_non_rf_rows(indicator_qs):
    return [{'row_type': 'indicator', 'row_data': row(i)} for i in indicator_qs]

def columns_by_category():
    """
    Returns an iterable of column objects by category name
    :return:
    """
    return groupby(COLUMNS, lambda c: c['category'])


# XLS generation

START_ROW = 2
START_COLUMN = 2

DARK_RED = '9e1b32'
TAN = 'dad6cb'
WHITE = 'ffffff'
BLACK = '000000'


def _apply_title_styling(cell):
    cell.font = Font(bold=True, color=WHITE, size=13)
    cell.fill = PatternFill(fill_type='solid', start_color=DARK_RED, end_color=DARK_RED)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def _apply_label_styling(cell):
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(fill_type='solid', start_color=TAN, end_color=TAN)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bd = Side(style='thin', color=BLACK)
    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return cell


def _apply_body_styling(cell):
    cell.font = Font(size=10)
    bd = Side(style='thin', color=BLACK)
    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def _set_column_widths(ws, col_num, width):
    ws.column_dimensions[get_column_letter(col_num)].width = width


def _set_row_height(ws, row_num, height):
    ws.row_dimensions[row_num].height = height


def _get_formatted_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = _('Indicator plan')

    row_num = START_ROW
    col_num = START_COLUMN

    # Set widths of columns
    _set_column_widths(ws, 1, 3)  # spacer col
    for col in COLUMNS:
        width = col.get('cell_width')
        if width:
            _set_column_widths(ws, col_num, width)
        col_num += 1

    # Title
    col_num = START_COLUMN
    cell = ws.cell(row_num, col_num, _('Indicator plan').encode('utf-8'))
    _apply_title_styling(cell)
    ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=col_num + len(COLUMNS)-1)
    row_num += 1

    # Category groupings
    col_num = START_COLUMN
    for category_name, columns in columns_by_category():
        num_columns = len(list(columns))
        cell = ws.cell(row_num, col_num, category_name.encode('utf-8'))
        ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=col_num + num_columns-1)
        _apply_label_styling(cell)
        col_num += num_columns
    row_num += 1

    # Column names
    col_num = START_COLUMN
    for column_name in column_names():
        cell = ws.cell(row_num, col_num, column_name.encode('utf-8'))
        _apply_label_styling(cell)
        col_num += 1
    row_num += 1
    return wb, row_num


def _style_level_row(ws, row_num):
    ws.merge_cells(start_row=row_num, start_column=START_COLUMN, end_row=row_num, end_column=len(COLUMNS))
    cell = ws.cell(row_num, START_COLUMN)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(fill_type='solid', start_color=TAN, end_color=TAN)
    bd = Side(style='thin', color=BLACK)
    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    ws.row_dimensions[row_num].height = 30


def create_rf_workbook(levels, program_id):
    """
    Take an iterable of levels and return a workbook
    """
    wb, row_num = _get_formatted_workbook()
    ws = wb.active
    col_num = START_COLUMN
    for level in levels:
        indicators = level.indicator_set.all()
        if indicators:
            cell = ws.cell(row_num, col_num)
            cell.value = level.display_name
            _style_level_row(ws, row_num)
            row_num += 1
            for indicator in indicators:
                for i, val in enumerate(row(indicator)):
                    cell = ws.cell(row_num, col_num, val)
                    if i == 0:
                        _apply_label_styling(cell)
                    else:
                        _apply_body_styling(cell)
                    col_num += 1
                col_num = START_COLUMN
                row_num += 1
    unassigned_indicators = models.Indicator.objects.filter(
        program_id=program_id, level_id__isnull=True
        ).with_logframe_sorting()
    if unassigned_indicators:
        cell = ws.cell(row_num, col_num)
        cell.value = ugettext('Indicators unassigned to a results framework level')
        _style_level_row(ws, row_num)
        row_num += 1
        for indicator in unassigned_indicators:
            for i, val in enumerate(row(indicator)):
                cell = ws.cell(row_num, col_num, val)
                if i == 0:
                    _apply_label_styling(cell)
                else:
                    _apply_body_styling(cell)
                col_num += 1
            col_num = START_COLUMN
            row_num += 1

    # fix issue with borders applied to multi-column cells
    update_borders(wb)

    return wb


def create_non_rf_workbook(indicators):
    """
    Take an iterable of indicators and return a workbook
    """
    wb, row_num = _get_formatted_workbook()
    ws = wb.active

    # rows
    col_num = START_COLUMN
    for level, level_indicators in groupby(indicators, lambda i: i.level):
        for indicator in level_indicators:
            for i, val in enumerate(row(indicator)):
                cell = ws.cell(row_num, col_num, val)
                if i == 0:
                    # first column has different styling
                    _apply_label_styling(cell)
                else:
                    _apply_body_styling(cell)
                col_num += 1
            col_num = START_COLUMN
            row_num += 1

        # spacer row
        for i in range(len(COLUMNS)):
            cell = ws.cell(row_num, col_num, '')
            _apply_label_styling(cell)
            _set_row_height(ws, row_num, 5)
            col_num += 1

        col_num = START_COLUMN
        row_num += 1

    # delete last space row
    ws.delete_rows(row_num-1)
    _set_row_height(ws, row_num-1, None)

    # fix issue with borders applied to multi-column cells
    update_borders(wb)

    return wb


# openpyxl fix
# below is a terrible massive hack to make borders work on merged cells
# see: https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working


def get_border(border_style=None, color=None):
    return Border(left=Side(border_style=border_style, color=color),
                  right=Side(border_style=border_style, color=color),
                  top=Side(border_style=border_style, color=color),
                  bottom=Side(border_style=border_style, color=color), )


def update_style(ws, cell_range):
    start_cell, end_cell = str(cell_range).split(':')
    start_coord = coordinate_from_string(start_cell)
    end_coord = coordinate_from_string(end_cell)

    start_row = start_coord[1]
    end_row = end_coord[1]

    start_col = column_index_from_string(start_coord[0])
    end_col = column_index_from_string(end_coord[0])

    border_style = ws.cell(row=start_row, column=start_col).border.left.style
    color = ws.cell(row=start_row, column=start_col).border.left.color

    cellborder = get_border(border_style, color)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = cellborder


def update_borders(wb):
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for each_range in ws.merged_cells.ranges:
            update_style(ws, each_range)
