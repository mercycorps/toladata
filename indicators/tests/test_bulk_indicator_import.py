import json
import string
import time
import os
import unittest
import copy
from tempfile import NamedTemporaryFile
import openpyxl
from openpyxl.cell import MergedCell
from django import test
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.serializers.json import DjangoJSONEncoder
from django.urls import reverse
from django.utils.translation import gettext, activate
from factories import (
    workflow_models as w_factories,
    indicators_models as i_factories
)
from tola_management.models import ProgramAuditLog
from tola.serializers import make_quantized_decimal
from indicators.models import Indicator, Level, LevelTier, BulkIndicatorImportFile
from indicators.views.bulk_indicator_import_views import (
    COLUMNS, COLUMNS_FIELD_INDEXES, FIRST_USED_COLUMN, DATA_START_ROW, PROGRAM_NAME_ROW, COLUMN_HEADER_ROW,
    HIDDEN_SHEET_NAME, BASE_TEMPLATE_NAME, LEVEL_HEADER_STYLE, EXISTING_INDICATOR_STYLE,
    UNPROTECTED_NEW_INDICATOR_STYLE, PROTECTED_NEW_INDICATOR_STYLE, RED_ERROR, ERROR_MSG_NAME_DUPLICATED,
    ERROR_MSG_NAME_IN_DB,
    ERROR_MISMATCHED_PROGRAM,
    ERROR_NO_NEW_INDICATORS,
    ERROR_UNDETERMINED_LEVEL,
    ERROR_TEMPLATE_NOT_FOUND,
    ERROR_MISMATCHED_TIERS,
    ERROR_INDICATOR_DATA_NOT_FOUND,
    ERROR_MISMATCHED_LEVEL_COUNT,
    ERROR_MISMATCHED_HEADERS,
    ERROR_SAVE_VALIDATION,
    ERROR_INVALID_LEVEL_HEADER,
    ERROR_MALFORMED_INDICATOR,
    ERROR_UNEXPECTED_INDICATOR_NUMBER,
    ERROR_INTERVENING_BLANK_ROW,
    ERROR_UNEXPECTED_LEVEL,
)

SAMPLE_INDICATOR_DATA = {
    'name': 'Lorem ipsum name', 'source': 'Lorem ipsum source', 'definition': 'Lorem ipsum definition',
    'justification': 'Lorem ipsum justification', 'unit_of_measure': 'Lorem ipsum unit_of_measure',
    'unit_of_measure_type': 1, 'rationale_for_target': 'Lorem ipsum rationale_for_target', 'baseline': '3.00',
    'direction_of_change': 1, 'target_frequency': 1, 'means_of_verification': 'Lorem ipsum means_of_verification',
    'data_collection_method': 'Lorem ipsum data_collection_method', 'data_points': 'Lorem ipsum data_points',
    'responsible_person': 'Lorem ipsum responsible_person', 'method_of_analysis': 'Lorem ipsum method_of_analysis',
    'information_use': 'Lorem ipsum information_use', 'quality_assurance': 'Lorem ipsum quality_assurance',
    'data_issues': 'Lorem ipsum data_issues', 'comments': 'Lorem ipsum comments', 'sector_id': 4, 'level_id': None,
    'program_id': None}


class TestBulkImportTemplateCreation(test.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.country = w_factories.CountryFactory()
        # cls.program = w_factories.RFProgramFactory(
        #     country=[cls.country], tiers=True, levels=6, indicators=150, indicators__levels=True)
        cls.program = w_factories.RFProgramFactory(
            country=[cls.country], tiers=True, levels=2, indicators=5, indicators__levels=True)
        levels = Level.objects.filter(program=cls.program)
        cls.ordered_levels = sorted(levels, key=lambda l: l.ontology)
        cls.tola_user = w_factories.TolaUserFactory()
        cls.client = test.client
        cls.first_used_column = FIRST_USED_COLUMN
        cls.data_start_row = DATA_START_ROW
        cls.program_name_row = PROGRAM_NAME_ROW
        cls.letters = list(string.ascii_lowercase) + [f'a{letter}' for letter in list(string.ascii_lowercase)]

    def get_template(self, program=None, request_params=None):
        if not program:
            program = self.program
        if not request_params:
            request_params = dict([(tier.name, 10) for tier in LevelTier.objects.filter(program=program)])
        response = self.client.get(reverse('bulk_import_indicators', args=[program.pk]), data=request_params)
        if response.status_code != 200:
            return response.content
        else:
            return ContentFile(response.content)

    def get_expected_indicator_numbers(self, config, request_params, has_indicators=False, auto_number=True):
        expected_numbers = []
        for level_conf in config:
            row_count = request_params[level_conf['tier']]
            row_count += level_conf['indicators'] if has_indicators else 0
            if auto_number:
                expected_numbers.append([f"{level_conf['ontology']}{letter}" for letter in self.letters[:row_count]])
            else:
                expected_numbers.append([None for i in range(row_count)])
        return expected_numbers

    def process_indicator_rows(self, wb, program):
        ws = wb.worksheets[0]
        counts = {'level_headers': 0, 'existing_indicators': [], 'new_indicators': []}
        level_names = []
        indicator_numbers = []
        level_index = -1
        for row_index in range(DATA_START_ROW, ws.max_row + 1):
            if isinstance(ws.cell(row_index, FIRST_USED_COLUMN + 1), MergedCell):
                # It's a level header
                level_names.append(ws.cell(row_index, FIRST_USED_COLUMN).value)
                indicator_numbers.append([])
                counts['level_headers'] += 1
                counts['existing_indicators'].append(0)
                counts['new_indicators'].append(0)
                level_index += 1
                self.assertEqual(ws.cell(row_index, FIRST_USED_COLUMN).style, LEVEL_HEADER_STYLE)
            elif ws.cell(row_index, FIRST_USED_COLUMN + 2).value is not None:
                counts['existing_indicators'][level_index] += 1
                if program.auto_number_indicators:
                    indicator_numbers[-1].append(ws.cell(row_index, FIRST_USED_COLUMN + 1).value)
                else:
                    indicator_numbers[-1].append(None)
                are_protected = [ws.cell(row_index, col).style == EXISTING_INDICATOR_STYLE
                                for col in range(FIRST_USED_COLUMN, len(COLUMNS))]
                self.assertTrue(all(are_protected))
            elif ws.cell(row_index, FIRST_USED_COLUMN).value is not None:
                counts['new_indicators'][level_index] += 1
                if program.auto_number_indicators:
                    indicator_numbers[-1].append(ws.cell(row_index, FIRST_USED_COLUMN + 1).value)
                else:
                    indicator_numbers[-1].append(None)
                # The first two columns are level and number, which have different protection than the rest of the row
                self.assertEqual(ws.cell(row_index, FIRST_USED_COLUMN).style, PROTECTED_NEW_INDICATOR_STYLE)
                expected_number_style = PROTECTED_NEW_INDICATOR_STYLE \
                    if program.auto_number_indicators else UNPROTECTED_NEW_INDICATOR_STYLE
                self.assertEqual(ws.cell(row_index, FIRST_USED_COLUMN + 1).style, expected_number_style)
                are_unprotected = [ws.cell(row_index, col).style == UNPROTECTED_NEW_INDICATOR_STYLE
                                   for col in range(FIRST_USED_COLUMN + 2, len(COLUMNS))]
                self.assertTrue(all(are_unprotected))
        return counts, level_names, indicator_numbers

    def compare_ws_lines(self, wb, expected_counts, expected_level_names, expected_indicator_numbers, program=None):
        program = program if program else self.program
        actual_counts, actual_level_names, actual_indicator_numbers = self.process_indicator_rows(wb, program)
        self.assertEqual(actual_counts, expected_counts)
        self.assertEqual(actual_level_names, expected_level_names)
        self.assertEqual(actual_indicator_numbers, expected_indicator_numbers)

    def test_permissions(self):
        response = self.client.get(reverse('bulk_import_indicators', args=[self.program.pk]))
        self.assertRedirects(response, '/accounts/login/')
        self.client.force_login(self.tola_user.user)
        response = self.client.get(reverse('bulk_import_indicators', args=[self.program.pk]))
        self.assertEqual(response.status_code, 403)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='medium')
        response = self.client.get(reverse('bulk_import_indicators', args=[self.program.pk]))
        self.assertEqual(response.status_code, 403)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        request_params = dict([(tier.name, 10) for tier in LevelTier.objects.filter(program=self.program)])
        response = self.client.get(reverse('bulk_import_indicators', args=[self.program.pk]), data=request_params)
        self.assertEqual(response.status_code, 200)

    def test_template_create(self):
        self.client.force_login(self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        request_params = dict([(tier.name, 10) for tier in LevelTier.objects.filter(program=self.program)])
        response = self.get_template(request_params=request_params)
        wb = openpyxl.load_workbook(response)
        ws = wb.worksheets[0]
        self.assertEqual(
            ws.cell(self.data_start_row, self.first_used_column).value,
            f'Goal: {self.ordered_levels[0].name}')

        request_params = {'bad tier 1': 10, 'bad tier 2': 80}
        response_content = self.get_template(request_params=request_params)
        self.assertEqual(json.loads(response_content)['error_codes'], [ERROR_MISMATCHED_TIERS])

    def test_row_counts(self):
        level_config = [
            {'name': 'Goal level', 'indicators': 2, 'ontology': '', 'customsort': 1, 'tier': 'Goal'},
            {'name': 'Outcome level 1', 'indicators': 0, 'ontology': '1', 'customsort': 1, 'tier': 'Outcome'},
            {'name': 'Output level 1', 'indicators': 2, 'ontology': '1.1', 'customsort': 1, 'tier': 'Output'},
            {'name': 'Activity level 1', 'indicators': 0, 'ontology': '1.1.1', 'customsort': 1, 'tier': 'Activity'},
            {'name': 'Outcome level 2', 'indicators': 3, 'ontology': '2', 'customsort': 2, 'tier': 'Outcome'},
            {'name': 'Output level 2', 'indicators': 0, 'ontology': '2.1', 'customsort': 1, 'tier': 'Output'},
            {'name': 'Activity level 2', 'indicators': 3, 'ontology': '2.1.1', 'customsort': 1, 'tier': 'Activity'},
        ]
        self.client.force_login(self.tola_user.user)
        small_program = w_factories.RFProgramFactory(country=[self.country], tiers=True)
        w_factories.grant_program_access(self.tola_user, small_program, self.country, role='high')
        request_params = {}
        tiers = self.program.level_tiers.all()
        for i, tier in enumerate(tiers):
            request_params[tier.name] = 10 if i < len(tiers) - 2 else 20

        parent_level = None
        for i, level in enumerate(level_config[:4]):
            level_obj = i_factories.LevelFactory(
                program=small_program,
                name=level_config[i]['name'],
                parent=parent_level,
                customsort=level_config[i]['customsort'])
            level_config[i]['level_obj'] = level_obj
            parent_level = level_obj

        # Test the defaults
        wb = openpyxl.load_workbook(self.get_template(program=small_program, request_params=request_params))
        expected_counts = {'level_headers': 4, 'existing_indicators': [0] * 4, 'new_indicators': [10, 10, 20, 20]}
        expected_level_names = [f"{item['tier']}{' ' + item['ontology'] if item['ontology'] else ''}: {item['name']}"
                                for item in level_config[:4]]
        expected_indicator_numbers = self.get_expected_indicator_numbers(level_config[:4], request_params)
        self.compare_ws_lines(wb, expected_counts, expected_level_names, expected_indicator_numbers)

        # Test protection of number column adjusts to program autonumber
        small_program.auto_number_indicators = False
        small_program.save()
        wb = openpyxl.load_workbook(self.get_template(program=small_program, request_params=request_params))
        expected_counts = {'level_headers': 4, 'existing_indicators': [0] * 4, 'new_indicators': [10, 10, 20, 20]}
        expected_level_names = [f"{item['tier']}{' ' + item['ontology'] if item['ontology'] else ''}: {item['name']}"
                                for item in level_config[:4]]
        expected_indicator_numbers = self.get_expected_indicator_numbers(
            level_config[:4], request_params, auto_number=small_program.auto_number_indicators)
        self.compare_ws_lines(
            wb, expected_counts, expected_level_names, expected_indicator_numbers, program=small_program)
        small_program.auto_number_indicators = True
        small_program.save()

        # Test with additional levels
        parent_level = level_config[0]['level_obj']
        for i in range(4, len(level_config)):
            level_obj = i_factories.LevelFactory(
                program=small_program,
                name=level_config[i]['name'],
                parent=parent_level,
                customsort=level_config[i]['customsort'])
            level_config[i]['level_obj'] = level_obj
            parent_level = level_obj

        wb = openpyxl.load_workbook(self.get_template(program=small_program, request_params=request_params))
        expected_counts = {
            'level_headers': 7, 'existing_indicators': [0] * 7, 'new_indicators': [10, 10, 20, 20, 10, 20, 20]}
        expected_level_names = [f"{item['tier']}{' ' + item['ontology'] if item['ontology'] else ''}: {item['name']}"
                                for item in level_config]
        expected_indicator_numbers = self.get_expected_indicator_numbers(level_config, request_params)
        self.compare_ws_lines(wb, expected_counts, expected_level_names, expected_indicator_numbers)

        # Add some indicators
        for level_dict in level_config:
            if level_dict['indicators'] > 0:
                i_factories.IndicatorFactory.create_batch(
                    level_dict['indicators'], program=small_program, level=level_dict['level_obj'])
        wb = openpyxl.load_workbook(self.get_template(program=small_program, request_params=request_params))
        expected_indicator_counts = [conf['indicators'] for conf in level_config]
        expected_counts = {
            'level_headers': 7,
            'existing_indicators': expected_indicator_counts,
            'new_indicators': [10, 10, 20, 20, 10, 20, 20]}
        expected_level_names = [f"{item['tier']}{' ' + item['ontology'] if item['ontology'] else ''}: {item['name']}"
                                for item in level_config]
        expected_indicator_numbers = self.get_expected_indicator_numbers(
            level_config, request_params, has_indicators=True)
        self.compare_ws_lines(wb, expected_counts, expected_level_names, expected_indicator_numbers)

        # Do some other request params
        for i, tier in enumerate(tiers):
            request_params[tier.name] = 5 if i < len(tiers) - 2 else 30
        wb = openpyxl.load_workbook(self.get_template(program=small_program, request_params=request_params))
        expected_counts = {
            'level_headers': 7,
            'existing_indicators': expected_indicator_counts,
            'new_indicators': [5, 5, 30, 30, 5, 30, 30]}
        expected_level_names = [f"{item['tier']}{' ' + item['ontology'] if item['ontology'] else ''}: {item['name']}"
                                for item in level_config]
        expected_indicator_numbers = self.get_expected_indicator_numbers(
            level_config, request_params, has_indicators=True)
        self.compare_ws_lines(wb, expected_counts, expected_level_names, expected_indicator_numbers)

        for i, tier in enumerate(tiers):
            request_params[tier.name] = 5 if i < len(tiers) - 2 else 0
        wb = openpyxl.load_workbook(self.get_template(program=small_program, request_params=request_params))
        expected_counts = {
            'level_headers': 7,
            'existing_indicators': expected_indicator_counts,
            'new_indicators': [5, 5, 0, 0, 5, 0, 0]}
        expected_level_names = [f"{item['tier']}{' ' + item['ontology'] if item['ontology'] else ''}: {item['name']}"
                                for item in level_config]
        expected_indicator_numbers = self.get_expected_indicator_numbers(
            level_config, request_params, has_indicators=True)
        self.compare_ws_lines(wb, expected_counts, expected_level_names, expected_indicator_numbers)

        for i, tier in enumerate(tiers):
            request_params[tier.name] = 0
        wb = openpyxl.load_workbook(self.get_template(program=small_program, request_params=request_params))
        expected_counts = {
            'level_headers': 7,
            'existing_indicators': expected_indicator_counts,
            'new_indicators': ([0] * 7)}
        expected_level_names = [f"{item['tier']}{' ' + item['ontology'] if item['ontology'] else ''}: {item['name']}"
                                for item in level_config]
        expected_indicator_numbers = self.get_expected_indicator_numbers(
            level_config, request_params, has_indicators=True)
        self.compare_ws_lines(wb, expected_counts, expected_level_names, expected_indicator_numbers)

        # Check a custom set of tiers
        level_obj = i_factories.LevelFactory(
            program=small_program,
            name='SubActivity 1',
            parent=level_config[-1]['level_obj'],
            customsort=1)
        i_factories.LevelTierFactory(program=small_program, name='SubActivity', tier_depth=5)
        i_factories.IndicatorFactory.create_batch(2, program=small_program, level=level_obj)
        new_tiers = LevelTier.objects.filter(program=small_program)
        new_level_config = copy.deepcopy(level_config)
        new_level_config.append(
            {'name': 'SubActivity 1', 'indicators': 2, 'ontology': '2.1.1.1',
             'customsort': 1, 'tier': 'SubActivity'})
        for i, tier in enumerate(new_tiers):
            request_params[tier.name] = 10 if i < len(new_tiers) - 2 else 20
        wb = openpyxl.load_workbook(self.get_template(program=small_program, request_params=request_params))
        expected_counts = {
            'level_headers': 8,
            'existing_indicators': expected_indicator_counts + [2],
            'new_indicators': [10, 10, 10, 20, 10, 10, 20, 20]}
        expected_level_names = [f"{item['tier']}{' ' + item['ontology'] if item['ontology'] else ''}: {item['name']}"
                                for item in new_level_config]
        expected_indicator_numbers = self.get_expected_indicator_numbers(
            new_level_config, request_params, has_indicators=True)
        self.compare_ws_lines(wb, expected_counts, expected_level_names, expected_indicator_numbers)

    def test_non_english_template_create(self):
        self.client.force_login(self.tola_user.user)
        self.tola_user.language = 'fr'
        self.tola_user.save()
        activate('fr')
        for sector in ['Agriculture', 'Conflict Management', 'Agribusiness']:
            w_factories.SectorFactory(sector=sector)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')

        wb = openpyxl.load_workbook(self.get_template())
        ws = wb.worksheets[0]
        first_level_header = ws.cell(DATA_START_ROW, FIRST_USED_COLUMN).value
        self.assertEqual(first_level_header[:3], 'But')
        hidden_ws = wb.get_sheet_by_name(HIDDEN_SHEET_NAME)
        self.assertEqual(hidden_ws['A3'].value, 'éTranslated Agribusiness')

        activate('en')
        self.tola_user.language = 'en'
        self.tola_user.save()

    @unittest.skip
    def test_data_validation(self):
        # openpyxl apparently does not allow you to read validations from a spreadsheet, only create new ones.
        # So for now, testing of validations is being deferred until openpyxl allows read and/or another package is
        # used.
        pass


class TestBulkImportTemplateProcessing(test.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.country = w_factories.CountryFactory()
        # cls.program = w_factories.RFProgramFactory(
        #     country=[cls.country], tiers=True, levels=6, indicators=150, indicators__levels=True)
        cls.program = w_factories.RFProgramFactory(
            country=[cls.country], tiers=True, levels=2, indicators=5, indicators__levels=True)
        cls.tola_user = w_factories.TolaUserFactory()
        w_factories.grant_program_access(cls.tola_user, cls.program, cls.country, role='high')
        cls.client = test.client
        cls.first_used_column = FIRST_USED_COLUMN
        cls.data_start_row = DATA_START_ROW
        cls.program_name_row = PROGRAM_NAME_ROW
        cls.sector = w_factories.SectorFactory()
        cls.template_file_path = os.path.join(settings.SITE_ROOT, BulkIndicatorImportFile.FILE_STORAGE_PATH)
        cls.name_sequence = 0

    def tearDown(self):
        for file_entry in os.scandir(self.template_file_path):
            if file_entry.name != BASE_TEMPLATE_NAME:
                file_path = os.path.join(self.template_file_path, file_entry.name)
                os.remove(file_path)

    def get_template(self):
        request_params = dict([(tier.name, 10) for tier in LevelTier.objects.filter(program=self.program)])
        response = self.client.get(reverse('bulk_import_indicators', args=[self.program.pk]), data=request_params)
        return ContentFile(response.content)

    def post_template(self, wb):
        with NamedTemporaryFile() as temp_file:
            wb.save(temp_file.name)
            return self.client.post(reverse('bulk_import_indicators', args=[self.program.id]), {'file': temp_file})

    def fill_worksheet_row(self, ws, row_index, custom_values=None):
        self.name_sequence += 1
        custom_values = [] if custom_values is None else custom_values
        for i, column in enumerate(COLUMNS):
            current_cell = ws.cell(row_index, FIRST_USED_COLUMN + i)
            if column['field_name'] in custom_values:
                current_cell.value = custom_values[column['field_name']]
                continue
            if column['field_name'] in ['level', 'number']:
                # These are pre-filled in the template so we can skip them
                continue
            if column['field_name'] == 'name':
                current_cell.value = f'Lorem ipsum name {self.name_sequence}'
            elif column['field_name'] == 'unit_of_measure_type':
                current_cell.value = str(Indicator.UNIT_OF_MEASURE_TYPES[0][1])
            elif column['field_name'] == 'direction_of_change':
                current_cell.value = str(Indicator.DIRECTION_OF_CHANGE[0][1])
            elif column['field_name'] == 'target_frequency':
                current_cell.value = str(Indicator.TARGET_FREQUENCIES[0][1])
            elif column['field_name'] == 'sector':
                current_cell.value = gettext(self.sector.sector)
            elif column['field_name'] == 'level':
                current_cell.value = gettext(Level.objects.first().name)
            elif column['field_name'] == 'baseline':
                current_cell.value = 3
            else:
                current_cell.value = 'Lorem ipsum ' + column['field_name']

    def check_row_highlights(self, ws, row_index, highlighted_cols=None):
        if highlighted_cols is None:
            highlighted_cols = []
        for col_index in range(1, len(COLUMNS) + 2):
            current_cell_fill = ws.cell(row_index, col_index).fill
            if col_index in highlighted_cols:
                self.assertEqual(current_cell_fill.fgColor.rgb, RED_ERROR)
            else:
                self.assertTrue(current_cell_fill is None or current_cell_fill.fgColor.rgb != RED_ERROR)

    def test_template_import_with_structural_problems(self):
        self.client.force_login(user=self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        self.assertEqual(Indicator.objects.count(), 5)

        wb = openpyxl.load_workbook(self.get_template())
        ws = wb.worksheets[0]

        ws.cell(self.program_name_row, self.first_used_column).value = "Wrong program name"
        response = self.post_template(wb)
        self.assertEqual(ProgramAuditLog.objects.count(), 1)
        self.assertEqual(ProgramAuditLog.objects.last().change_type, 'template_uploaded')
        self.assertEqual(response.status_code, 406)
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_MISMATCHED_PROGRAM])
        ws.cell(self.program_name_row, self.first_used_column).value = self.program.name
        response = self.post_template(wb)
        self.assertEqual(ProgramAuditLog.objects.count(), 2)
        self.assertEqual(ProgramAuditLog.objects.last().change_type, 'template_uploaded')
        self.assertEqual(response.status_code, 406)
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_NO_NEW_INDICATORS])

    def test_successful_template_upload(self):
        self.client.force_login(user=self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        wb = openpyxl.load_workbook(self.get_template())
        existing_goal_indicator_count = Indicator.objects.filter(level__parent=None).count()
        first_blank_goal_row = DATA_START_ROW + existing_goal_indicator_count + 1
        ws = wb.worksheets[0]
        self.fill_worksheet_row(ws, first_blank_goal_row)
        self.fill_worksheet_row(ws, first_blank_goal_row + 1)
        response = self.post_template(wb)
        self.assertEqual(ProgramAuditLog.objects.count(), 1)
        self.assertEqual(ProgramAuditLog.objects.last().change_type, 'template_uploaded')
        self.assertEqual(response.status_code, 200)
        bulk_upload_objs = BulkIndicatorImportFile.objects.filter(
            program=self.program, user=self.tola_user, file_type=BulkIndicatorImportFile.INDICATOR_DATA_TYPE)
        self.assertEqual(len(bulk_upload_objs), 1, 'There should be a new Bulk...File entry in the DB')
        self.assertEqual(
            len(os.listdir(self.template_file_path)),
            2,
            'There should be a indicator json file saved to the file system.')
        # Need to sleep to avoid file name collision
        time.sleep(1)
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 200)
        bulk_upload_objs = BulkIndicatorImportFile.objects.filter(
            program=self.program, user=self.tola_user, file_type=BulkIndicatorImportFile.INDICATOR_DATA_TYPE)
        self.assertEqual(len(bulk_upload_objs), 1, 'Old template file entry should have been deleted')
        self.assertEqual(
            len(os.listdir(self.template_file_path)), 2, 'Old indicator json file should have been deleted.')

    def test_non_fatal_import_errors(self):
        self.client.force_login(user=self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        wb = openpyxl.load_workbook(self.get_template())
        goal_level = Level.objects.get(parent=None, program=self.program)
        existing_goal_indicator_count = Indicator.objects.filter(level__parent=None).count()
        existing_outcome_indicator_count = Indicator.objects.filter(level__parent=goal_level).count()
        first_blank_goal_row = DATA_START_ROW + existing_goal_indicator_count + 1
        first_blank_outcome_row = first_blank_goal_row + 12 + existing_outcome_indicator_count
        ws = wb.worksheets[0]

        # Make sure we got the outcome row calc right
        self.assertEqual(ws.cell(
            first_blank_outcome_row - existing_outcome_indicator_count - 1,
            FIRST_USED_COLUMN
        ).style,  LEVEL_HEADER_STYLE)

        # Create one row for each required column where the column value is missing
        expected_errors = {}
        current_row_index = first_blank_goal_row + 1
        for required_col_index, col in enumerate(COLUMNS):
            if col['required'] and col['field_name'] != 'level':
                self.fill_worksheet_row(ws, current_row_index, custom_values={col['field_name']: None})
                expected_errors[current_row_index] = col['field_name']
                current_row_index += 1

        # Fill the next row with bad values for all of the validated fields
        validated_fields_bad_values = {
            'sector': 'bad sector',
            'unit_of_measure_type': 'therbligs',
            'direction_of_change': 'north by northwest',
            'target_frequency': 'whats the frequency kenneth',
        }
        self.fill_worksheet_row(ws, current_row_index, validated_fields_bad_values)
        bad_values_row = current_row_index

        # In the first outcome section, skip a couple of rows and then add valid data
        self.fill_worksheet_row(ws, first_blank_outcome_row + 2)

        # Check if upload was logged
        response = self.post_template(wb)
        self.assertEqual(ProgramAuditLog.objects.count(), 1)
        self.assertEqual(ProgramAuditLog.objects.last().change_type, 'template_uploaded')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            json.loads(response.content)['error_codes'],
            [ERROR_MALFORMED_INDICATOR, ERROR_INTERVENING_BLANK_ROW])

        # Get the feedback template and see if it looks right
        response = self.client.get(reverse('get_feedback_bulk_import_template', args=[self.program.id]))
        self.assertEqual(response.get('Content-Disposition'), 'attachment; filename="Marked-up-template.xlsx"')
        feedback_wb = openpyxl.load_workbook(ContentFile(response.content))
        feedback_ws = feedback_wb.worksheets[0]

        for required_row_index, field_name in expected_errors.items():
            self.check_row_highlights(
                feedback_ws,
                required_row_index,
                highlighted_cols=[1, self.first_used_column + COLUMNS_FIELD_INDEXES[field_name]]
            )

        highlighted_cols = [1] + [self.first_used_column + COLUMNS_FIELD_INDEXES[field_name]
            for field_name in validated_fields_bad_values.keys()]
        self.check_row_highlights(feedback_ws, bad_values_row, highlighted_cols)

        self.check_row_highlights(feedback_ws, first_blank_outcome_row, list(range(1, 2+len(COLUMNS))))
        self.check_row_highlights(feedback_ws, first_blank_outcome_row + 1, list(range(1, 2 + len(COLUMNS))))
        self.check_row_highlights(feedback_ws, first_blank_outcome_row + 2, [])

        wb = openpyxl.load_workbook(self.get_template())
        existing_goal_indicator_count = Indicator.objects.filter(level__parent=None).count()
        first_blank_goal_row = DATA_START_ROW + existing_goal_indicator_count + 1
        ws = wb.worksheets[0]
        self.fill_worksheet_row(ws, first_blank_goal_row, custom_values={
            'sector': 'bad sector name'})
        response = self.post_template(wb)
        self.assertEqual(ProgramAuditLog.objects.count(), 2)
        self.assertEqual(response.status_code, 400)
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_MALFORMED_INDICATOR])

    def test_non_english_imports(self):
        self.tola_user.language = 'fr'
        self.tola_user.save()
        self.client.force_login(user=self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        wb = openpyxl.load_workbook(self.get_template())
        existing_goal_indicator_count = Indicator.objects.filter(level__parent=None).count()
        first_blank_goal_row = DATA_START_ROW + existing_goal_indicator_count + 1
        ws = wb.worksheets[0]
        self.fill_worksheet_row(ws, first_blank_goal_row)
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(json.loads(response.content), {'message': 'success', 'valid': 1, 'invalid': 0})
        self.tola_user.language = 'en'
        self.tola_user.save()

    def test_fatal_import_errors(self):
        self.client.force_login(user=self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        wb = openpyxl.load_workbook(self.get_template())
        existing_goal_indicator_count = Indicator.objects.filter(level__parent=None).count()
        first_blank_goal_row = DATA_START_ROW + existing_goal_indicator_count + 1
        ws = wb.worksheets[0]

        ws.cell(self.program_name_row, self.first_used_column).value = "bad program name"
        response = self.post_template(wb)
        self.assertEqual(ProgramAuditLog.objects.count(), 1)
        self.assertEqual(ProgramAuditLog.objects.last().change_type, 'template_uploaded')
        self.assertEqual(response.status_code, 406)
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_MISMATCHED_PROGRAM])
        ws.cell(self.program_name_row, self.first_used_column).value = self.program.name

        response = self.post_template(wb)
        self.assertEqual(response.status_code, 406)
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_NO_NEW_INDICATORS])

        header_cell = ws.cell(COLUMN_HEADER_ROW, self.first_used_column)
        orig_header_value = header_cell.value
        header_cell.value = "bad column header"
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 406)
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_MISMATCHED_HEADERS])
        header_cell.value = orig_header_value

        should_be_indicator = ws.cell(first_blank_goal_row + 6, self.first_used_column)
        orig_style = should_be_indicator.style
        should_be_indicator.value = 'bad_level_name'
        should_be_indicator.style = LEVEL_HEADER_STYLE
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 406)
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_INVALID_LEVEL_HEADER])
        should_be_indicator.style = orig_style

        second_level_header = ws.cell(first_blank_goal_row + 11, self.first_used_column)
        orig_value = second_level_header.value
        second_level_header.value = 'bad_level_name'
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 406)
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_MISMATCHED_LEVEL_COUNT])
        second_level_header.value = orig_value

        first_header_cell = ws.cell(DATA_START_ROW, FIRST_USED_COLUMN)
        orig_header_value = first_header_cell.value
        first_header_cell.value = 'bad level name'
        self.fill_worksheet_row(ws, first_blank_goal_row)
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 406)
        self.assertEqual(
            json.loads(response.content)['error_codes'],[ERROR_INVALID_LEVEL_HEADER, ERROR_UNDETERMINED_LEVEL])
        first_header_cell.value = orig_header_value

        number_cell = ws.cell(first_blank_goal_row + 2, FIRST_USED_COLUMN + COLUMNS_FIELD_INDEXES['number'])
        orig_value = number_cell.value
        self.fill_worksheet_row(ws, first_blank_goal_row + 2, {'number': 'bad number'})
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 406)
        self.assertEqual(
            json.loads(response.content)['error_codes'], [ERROR_UNEXPECTED_INDICATOR_NUMBER])
        number_cell.value = orig_value

        self.fill_worksheet_row(ws, first_blank_goal_row + 3, {'level': 'bad level'})
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 406)
        self.assertEqual(
            json.loads(response.content)['error_codes'], [ERROR_UNEXPECTED_LEVEL])
        self.fill_worksheet_row(ws, first_blank_goal_row + 3)

    def test_saving_indicators(self):
        self.client.force_login(user=self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        wb = openpyxl.load_workbook(self.get_template())
        ws = wb.worksheets[0]
        ws.cell(PROGRAM_NAME_ROW, FIRST_USED_COLUMN).value = self.program.name
        self.fill_worksheet_row(ws, DATA_START_ROW + 2, custom_values={
            'name': 'Test Number',
            'unit_of_measure_type': str(dict(Indicator.UNIT_OF_MEASURE_TYPES)[Indicator.NUMBER]),
            'baseline': .3
        })
        self.fill_worksheet_row(ws, DATA_START_ROW + 3, custom_values={
            'name': 'Test Percent',
            'unit_of_measure_type': str(dict(Indicator.UNIT_OF_MEASURE_TYPES)[Indicator.PERCENTAGE]),
            'baseline': .3
        })
        self.post_template(wb)
        self.assertEqual(ProgramAuditLog.objects.count(), 1, 'After upload, there should be a single audit log record')
        self.assertEqual(
            ProgramAuditLog.objects.last().change_type,
            'template_uploaded',
            'The audit log record should be of the right type')
        self.assertEqual(Indicator.objects.count(), 5, 'We should be starting with the expected number of indicators')
        self.client.post(reverse('save_bulk_import_data', args=[self.program.id]))
        self.assertEqual(
            ProgramAuditLog.objects.count(),
            3,
            'There should be two audit log records for the indicator saves and one for the upload')
        self.assertEqual(ProgramAuditLog.objects.last().change_type, 'indicator_imported')
        self.assertEqual(Indicator.objects.count(), 7, 'There should be two more indicators than before')
        i_number = Indicator.objects.get(name="Test Number")
        self.assertEqual(
            make_quantized_decimal(i_number.baseline),
            make_quantized_decimal(.3),
            "The number value should be saved in its raw form")
        i_percent = Indicator.objects.get(name="Test Percent")
        self.assertEqual(
            make_quantized_decimal(i_percent.baseline),
            make_quantized_decimal(30),
            "The percent value should be saved as 100 x its input form, since its a percent")

        response = self.client.post(reverse('save_bulk_import_data', args=[self.program.id]))
        self.assertEqual(
            json.loads(response.content)['error_codes'],
            [ERROR_INDICATOR_DATA_NOT_FOUND],
            'Hitting the save endpoint should error because the stored json data has been deleted by the prior save')

        bulk_import_json = BulkIndicatorImportFile.objects.create(
            user=self.tola_user, program=self.program, file_type=BulkIndicatorImportFile.INDICATOR_DATA_TYPE,
            file_name='test.json')
        indicator_data = copy.deepcopy(SAMPLE_INDICATOR_DATA)
        indicator_data.pop('program_id')
        indicator_data.pop('sector_id')
        with open(bulk_import_json.file_path, 'w+') as fh:
            fh.write(json.dumps([indicator_data]))
        response = self.client.post(reverse('save_bulk_import_data', args=[self.program.id]))
        self.assertEqual(
            response.status_code, 400, "If the json file fails to save properly, it should produce and error")
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_SAVE_VALIDATION])
        self.assertEqual(Indicator.objects.count(), 7)

    def test_duplicated_names(self):
        self.client.force_login(user=self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')

        # Now make sure we have two existing indicators that have the same name
        goal_level = Level.objects.get(parent=None, program=self.program)
        target_indicator = Indicator.objects.filter(program=self.program, level=goal_level).first()
        i_factories.RFIndicatorFactory(program=self.program, level=goal_level, name=target_indicator.name)

        wb = openpyxl.load_workbook(self.get_template())
        ws = wb.worksheets[0]
        ws.cell(PROGRAM_NAME_ROW, FIRST_USED_COLUMN).value = self.program.name

        name_column = FIRST_USED_COLUMN + COLUMNS_FIELD_INDEXES['name']
        existing_goal_indicator_count = Indicator.objects.filter(level__parent=None).count()
        last_goal_row = DATA_START_ROW + existing_goal_indicator_count

        self.fill_worksheet_row(ws, last_goal_row + 1, custom_values={'name': 'Unique Name'})
        response = self.post_template(wb)
        self.assertEqual(
            response.status_code,
            200,
            "Template should have been submitted successfully even though two existing indicators have the same name")

        last_goal_row_name = ws.cell(last_goal_row, name_column).value
        self.fill_worksheet_row(ws, last_goal_row + 1, custom_values={'name': last_goal_row_name})
        self.fill_worksheet_row(ws, last_goal_row + 2, custom_values={'name': last_goal_row_name})
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 400)
        response = self.client.get(reverse('get_feedback_bulk_import_template', args=[self.program.id]))
        feedback_wb = openpyxl.load_workbook(ContentFile(response.content))
        feedback_ws = feedback_wb.worksheets[0]
        self.assertEqual(
            feedback_ws.cell(last_goal_row + 1, name_column).comment.text,
            ERROR_MSG_NAME_IN_DB,
            'Indicator name already in DB, comment should reflect appropriate error message')
        self.assertEqual(
            feedback_ws.cell(last_goal_row + 2, name_column).comment.text,
            ERROR_MSG_NAME_IN_DB,
            'Indicator name already in DB, comment should reflect appropriate error message')

        self.fill_worksheet_row(ws, last_goal_row + 1, custom_values={'name': 'Duplicate name'})
        self.fill_worksheet_row(ws, last_goal_row + 2, custom_values={'name': 'Duplicate name'})
        response = self.post_template(wb)
        self.assertEqual(response.status_code, 400)
        response = self.client.get(reverse('get_feedback_bulk_import_template', args=[self.program.id]))
        feedback_wb = openpyxl.load_workbook(ContentFile(response.content))
        feedback_ws = feedback_wb.worksheets[0]
        self.assertEqual(
            feedback_ws.cell(last_goal_row + 1, name_column).comment.text,
            ERROR_MSG_NAME_DUPLICATED,
            "Duplicate indicator names should cause duplicate name comment to be shown on cell")
        self.assertEqual(
            feedback_ws.cell(last_goal_row + 2, name_column).comment.text,
            ERROR_MSG_NAME_DUPLICATED,
            "Duplicate indicator names should cause duplicate name comment to be shown on cell")


    def test_bad_lookup_values(self):
        self.client.force_login(user=self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        indicator_data = copy.deepcopy(SAMPLE_INDICATOR_DATA)
        indicator_data['program_id'] = self.program.pk
        indicator_data['sector_id'] = 9999
        indicator_data = [indicator_data]
        file_obj = BulkIndicatorImportFile.objects.create(
            user=self.tola_user,
            program=self.program,
            file_name='test_file',
            file_type=BulkIndicatorImportFile.INDICATOR_DATA_TYPE)
        with open(file_obj.file_path, 'w') as fh:
            fh.write(json.dumps(indicator_data, cls=DjangoJSONEncoder))
        response = self.client.post(reverse('save_bulk_import_data', args=[self.program.id]))
        self.assertEqual(
            response.status_code, 400, "If the json file fails to save properly, it should produce and error")
        self.assertEqual(json.loads(response.content)['error_codes'], [ERROR_SAVE_VALIDATION])

    def test_retrieving_feedback_template(self):
        self.client.force_login(user=self.tola_user.user)
        w_factories.grant_program_access(self.tola_user, self.program, self.country, role='high')
        wb = openpyxl.load_workbook(self.get_template())
        ws = wb.worksheets[0]
        ws.cell(PROGRAM_NAME_ROW, FIRST_USED_COLUMN).value = self.program.name
        self.fill_worksheet_row(ws, DATA_START_ROW + 2, {'name': None})
        self.fill_worksheet_row(ws, DATA_START_ROW + 3)
        self.post_template(wb)
        response = self.client.get(reverse('get_feedback_bulk_import_template', args=[self.program.id]))
        self.assertEqual(response.get('Content-Disposition'),'attachment; filename="Marked-up-template.xlsx"')

        # Make sure feedback template was deleted by the first request in this test.
        response = self.client.get(reverse('get_feedback_bulk_import_template', args=[self.program.id]))
        self.assertEqual(
            json.loads(response.content)['error_codes'], [ERROR_TEMPLATE_NOT_FOUND])

