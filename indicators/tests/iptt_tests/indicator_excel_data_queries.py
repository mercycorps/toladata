"""Tests for the IPTT excel endpoint serializer/renderer classes (checking query counts and speeds)

    Excel Endpoint functional structure:
    indicators.views.views_reports.IPTTExcelReport
        - processes request data, instances correct IPTTSerializer
    indicators.serializers.IPTTSerializer
        - instanced with a report type class and request data
        - new - returns class with request data in class dict
        - init - prefetches appropriate data

"""

import unittest
import openpyxl

from django import test

from indicators.views.views_reports import IPTTExcelReport
from indicators.serializers import IPTTSerializer
from indicators.queries.iptt_queries import IPTTIndicator
from indicators.models import Indicator
from factories.workflow_models import CountryFactory, RFProgramFactory
from factories.indicators_models import RFIndicatorFactory

SPECIAL_CHARACTERS_NAME = "Thîs Nåmé has ßpeçial Chars"
SPECIAL_CHARACTERS_TIER = "Tîér 2"
SPECIAL_CHARACTERS_NUMBER = "Cüstom Nümbéring"

class TestRequest:
    defaults = {
        'frequency': Indicator.ANNUAL
    }
    def __init__(self, **kwargs):
        self._data = {**self.defaults, **kwargs}

    def get(self, *args):
        default = [] if len(args) == 1 else args[1]
        return self._data.get(args[0], default)

    def getlist(self, *args):
        default = [] if len(args) == 1 else args[1]
        value = self._data.get(args[0], default)
        if not isinstance(value, list):
            value = list(value)
        return value

    def keys(self):
        return self._data.keys()

def get_timeperiods_report(program_id, **kwargs):
    request = TestRequest(programId=program_id, **kwargs)
    return IPTTSerializer(IPTTSerializer.TIMEPERIODS_EXCEL, request)

def get_tva_report(program_id, **kwargs):
    request = TestRequest(programId=program_id, **kwargs)
    return IPTTSerializer(IPTTSerializer.TVA_EXCEL, request)

def get_full_tva_report(program_id, **kwargs):
    request = TestRequest(programId=program_id, **kwargs)
    return IPTTSerializer(IPTTSerializer.TVA_FULL_EXCEL, request)


def get_timeperiods_report_indicators(program_id, **kwargs):
    report = get_timeperiods_report(program_id, **kwargs)
    report.initialize()
    return {indicator.pk: indicator for indicator in report.indicators}

# QUERY COUNTS:
PROGRAM_DATA_PREFETCH = 1
PROGRAM_LEVELS_PREFETCH = 1
TP_INDICATOR_PREFETCH = 6

class RFProgramMixin:
    @classmethod
    def get_rf_program(cls):
        cls.country = CountryFactory(country="TestLand", code="TL")
        cls.program = RFProgramFactory(
            name=SPECIAL_CHARACTERS_NAME,
            tiers=["Tier 1", SPECIAL_CHARACTERS_TIER],
            levels=1
        )
        cls.level1 = [level for level in cls.program.levels.all() if level.level_depth == 1][0]
        cls.level2 = [level for level in cls.program.levels.all() if level.level_depth == 2][0]
        cls.program.country.set([cls.country])


@unittest.skip("fixing")
class TestPrefetchQueryCounts(RFProgramMixin, test.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.get_rf_program()
        cls.indicator1 = RFIndicatorFactory(
            program=cls.program,
            level=cls.level1
        )
        cls.indicator2a = RFIndicatorFactory(
            program=cls.program,
            level=cls.level2
        )
        cls.indicator2b = RFIndicatorFactory(
            program=cls.program,
            level=cls.level2
        )
        cls.blank_row_indicator1 = RFIndicatorFactory(
            program=cls.program,
            level=None
        )
        cls.blank_row_indicator2 = RFIndicatorFactory(
            program=cls.program,
            level=None
        )
        cls.test_workbook = openpyxl.Workbook()
        cls.test_sheet = None

    def setUp(self):
        if self.test_sheet:
            self.test_workbook.remove(self.test_sheet)
        self.test_sheet = self.test_workbook.create_sheet('Test sheet')

    def get_all_serializers(self):
        with self.assertNumQueries(0):
            tp_serializer = get_timeperiods_report(self.program.pk)
            tva_serializer = get_tva_report(self.program.pk)
            full_tva_serializer = get_full_tva_report(self.program.pk)
        return [tp_serializer, tva_serializer, full_tva_serializer]

    def test_program_data_queries(self):
        for serializer in self.get_all_serializers():
            with self.assertNumQueries(PROGRAM_DATA_PREFETCH):
                serializer.init_program_data()
                self.assertEqual(serializer.program_data['pk'], self.program.pk)
                self.assertEqual(serializer.program_name, self.program.name)
                self.assertFalse(serializer.level_column)
                self.assertEqual(serializer.program_data['reporting_period_start'],
                                 self.program.reporting_period_start)
                self.assertEqual(serializer.program_data['reporting_period_end'],
                                 self.program.reporting_period_end)

    def test_program_level_queries(self):
        for serializer in self.get_all_serializers():
            serializer.init_program_data()
            with self.assertNumQueries(PROGRAM_LEVELS_PREFETCH):
                serializer.init_level_data()
                self.assertEqual([level for level in serializer._levels], [self.level1, self.level2])
                self.assertEqual([level._tier for level in serializer._levels], [self.level1.leveltier, self.level2.leveltier])

    @unittest.skip("fixing")
    def test_indicator_annotation_queries(self):
        (tp_serializer, tva_serializer, full_tva_serializer) = self.get_all_serializers()
        tp_serializer.init_program_data()
        tp_serializer.init_level_data()
        with self.assertNumQueries(TP_INDICATOR_PREFETCH):
            tp_serializer.init_indicator_data()
            self.assertEqual(len(tp_serializer.indicators), 5)
            tp_serializer.renderer_class(tp_serializer).add_headers(self.test_sheet)
            self.assertEqual(self.test_sheet.cell(row=1, column=3).value, "Indicator Performance Tracking Report")
            self.assertEqual(len(tp_serializer.blank_level_row), 2)


@unittest.skip("fixing")
class TestNumberingQueries(RFProgramMixin, test.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.get_rf_program()
        RFIndicatorFactory(
            program=cls.program,
            level=cls.level1
        )
        cls.indicator1 = RFIndicatorFactory(
            program=cls.program,
            level=cls.level1
        )
        cls.indicator2 = RFIndicatorFactory(
            program=cls.program,
            level=cls.level2
        )
        cls.custom_number_program = RFProgramFactory(
            tiers=["Tier 1", "Tier 2"],
            levels=1,
            auto_number_indicators=False
        )
        cls.custom_number_program.country.set([cls.country])
        cls.level2_custom = [level for level in cls.custom_number_program.levels.all() if level.level_depth == 2][0]
        cls.indicator_custom = RFIndicatorFactory(
            program=cls.custom_number_program,
            level=cls.level2_custom,
            number=SPECIAL_CHARACTERS_NUMBER
        )

    def test_indicator_numbers_no_queries(self):
        self.assertEqual(self.indicator1.number_display, 'Tier 1 b')
        self.assertEqual(self.indicator2.number_display, f'{SPECIAL_CHARACTERS_TIER} 1a')
        self.assertEqual(self.indicator_custom.number_display, SPECIAL_CHARACTERS_NUMBER)

    def test_serializer_numbers_are_correct(self):
        tp_report_indicators = get_timeperiods_report_indicators(self.program.pk)
        self.assertEqual(tp_report_indicators[self.indicator1.pk].number_display, 'Tier 1 b')
        self.assertEqual(tp_report_indicators[self.indicator2.pk].number_display, f'{SPECIAL_CHARACTERS_TIER} 1a')
        tp_custom_report_indicators = get_timeperiods_report_indicators(self.custom_number_program.pk)
        self.assertEqual(
            tp_custom_report_indicators[self.indicator_custom.pk].number_display, SPECIAL_CHARACTERS_NUMBER
        )
