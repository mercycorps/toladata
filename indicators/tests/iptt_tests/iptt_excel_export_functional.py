"""Functional tests for the IPTT Excel export - test overall excel export for completeness and accuracy

    Report types:
        TvA full
        TvA LoP
        TvA MidEnd
        TvA regular (SEMI_ANNUAL)
        TP ANNUAL
        TP Monthly
    Languages:
        English
        French
        Spanish
    Programs:
        Auto-numbering
        Non-auto-numbering
    Indicators:
        Numeric NC
        Numeric Cumulative
        Percent
    Results:
        None
        One in a period
        Many in a period
    Disaggregations:
        None
        One with no results
        One with some results
        Many with no results
        Many with some results
    Display Filters:
        Start period
        End period
        Grouping
        Columns Hidden
        Rows expanded
    Indicator Filters:
        Sites
        Types
        Sectors
        Indicators
        Disaggregations (with disaggs shown accordingly)


Indicators:
    LoP - a - "1.1.1" - Numeric NC - baseline 0 - No results - standard disagg - no sites - no types - sector 1
    MID_END - b - "aaa" - Percent - baseline na - One result Mid (month 2) - country disagg (no disagg results) - site 1 - type 1 - no sector
    ANNUAL - 1a - "x4.1" - Numeric Cum - baseline 100 - Two results month 4,5 - no disaggs - sites 1 and 2 - types 1 + 2 - sector 2
    QUARTERLY - 2a - "4a" - Numeric NC - baseline 0 - Two results month 2 One result month 10 - standard disagg and country disagg (all results disagg'd) - no sites - type 2 - sector 1
    TRI_ANNUAL - 2b - "4b" - Percent - baseline na - No results - no disaggs - site 1 - no types - no sector
    MONTHLY - 1.1a - "21.4" - Numeric Cum - baseline 100 - One result month 3 One result month 6 - standard disagg (disagg'd result) - sites 1 and 2 - type 1 - sector 2
    EVENT - 1.1b - "19.1" - Numeric NC - baseline 0 - One result month 1 event 1 - country disagg (disagg'd result) - no sites - types 1 + 2 - sector 1

"""

import io
import datetime
import locale
import itertools
import unittest
import openpyxl
from django import test
from django.urls import reverse
from django.utils import translation
from indicators.models import Indicator
from indicators.tests.iptt_tests.iptt_scenario import IPTTScenarioBuilder
from factories.workflow_models import (
    CountryFactory,
    RFProgramFactory,
    TolaUserFactory,
    SectorFactory,
    SiteProfileFactory,
)
from factories.indicators_models import (
    RFIndicatorFactory,
    LevelFactory,
    LevelTierFactory,
    DisaggregationTypeFactory,
    IndicatorTypeFactory,
    PeriodicTargetFactory,
    ResultFactory,
    DisaggregatedValueFactory,
)



ENGLISH = 1
FRENCH = 2
SPANISH = 3

DATE_FORMATS = {
    ENGLISH: lambda d: d.strftime('%b %-d, %Y'),
    FRENCH: lambda d: d.strftime('%-d %b. %Y'),
    SPANISH: lambda d: d.strftime('%-d %b. %Y').title()
}

ENGLISH_COLUMNS = [
    'No.',
    'Indicator',
    'Unit of measure',
    'Change',
    'C / NC',
    '# / %',
    'Baseline'
]

FRENCH_COLUMNS = [
    'Nº',
    'Indicateur',
    'Unité de mesure',
    'Changement',
    'C / NC',
    '# / %',
    'Base de référence'
]

SPANISH_COLUMNS = [
    'No.',
    'Indicador',
    'Unidad de medida',
    'Cambio',
    'C / NC',
    '# / %',
    'Base'
]

REPORT_TITLE = {
    ENGLISH: "Indicator Performance Tracking Report",
    FRENCH: "Rapport de suivi des performances de l’indicateur",
    SPANISH: "Informe de seguimiento del rendimiento del indicador",
}

LOP_PERIOD = {
    ENGLISH: "Life of Program",
    FRENCH: "Vie du programme",
    SPANISH: "Vida del programa",
}

TARGET = {
    ENGLISH: "Target",
    FRENCH: "Cible",
    SPANISH: "Objetivo",
}

ACTUAL = {
    ENGLISH: "Actual",
    FRENCH: "Réel",
    SPANISH: "Real",
}

MET = {
    ENGLISH: "% Met",
    FRENCH: "% Atteint",
    SPANISH: "% Cumplido",
}


SPECIAL_CHARS = "Spécîal Chårs"


class TestScenarioBuilder(test.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.scenario = IPTTScenarioBuilder()

    def test_indicators(self):
        indicators = self.scenario.indicators
        for c, frequency in enumerate([
            Indicator.LOP,
            Indicator.MID_END,
            Indicator.ANNUAL,
            Indicator.QUARTERLY,
            Indicator.TRI_ANNUAL,
            Indicator.MONTHLY,
            Indicator.EVENT
        ]):
            self.assertEqual(indicators[c].target_frequency, frequency)


class TestIPTTExcelExports(test.TestCase):
    iptt_url = reverse('iptt_excel')

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.scenario = IPTTScenarioBuilder()
        cls.tolauser = TolaUserFactory()
        cls.tolauser.user.is_superuser = True
        cls.tolauser.user.save()
        cls.client = test.Client()

    def login_client(self, language):
        if language == FRENCH:
            self.tolauser.language = 'fr'
            self.tolauser.save()
        elif language == SPANISH:
            self.tolauser.language = 'es'
            self.tolauser.save()
        self.client.force_login(user=self.tolauser.user)

    def get_tva_report_full(self, response=False, language=ENGLISH, **kwargs):
        self.login_client(language)
        get_response = self.client.get(
            self.iptt_url,
            {
                **{
                    'programId': self.scenario.program.pk,
                    'fullTVA': 'true'
                    },
                **kwargs
        })
        self.assertEqual(get_response.status_code, 200)
        self.client.logout()
        self.tolauser.language = 'en'
        self.tolauser.save()
        if response:
            return get_response
        return openpyxl.load_workbook(io.BytesIO(get_response.content))

    def get_tva_report(self, frequency=Indicator.LOP, response=False, language=ENGLISH, **kwargs):
        self.login_client(language)
        get_response = self.client.get(
            self.iptt_url,
            { **{
                'programId': self.scenario.program.pk,
                'fullTVA': 'false',
                'reportType': '1',
                'frequency': frequency,
                },
             **kwargs
            }
        )
        self.assertEqual(get_response.status_code, 200)
        self.client.logout()
        self.tolauser.language = 'en'
        self.tolauser.save()
        if response:
            return get_response
        return openpyxl.load_workbook(io.BytesIO(get_response.content))

    def get_tp_report(self, frequency=Indicator.ANNUAL, response=False, language=ENGLISH, **kwargs):
        self.login_client(language)
        get_response = self.client.get(
            self.iptt_url,
            { **{
                'programId': self.scenario.program.pk,
                'fullTVA': 'false',
                'reportType': '2',
                'frequency': frequency,
                },
             **kwargs
            }
        )
        self.assertEqual(get_response.status_code, 200)
        self.client.logout()
        self.tolauser.language = 'en'
        self.tolauser.save()
        if response:
            return get_response
        return openpyxl.load_workbook(io.BytesIO(get_response.content))

    def assert_response_headers(self, response, filename):
        self.assertEqual(
            response.get('Content-Type'),
            "application/ms-excel"
        )
        self.assertEqual(
            response.get('Content-Disposition'),
            f'attachment; filename="{filename}"'
        )

    def test_responses(self):
        tva_full_response = self.get_tva_report_full(response=True)
        filename = f"IPTT TvA full program report {datetime.date.today().strftime('%b %-d, %Y')}.xlsx"
        self.assert_response_headers(tva_full_response, filename)
        tva_response = self.get_tva_report(response=True)
        filename = f"IPTT TvA report {datetime.date.today().strftime('%b %-d, %Y')}.xlsx"
        self.assert_response_headers(tva_response, filename)
        tp_response = self.get_tp_report(response=True)
        filename = f"IPTT Actuals only report {datetime.date.today().strftime('%b %-d, %Y')}.xlsx"
        self.assert_response_headers(tp_response, filename)

    def test_responses_french(self):
        default_locale = locale.getlocale()
        locale.setlocale(locale.LC_ALL, 'fr_FR')
        tva_full_response = self.get_tva_report_full(response=True, language=FRENCH)
        filename = f"Rapport IPTT relatif à la totalité de la TVA du programme {datetime.date.today().strftime('%-d %b. %Y').lower()}.xlsx"
        self.assert_response_headers(tva_full_response, filename)
        tva_response = self.get_tva_report(response=True, language=FRENCH)
        filename = f"Rapport TVA IPTT {datetime.date.today().strftime('%-d %b. %Y').lower()}.xlsx"
        self.assert_response_headers(tva_response, filename)
        tp_response = self.get_tp_report(response=True, language=FRENCH)
        filename = f"Rapport IPTT relatif aux valeurs réelles {datetime.date.today().strftime('%-d %b. %Y').lower()}.xlsx"
        self.assert_response_headers(tp_response, filename)
        locale.setlocale(locale.LC_ALL, default_locale)

    def test_responses_spanish(self):
        default_locale = locale.getlocale()
        locale.setlocale(locale.LC_ALL, 'es_ES')
        tva_full_response = self.get_tva_report_full(response=True, language=SPANISH)
        filename = f"Informe completo del programa del IPTT TvA {datetime.date.today().strftime('%-d %b. %Y').title()}.xlsx"
        self.assert_response_headers(tva_full_response, filename)
        tva_response = self.get_tva_report(response=True, language=SPANISH)
        filename = f"Informe del IPTT TvA {datetime.date.today().strftime('%-d %b. %Y').title()}.xlsx"
        self.assert_response_headers(tva_response, filename)
        tp_response = self.get_tp_report(response=True, language=SPANISH)
        filename = f"Informes de reales únicamente del IPTT {datetime.date.today().strftime('%-d %b. %Y').title()}.xlsx"
        self.assert_response_headers(tp_response, filename)
        locale.setlocale(locale.LC_ALL, default_locale)

    def assert_iptt_title_cells(self, sheet, language=ENGLISH):
        default_locale = locale.getlocale()
        if language == FRENCH:
            locale.setlocale(locale.LC_ALL, 'fr_FR')
        elif language == SPANISH:
            locale.setlocale(locale.LC_ALL, 'es_ES')
        self.assertEqual(sheet.cell(row=1, column=3).value, REPORT_TITLE[language])
        self.assertEqual(
            sheet.cell(row=2, column=3).value,
            f"{DATE_FORMATS[language](self.scenario.start_date)} – {DATE_FORMATS[language](self.scenario.end_date)}",
        )
        self.assertEqual(
            sheet.cell(row=3, column=3).value,
            self.scenario.program_name
        )
        
        locale.setlocale(locale.LC_ALL, default_locale)

    def assert_lop_header_cells(self, sheet, language=ENGLISH, column=10):
        self.assertEqual(
            sheet.cell(row=3, column=column).value,
            LOP_PERIOD[language]
        )
        self.assertEqual(
            sheet.cell(row=4, column=column).value,
            TARGET[language]
        )
        self.assertEqual(
            sheet.cell(row=4, column=column+1).value,
            ACTUAL[language]
        )
        self.assertEqual(
            sheet.cell(row=4, column=column+2).value,
            MET[language]
        )

    def assert_period_tva_header_cells(self, sheet, period, column, language=ENGLISH, subheader=None):
        if subheader is None:
            self.assertEqual(
                sheet.cell(row=3, column=column).value,
                period
            )
        else:
            self.assertEqual(
                sheet.cell(row=2, column=column).value,
                period
            )
            self.assertEqual(
                sheet.cell(row=3, column=column).value,
                subheader
            )
        self.assertEqual(
            sheet.cell(row=4, column=column).value,
            TARGET[language]
        )
        self.assertEqual(
            sheet.cell(row=4, column=column+1).value,
            ACTUAL[language]
        )
        self.assertEqual(
            sheet.cell(row=4, column=column+2).value,
            MET[language]
        )

    def assert_period_tp_header_cells(self, sheet, period, column, language=ENGLISH, subheader=None):
        if subheader is None:
            self.assertEqual(
                sheet.cell(row=3, column=column).value,
                period
            )
        else:
            self.assertEqual(
                sheet.cell(row=2, column=column).value,
                period
            )
            self.assertEqual(
                sheet.cell(row=3, column=column).value,
                subheader
            )
        self.assertEqual(
            sheet.cell(row=4, column=column).value,
            ACTUAL[language]
        )
        

    def test_tva_full_report_headers(self):
        tva_full_report = self.get_tva_report_full()
        self.assertEqual(len(tva_full_report.worksheets), 7)
        self.assertEqual(
            tva_full_report.sheetnames,
            [
                "Life of Program (LoP) only",
                "Midline and endline",
                "Annual",
                "Tri-annual",
                "Quarterly",
                "Monthly",
                "Change log"
            ])
        for sheet in tva_full_report.worksheets[:-1]:
            self.assert_iptt_title_cells(sheet)
            for column, header in enumerate(ENGLISH_COLUMNS):
                self.assertEqual(
                    sheet.cell(row=4, column=3+column).value,
                    header
                )
            self.assert_lop_header_cells(sheet)
        midend = tva_full_report.worksheets[1]
        for col, period in enumerate(["Midline", "Endline"]):
            self.assert_period_tva_header_cells(midend, period, 13+3*col)
        change_log = tva_full_report.worksheets[-1]
        self.assertEqual(change_log.cell(row=1, column=1).value, "Change log")
        self.assertEqual(change_log.cell(row=2, column=1).value, self.scenario.program_name)
        for column, header in enumerate(
            ['Date and time', 'Result level', 'Indicator', 'User', 'Organization', 'Change type',
             'Previous entry', 'New entry', 'Rationale']
        ):
            self.assertEqual(change_log.cell(row=3, column=1+column).value, header)


    def test_tva_lop_report(self):
        tva_lop_report = self.get_tva_report(frequency=Indicator.LOP)
        self.assertEqual(len(tva_lop_report.worksheets), 1)
        sheet = tva_lop_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)
        self.assertEqual(
            sheet.cell(row=5, column=3).value,
            self.scenario.goal_level_row
            )
        self.assertEqual(
            sheet.cell(row=6, column=3).value,
            "Tîér1 a"
        )
        for col, val in enumerate([
            self.scenario.indicators[0].name,
            self.scenario.indicators[0].unit_of_measure,
            None,
            "Not cumulative",
            "#",
            0,
            1000,
            "–",
            "–",
            None]):
            self.assertEqual(
                sheet.cell(row=6, column=4+col).value,
                val
            )
        self.assertIsNone(sheet.cell(row=7, column=3).value)

    def test_tva_midend_report_headers(self):
        tva_midend_report = self.get_tva_report(frequency=Indicator.MID_END)
        self.assertEqual(len(tva_midend_report.worksheets), 1)
        sheet = tva_midend_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)
        for col, period in enumerate(["Midline", "Endline"]):
            self.assert_period_tva_header_cells(sheet, period, 13+3*col)
        self.assertEqual(
            sheet.cell(row=5, column=3).value,
            self.scenario.goal_level_row
            )
        self.assertEqual(
            sheet.cell(row=6, column=3).value,
            "Tîér1 b"
        )
        self.assertEqual(
            sheet.cell(row=6, column=4).value,
            self.scenario.indicators[1].name
        )
        self.assertEqual(
            sheet.cell(row=7, column=3).value,
            self.scenario.country_disagg.disaggregation_type
        )
        for c, label in enumerate(["Label 1", "Låbél 2", "Label 3"]):
            self.assertEqual(
                sheet.cell(row=7+c, column=4).value,
                label
            )
            for col in range(10):
                self.assertEqual(
                    sheet.cell(row=7+c, column=9+col).value,
                    "–"
                )
            self.assertIsNone(sheet.cell(row=7+c, column=19).value)
            self.assertTrue(sheet.row_dimensions[7+c].hidden, f"row {c}")
        self.assertIsNone(sheet.cell(row=10, column=3).value)

    def test_tva_quarterly_report_headers(self):
        tva_quarterly_report = self.get_tva_report(frequency=Indicator.QUARTERLY)
        self.assertEqual(len(tva_quarterly_report.worksheets), 1)
        sheet = tva_quarterly_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)
        for c in range(12):
            start_date = datetime.date(
                self.scenario.start_date.year + c//4,
                ((c % 4)*3)+1,
                1
            )
            end_date = datetime.date(
                self.scenario.start_date.year + (c+1)//4,
                (((c+1)%4)*3)+1,
                1
            ) - datetime.timedelta(days=1)
            period = f"Quarter {c+1}"
            subheader = f"{start_date.strftime('%b %-d, %Y')} – {end_date.strftime('%b %-d, %Y')}"
            self.assert_period_tva_header_cells(sheet, period, 13+3*c, subheader=subheader)
        self.assertEqual(
            sheet.cell(row=5, column=3).value,
            self.scenario.goal_level_row
        )
        self.assertEqual(
            sheet.cell(row=6, column=3).value,
            "Tîér2 2: Lévêl outcome 1b"
        )
        for c, val in enumerate([
            "Tîér2 2a",
            self.scenario.indicators[3].name,
            self.scenario.indicators[3].unit_of_measure,
            None, "Not cumulative", "#", 0, 10, 20, 2, 0.83, 11, 13.253,
            0.83, "–", "–", 0.83, "–", "–", 0.83, 9, 10.8434
            ]):
            self.assertEqual(
                sheet.cell(row=7, column=3+c).value,
                val
            )
        self.assertIsNone(sheet.cell(row=7, column=49).value)
        self.assertEqual(
            sheet.cell(row=8, column=3).value,
            self.scenario.country_disagg.disaggregation_type
        )
        target_columns = [10, 13, 16, 19, 22]
        met_columns = [12, 15, 18, 21]
        for c, label in enumerate(["Label 1", "Låbél 2", "Label 3"]):
            self.assertEqual(sheet.cell(row=8+c, column=4).value, label)
            # baseline:
            self.assertEqual(sheet.cell(row=8+c, column=9).value, "–")
            for column in target_columns:
                self.assertEqual(sheet.cell(row=8+c, column=column).value, "–")
            for column in met_columns:
                self.assertEqual(sheet.cell(row=8+c, column=column).value, "–")
            self.assertEqual(sheet.cell(row=8+c, column=11).value, 14 if c == 2 else 3)
            self.assertEqual(sheet.cell(row=8+c, column=14).value, 7 if c == 2 else 2)
            self.assertEqual(sheet.cell(row=8+c, column=17).value, "–")
            self.assertEqual(sheet.cell(row=8+c, column=20).value, "–")
            self.assertEqual(sheet.cell(row=8+c, column=23).value, 7 if c == 2 else 1)
        self.assertEqual(
            sheet.cell(row=11, column=3).value,
            self.scenario.standard_disagg.disaggregation_type
        )
        for c, label in enumerate(["Label 1", "Låbél 2"]):
            self.assertEqual(sheet.cell(row=11+c, column=4).value, label)
            # baseline:
            self.assertEqual(sheet.cell(row=11+c, column=9).value, "–")
            for column in target_columns:
                self.assertEqual(sheet.cell(row=8+c, column=column).value, "–")
            for column in met_columns:
                self.assertEqual(sheet.cell(row=8+c, column=column).value, "–")
            self.assertEqual(sheet.cell(row=11+c, column=11).value, 17 if c == 1 else 3)
            self.assertEqual(sheet.cell(row=11+c, column=14).value, 9 if c == 1 else 2)
            self.assertEqual(sheet.cell(row=11+c, column=17).value, "–")
            self.assertEqual(sheet.cell(row=11+c, column=20).value, "–")
            self.assertEqual(sheet.cell(row=11+c, column=23).value, 8 if c == 1 else 1)
        self.assertIsNone(sheet.cell(row=13, column=3).value)
        

    def test_tva_semi_annual_report_headers_spanish(self):
        default_locale = locale.getlocale()
        locale.setlocale(locale.LC_ALL, 'es_ES')
        tva_semi_annual_report = self.get_tva_report(frequency=Indicator.SEMI_ANNUAL, language=SPANISH)
        self.assertEqual(len(tva_semi_annual_report.worksheets), 1)
        sheet = tva_semi_annual_report.worksheets[0]
        self.assert_iptt_title_cells(sheet, language=SPANISH)
        for column, header in enumerate(SPANISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet, language=SPANISH)
        for c in range(6):
            start_date = datetime.date(
                self.scenario.start_date.year + c//2,
                1 if c % 2 == 0 else 7,
                1
            )
            end_date = datetime.date(
                self.scenario.start_date.year + (c+1)//2,
                1 if c % 2 == 1 else 7,
                1
            ) - datetime.timedelta(days=1)
            period = f"Períodos semestrales {c+1}"
            subheader = f"{start_date.strftime('%-d %b. %Y').title()} – {end_date.strftime('%-d %b. %Y').title()}"
            self.assert_period_tva_header_cells(sheet, period, 13+3*c, subheader=subheader, language=SPANISH)
        locale.setlocale(locale.LC_ALL, default_locale)

    def test_tp_annual_report_headers(self):
        tp_annual_report = self.get_tp_report(frequency=Indicator.ANNUAL)
        self.assertEqual(len(tp_annual_report.worksheets), 1)
        sheet = tp_annual_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)
        for c in range(3):
            start_date = datetime.date(self.scenario.start_date.year + c, 1, 1)
            end_date = datetime.date(self.scenario.start_date.year + c + 1, 1, 1) - datetime.timedelta(days=1)
            period = f"Year {c+1}"
            subheader = f"{start_date.strftime('%b %-d, %Y')} – {end_date.strftime('%b %-d, %Y')}"
            self.assert_period_tp_header_cells(sheet, period, 13+c, subheader=subheader)
        for row_num, level_name in [
            (5, self.scenario.goal_level_row),
            (11, "Tîér2 1: Lévêl outcome 1a"),
            (13, "Tîér3 1.1: Lévêl output 1.1a"),
            (21, "Tîér2 2: Lévêl outcome 1b"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, level_name)
        for row_num, indicator_name, indicator_number, year_1 in [
            (6, self.scenario.indicators[0].name, "Tîér1 a", "–"),
            (7, self.scenario.indicators[1].name, "Tîér1 b", 0.95),
            (12, self.scenario.indicators[2].name, "Tîér2 1a", 70),
            (14, self.scenario.indicators[5].name, "Tîér3 1.1a", 110),
            (17, self.scenario.indicators[6].name, "Tîér3 1.1b", 42),
            (22, self.scenario.indicators[3].name, "Tîér2 2a", 20),
            (28, self.scenario.indicators[4].name, "Tîér2 2b", "–"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, indicator_number)
            self.assertEqual(sheet.cell(row=row_num, column=4).value, indicator_name)
            self.assertEqual(sheet.cell(row=row_num, column=13).value, year_1)

    def test_tp_annual_report_headers_manual_numbering(self):
        self.scenario.program.auto_number_indicators = False
        self.scenario.program.save()
        tp_annual_report = self.get_tp_report(frequency=Indicator.ANNUAL)
        self.scenario.program.auto_number_indicators = True
        self.scenario.program.save()
        self.assertEqual(len(tp_annual_report.worksheets), 1)
        sheet = tp_annual_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)
        for c in range(3):
            start_date = datetime.date(self.scenario.start_date.year + c, 1, 1)
            end_date = datetime.date(self.scenario.start_date.year + c + 1, 1, 1) - datetime.timedelta(days=1)
            period = f"Year {c+1}"
            subheader = f"{start_date.strftime('%b %-d, %Y')} – {end_date.strftime('%b %-d, %Y')}"
            self.assert_period_tp_header_cells(sheet, period, 13+c, subheader=subheader)
        for row_num, level_name in [
            (5, self.scenario.goal_level_row),
            (11, "Tîér2 1: Lévêl outcome 1a"),
            (13, "Tîér3 1.1: Lévêl output 1.1a"),
            (21, "Tîér2 2: Lévêl outcome 1b"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, level_name)
        for row_num, indicator_name, indicator_number in [
            (6, self.scenario.indicators[0].name, "1.1.1"),
            (7, self.scenario.indicators[1].name, "aaa"),
            (12, self.scenario.indicators[2].name, "x4.1"),
            (14, self.scenario.indicators[5].name, "21.4"),
            (17, self.scenario.indicators[6].name, "19.1"),
            (22, self.scenario.indicators[3].name, "4a"),
            (28, self.scenario.indicators[4].name, "4b"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, indicator_number)
            self.assertEqual(sheet.cell(row=row_num, column=4).value, indicator_name)

    def test_tp_semi_annual_report_headers_sites_filter(self):
        tp_semi_annual_report = self.get_tp_report(frequency=Indicator.SEMI_ANNUAL, sites=[self.scenario.site1.pk])
        self.assertEqual(len(tp_semi_annual_report.worksheets), 1)
        sheet = tp_semi_annual_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)
        for row_num, level_name in [
            (5, self.scenario.goal_level_row),
            (10, "Tîér2 1: Lévêl outcome 1a"),
            (12, "Tîér3 1.1: Lévêl output 1.1a"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, level_name)
        for row_num, indicator_name, indicator_number in [
            (6, self.scenario.indicators[1].name, "Tîér1 b"),
            (11, self.scenario.indicators[2].name, "Tîér2 1a"),
            (13, self.scenario.indicators[5].name, "Tîér3 1.1a"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, indicator_number)
            self.assertEqual(sheet.cell(row=row_num, column=4).value, indicator_name)
        self.assertIsNone(sheet.cell(row=16, column=3).value)

    def test_tp_tri_annual_report_headers_disaggs_filter(self):
        tp_tri_annual_report = self.get_tp_report(
            frequency=Indicator.TRI_ANNUAL, disaggregations=[self.scenario.standard_disagg.pk])
        self.assertEqual(len(tp_tri_annual_report.worksheets), 1)
        sheet = tp_tri_annual_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)
        for row_num, level_name in [
            (5, self.scenario.goal_level_row),
            (6, "Tîér3 1.1: Lévêl output 1.1a"),
            (10, "Tîér2 2: Lévêl outcome 1b"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, level_name)
        for row_num, indicator_name, indicator_number in [
            (7, self.scenario.indicators[5].name, "Tîér3 1.1a"),
            (11, self.scenario.indicators[3].name, "Tîér2 2a"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, indicator_number)
            self.assertEqual(sheet.cell(row=row_num, column=4).value, indicator_name)
            self.assertEqual(sheet.cell(row=row_num+1, column=3).value, self.scenario.standard_disagg.disaggregation_type)
            self.assertEqual(sheet.cell(row=row_num+1, column=4).value, "Label 1")
            self.assertFalse(sheet.row_dimensions[row_num+1].hidden)
            self.assertEqual(sheet.cell(row=row_num+2, column=4).value, "Låbél 2")
            self.assertFalse(sheet.row_dimensions[row_num+2].hidden) # don't collapse disagg rows when filtered by disaggs
        self.assertIsNone(sheet.cell(row=14, column=3).value)
        

    def test_tp_annual_report_headers_columns_missing(self):
        tp_annual_report = self.get_tp_report(frequency=Indicator.ANNUAL, columns=['1', '2', '4'])
        self.assertEqual(len(tp_annual_report.worksheets), 1)
        sheet = tp_annual_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        columns = [column for c, column in enumerate(ENGLISH_COLUMNS) if c not in (3, 4, 6)]
        for column, header in enumerate(columns):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet, column=7)
        for c in range(3):
            start_date = datetime.date(self.scenario.start_date.year + c, 1, 1)
            end_date = datetime.date(self.scenario.start_date.year + c + 1, 1, 1) - datetime.timedelta(days=1)
            period = f"Year {c+1}"
            subheader = f"{start_date.strftime('%b %-d, %Y')} – {end_date.strftime('%b %-d, %Y')}"
            self.assert_period_tp_header_cells(sheet, period, 10+c, subheader=subheader)
        for row_num, level_name in [
            (5, self.scenario.goal_level_row),
            (11, "Tîér2 1: Lévêl outcome 1a"),
            (13, "Tîér3 1.1: Lévêl output 1.1a"),
            (21, "Tîér2 2: Lévêl outcome 1b"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, level_name)

    def test_tp_monthly_report_headers_group_by_level(self):
        tp_monthly_report = self.get_tp_report(frequency=Indicator.MONTHLY, groupby='2')
        self.assertEqual(len(tp_monthly_report.worksheets), 1)
        sheet = tp_monthly_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)
        for c in range(36):
            start_date = datetime.date(self.scenario.start_date.year + c//12, 1 + (c % 12), 1)
            period = f"{start_date.strftime('%B %Y')}"
            self.assert_period_tp_header_cells(sheet, period, 13+c)
        for row_num, level_name in [
            (5, self.scenario.goal_level_row),
            (11, "Tîér2 1: Lévêl outcome 1a"),
            (13, "Tîér2 2: Lévêl outcome 1b"),
            (21, "Tîér3 1.1: Lévêl output 1.1a"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, level_name)
        for row_num, indicator_name, indicator_number in [
            (6, self.scenario.indicators[0].name, "Tîér1 a"),
            (7, self.scenario.indicators[1].name, "Tîér1 b"),
            (12, self.scenario.indicators[2].name, "Tîér2 1a"),
            (14, self.scenario.indicators[3].name, "Tîér2 2a"),
            (20, self.scenario.indicators[4].name, "Tîér2 2b"),
            (22, self.scenario.indicators[5].name, "Tîér3 1.1a"),
            (25, self.scenario.indicators[6].name, "Tîér3 1.1b"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, indicator_number)
            self.assertEqual(sheet.cell(row=row_num, column=4).value, indicator_name)

    def test_tp_report_headers_group_by_level_and_type_filter(self):
        tp_monthly_report = self.get_tp_report(
            frequency=Indicator.MONTHLY, groupby='2', types=[self.scenario.type1.pk]
        )
        self.assertEqual(len(tp_monthly_report.worksheets), 1)
        sheet = tp_monthly_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)
        for row_num, level_name in [
            (5, self.scenario.goal_level_row),
            (10, "Tîér2 1: Lévêl outcome 1a"),
            (12, "Tîér3 1.1: Lévêl output 1.1a"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, level_name)
        for row_num, indicator_name, indicator_number in [
            (6, self.scenario.indicators[1].name, "Tîér1 b"),
            (11, self.scenario.indicators[2].name, "Tîér2 1a"),
            (13, self.scenario.indicators[5].name, "Tîér3 1.1a"),
            (16, self.scenario.indicators[6].name, "Tîér3 1.1b"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, indicator_number)
            self.assertEqual(sheet.cell(row=row_num, column=4).value, indicator_name)
        self.assertIsNone(sheet.cell(row=20, column=4).value)

    def test_tp_monthly_report_headers_sector_filter(self):
        tp_monthly_report = self.get_tp_report(
            frequency=Indicator.MONTHLY, sectors=[self.scenario.sector1.pk, self.scenario.sector2.pk]
        )
        self.assertEqual(len(tp_monthly_report.worksheets), 1)
        sheet = tp_monthly_report.worksheets[0]
        self.assert_iptt_title_cells(sheet)
        for column, header in enumerate(ENGLISH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet)

        for row_num, level_name in [
            (5, self.scenario.goal_level_row),
            (7, "Tîér2 1: Lévêl outcome 1a"),
            (9, "Tîér3 1.1: Lévêl output 1.1a"),
            (17, "Tîér2 2: Lévêl outcome 1b"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, level_name)
        for row_num, indicator_name, indicator_number in [
            (6, self.scenario.indicators[0].name, "Tîér1 a"),
            (8, self.scenario.indicators[2].name, "Tîér2 1a"),
            (10, self.scenario.indicators[5].name, "Tîér3 1.1a"),
            (13, self.scenario.indicators[6].name, "Tîér3 1.1b"),
            (18, self.scenario.indicators[3].name, "Tîér2 2a"),
        ]:
            self.assertEqual(sheet.cell(row=row_num, column=3).value, indicator_number)
            self.assertEqual(sheet.cell(row=row_num, column=4).value, indicator_name)
        self.assertIsNone(sheet.cell(row=24, column=3).value)

    def test_tp_monthly_report_headers_french(self):
        default_locale = locale.getlocale()
        locale.setlocale(locale.LC_ALL, 'fr_FR')
        tp_monthly_report = self.get_tp_report(frequency=Indicator.MONTHLY, language=FRENCH)
        self.assertEqual(len(tp_monthly_report.worksheets), 1)
        sheet = tp_monthly_report.worksheets[0]
        self.assert_iptt_title_cells(sheet, language=FRENCH)
        for column, header in enumerate(FRENCH_COLUMNS):
            self.assertEqual(
                sheet.cell(row=4, column=3+column).value,
                header
            )
        self.assert_lop_header_cells(sheet, language=FRENCH)
        for c in range(36):
            start_date = datetime.date(self.scenario.start_date.year + c//12, 1 + (c % 12), 1)
            period = f"{start_date.strftime('%B %Y')}"
            self.assert_period_tp_header_cells(sheet, period, 13+c, language=FRENCH)
        locale.setlocale(locale.LC_ALL, default_locale)