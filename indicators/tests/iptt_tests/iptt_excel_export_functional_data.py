"""Tests for the numeric elements of IPTT exports including disaggregated data"""


import io
import openpyxl
from django import test
from django.urls import reverse
from indicators.tests.iptt_tests.iptt_scenario import (
    IPTTScenarioSums
)
from factories.workflow_models import TolaUserFactory

from indicators.models import Indicator

BLANK_CELL='–'

class TestIPTTDisaggregatedSummativeData(test.TestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.scenario = IPTTScenarioSums()
        cls.tolauser = TolaUserFactory()
        cls.tolauser.user.is_superuser = True
        cls.tolauser.user.save()
        cls.client = test.Client()

    def test_timeperiods_report(self):
        self.client.force_login(user=self.tolauser.user)
        response = self.client.get(
            reverse('iptt_excel'),
            {
                'programId': self.scenario.program.pk,
                'fullTVA': 'false',
                'reportType': '2',
                'frequency': Indicator.SEMI_ANNUAL
            })
        self.assertEqual(response.status_code, 200)
        report = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = report.worksheets[0]
        assert self.lop_column_tests(sheet)
        assert self.tp_period_column_tests(sheet)

    def test_tva_report(self):
        self.client.force_login(user=self.tolauser.user)
        response = self.client.get(
            reverse('iptt_excel'),
            {
                'programId': self.scenario.program.pk,
                'fullTVA': 'false',
                'reportType': '1',
                'frequency': Indicator.QUARTERLY
            })
        self.assertEqual(response.status_code, 200)
        report = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = report.worksheets[0]
        assert self.lop_column_tests(sheet)
        assert self.tva_period_column_tests(sheet)

    def lop_column_tests(self, sheet):
        # Numeric / Non-cumulative / No results
        row = 6
        # baseline should be 100 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=9).value, 100)
        # baseline should be - for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=9).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=9).value, BLANK_CELL)
        # lop target should be 1000 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=10).value, 1000)
        # lop target should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=10).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=10).value, BLANK_CELL)
        # lop actual should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=11).value, BLANK_CELL)
        # lop actual should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=11).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=11).value, BLANK_CELL)
        # lop percent met should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=12).value, BLANK_CELL)
        # lop percent met should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=12).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=12).value, BLANK_CELL)

        # Numeric / Non-cumulative / One result
        row = 9
        # baseline should be 100 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=9).value, 100)
        # baseline should be - for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=9).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=9).value, BLANK_CELL)
        # lop target should be 1000 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=10).value, 1000)
        # lop target should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=10).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=10).value, BLANK_CELL)
        # lop actual should be 500 for indicator row
        self.assertEqual(sheet.cell(row=row, column=11).value, 500)
        # lop actual should be 100,400 for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=11).value, 100)
        self.assertEqual(sheet.cell(row=row+2, column=11).value, 400)
        # lop percent met should be .5 (50%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=12).value, 0.5)
        self.assertEqual(sheet.cell(row=row, column=12).number_format, '0%')
        # lop percent met should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=12).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=12).value, BLANK_CELL)

        # Numeric / Non-cumulative / Many results
        row = 12
        # baseline should be 100 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=9).value, 100)
        # baseline should be - for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=9).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=9).value, BLANK_CELL)
        # lop target should be 1000 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=10).value, 1000)
        # lop target should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=10).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=10).value, BLANK_CELL)
        # lop actual should be 1500 for indicator row
        self.assertEqual(sheet.cell(row=row, column=11).value, 1500)
        # lop actual should be 300,1200 for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=11).value, 300)
        self.assertEqual(sheet.cell(row=row+2, column=11).value, 1200)
        # lop percent met should be 1.5 (150%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=12).value, 1.5)
        self.assertEqual(sheet.cell(row=row, column=12).number_format, '0%')
        # lop percent met should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=12).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=12).value, BLANK_CELL)

        # Numeric / cumulative / No results
        row = 15
        # baseline should be 100 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=9).value, 100)
        # baseline should be - for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=9).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=9).value, BLANK_CELL)
        # lop target should be 1000 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=10).value, 250)
        # lop target should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=10).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=10).value, BLANK_CELL)
        # lop actual should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=11).value, BLANK_CELL)
        # lop actual should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=11).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=11).value, BLANK_CELL)
        # lop percent met should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=12).value, BLANK_CELL)
        # lop percent met should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=12).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=12).value, BLANK_CELL)

        # Numeric / cumulative / One result
        row = 18
        # baseline should be 100 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=9).value, 100)
        # baseline should be - for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=9).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=9).value, BLANK_CELL)
        # lop target should be 1000 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=10).value, 250)
        # lop target should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=10).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=10).value, BLANK_CELL)
        # lop actual should be 500 for indicator row
        self.assertEqual(sheet.cell(row=row, column=11).value, 500)
        # lop actual should be 100,400 for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=11).value, 100)
        self.assertEqual(sheet.cell(row=row+2, column=11).value, 400)
        # lop percent met should be 2 (200%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=12).value, 2)
        self.assertEqual(sheet.cell(row=row, column=12).number_format, '0%')
        # lop percent met should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=12).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=12).value, BLANK_CELL)

        # Numeric / cumulative / Many results
        row = 21
        # baseline should be 100 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=9).value, 100)
        # baseline should be - for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=9).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=9).value, BLANK_CELL)
        # lop target should be 1000 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=10).value, 250)
        # lop target should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=10).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=10).value, BLANK_CELL)
        # lop actual should be 1500 for indicator row
        self.assertEqual(sheet.cell(row=row, column=11).value, 1500)
        # lop actual should be 300,1200 for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=11).value, 300)
        self.assertEqual(sheet.cell(row=row+2, column=11).value, 1200)
        # lop percent met should be 6 (600%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=12).value, 6)
        self.assertEqual(sheet.cell(row=row, column=12).number_format, '0%')
        # lop percent met should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=12).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=12).value, BLANK_CELL)

        # Percentage / No results
        row = 24
        # baseline should be 100 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=9).value, 1)
        self.assertEqual(sheet.cell(row=row, column=9).number_format, '0%')
        # baseline should be - for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=9).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=9).value, BLANK_CELL)
        # lop target should be 10 (1000%) for indicator row:
        self.assertEqual(sheet.cell(row=row, column=10).value, 2.5)
        self.assertEqual(sheet.cell(row=row, column=10).number_format, '0%')
        # lop target should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=10).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=10).value, BLANK_CELL)
        # lop actual should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=11).value, BLANK_CELL)
        # lop actual should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=11).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=11).value, BLANK_CELL)
        # lop percent met should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=12).value, BLANK_CELL)
        # lop percent met should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=12).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=12).value, BLANK_CELL)

        # Percentage / One result
        row = 27
        # baseline should be 100 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=9).value, 1)
        self.assertEqual(sheet.cell(row=row, column=9).number_format, '0%')
        # baseline should be - for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=9).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=9).value, BLANK_CELL)
        # lop target should be 10 (1000%) for indicator row:
        self.assertEqual(sheet.cell(row=row, column=10).value, 2.5)
        self.assertEqual(sheet.cell(row=row, column=10).number_format, '0%')
        # lop target should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=10).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=10).value, BLANK_CELL)
        # lop actual should be 5 (500%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=11).value, 5)
        self.assertEqual(sheet.cell(row=row, column=11).number_format, '0%')
        # lop actual should be 1 (100%), 4 (400%) for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=11).value, 1)
        self.assertEqual(sheet.cell(row=row+1, column=11).number_format, '0%')
        self.assertEqual(sheet.cell(row=row+2, column=11).value, 4)
        self.assertEqual(sheet.cell(row=row+2, column=11).number_format, '0%')
        # lop percent met should be 2 (200%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=12).value, 2)
        self.assertEqual(sheet.cell(row=row, column=12).number_format, '0%')
        # lop percent met should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=12).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=12).value, BLANK_CELL)

        # Percentage / Many results
        row = 30
        # baseline should be 100 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=9).value, 1)
        self.assertEqual(sheet.cell(row=row, column=9).number_format, '0%')
        # baseline should be - for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=9).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=9).value, BLANK_CELL)
        # lop target should be 10 (1000%) for indicator row:
        self.assertEqual(sheet.cell(row=row, column=10).value, 2.5)
        self.assertEqual(sheet.cell(row=row, column=10).number_format, '0%')
        # lop target should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=10).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=10).value, BLANK_CELL)
        # lop actual should be 5(500%) for indicator row (percentage takes most recent)
        self.assertEqual(sheet.cell(row=row, column=11).value, 5)
        self.assertEqual(sheet.cell(row=row, column=11).number_format, '0%')
        # lop actual should be 1(100%),4(400%) for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=11).value, 1)
        self.assertEqual(sheet.cell(row=row+1, column=11).number_format, '0%')
        self.assertEqual(sheet.cell(row=row+2, column=11).value, 4)
        self.assertEqual(sheet.cell(row=row+2, column=11).number_format, '0%')
        # lop percent met should be 2 (200%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=12).value, 2)
        self.assertEqual(sheet.cell(row=row, column=12).number_format, '0%')
        # lop percent met should be – for disaggregation rows
        self.assertEqual(sheet.cell(row=row+1, column=12).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=12).value, BLANK_CELL)

        return True

    def tp_period_column_tests(self, sheet):
        # sheet should only have 14 columns (2 semi annual periods, one year program)
        self.assertEqual(sheet.max_column, 14)
        # Numeric / Non-cumulative / No results
        row = 6
        # actual semi_annual period 1 should be – for indicator row:
        self.assertEqual(sheet.cell(row=row, column=13).value, BLANK_CELL)
        # actual semi_annual period 1 should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, BLANK_CELL)
        # actual semi_annual period 2 should be – for indicator row:
        self.assertEqual(sheet.cell(row=row, column=14).value, BLANK_CELL)
        # actual semi_annual period 2 should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, BLANK_CELL)

        # Numeric / Non-cumulative / One result
        row = 9
        # actual semi_annual period 1 should be 500 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=13).value, 500)
        # actual semi_annual period 1 should be 100, 400 for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, 100)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, 400)
        # actual semi_annual period 2 should be – for indicator row:
        self.assertEqual(sheet.cell(row=row, column=14).value, BLANK_CELL)
        # actual semi_annual period 2 should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, BLANK_CELL)

        # Numeric / Non-cumulative / Many results
        row = 12
        # actual semi_annual period 1 should be 1000 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=13).value, 1000)
        # actual semi_annual period 1 should be 200, 800 for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, 200)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, 800)
        # actual semi_annual period 2 should be 500 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=14).value, 500)
        # actual semi_annual period 2 should be 100,400 for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, 100)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, 400)

        # Numeric / cumulative / No results
        row = 15
        # actual semi_annual period 1 should be – for indicator row:
        self.assertEqual(sheet.cell(row=row, column=13).value, BLANK_CELL)
        # actual semi_annual period 1 should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, BLANK_CELL)
        # actual semi_annual period 2 should be – for indicator row:
        self.assertEqual(sheet.cell(row=row, column=14).value, BLANK_CELL)
        # actual semi_annual period 2 should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, BLANK_CELL)


        # Numeric / cumulative / One result
        row = 18
        # actual semi_annual period 1 should be 500 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=13).value, 500)
        # actual semi_annual period 1 should be 100, 400 for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, 100)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, 400)
        # actual semi_annual period 2 should be – for indicator row:
        self.assertEqual(sheet.cell(row=row, column=14).value, BLANK_CELL)
        # actual semi_annual period 2 should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, BLANK_CELL)

        # Numeric / cumulative / Many results
        row = 21
        # actual semi_annual period 1 should be 1000 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=13).value, 1000)
        # actual semi_annual period 1 should be 200, 800 for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, 200)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, 800)
        # actual semi_annual period 2 should be 1500 for indicator row:
        self.assertEqual(sheet.cell(row=row, column=14).value, 1500)
        # actual semi_annual period 2 should be 300,1200 for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, 300)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, 1200)

        # Percentage / No results
        row = 24
        # actual semi_annual period 1 should be – for indicator row:
        self.assertEqual(sheet.cell(row=row, column=13).value, BLANK_CELL)
        # actual semi_annual period 1 should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, BLANK_CELL)
        # actual semi_annual period 2 should be – for indicator row:
        self.assertEqual(sheet.cell(row=row, column=14).value, BLANK_CELL)
        # actual semi_annual period 2 should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, BLANK_CELL)


        # Percentage / One result
        row = 27
        # actual semi_annual period 1 should be 5(500%) for indicator row:
        self.assertEqual(sheet.cell(row=row, column=13).value, 5)
        self.assertEqual(sheet.cell(row=row, column=13).number_format, '0%')
        # actual semi_annual period 1 should be 1(100%), 4(400%) for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, 1)
        self.assertEqual(sheet.cell(row=row+1, column=13).number_format, '0%')
        self.assertEqual(sheet.cell(row=row+2, column=13).value, 4)
        self.assertEqual(sheet.cell(row=row+2, column=13).number_format, '0%')
        # actual semi_annual period 2 should be – for indicator row:
        self.assertEqual(sheet.cell(row=row, column=14).value, BLANK_CELL)
        # actual semi_annual period 2 should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, BLANK_CELL)

        # Percentage / Many results
        row = 30
        # actual semi_annual period 1 should be 5(500%) for indicator row:
        self.assertEqual(sheet.cell(row=row, column=13).value, 5)
        self.assertEqual(sheet.cell(row=row, column=13).number_format, '0%')
        # actual semi_annual period 1 should be 1(100%), 4(400%) for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, 1)
        self.assertEqual(sheet.cell(row=row+1, column=13).number_format, '0%')
        self.assertEqual(sheet.cell(row=row+2, column=13).value, 4)
        self.assertEqual(sheet.cell(row=row+2, column=13).number_format, '0%')
        # actual semi_annual period 2 should be 5(500%) for indicator row:
        self.assertEqual(sheet.cell(row=row, column=14).value, 5)
        self.assertEqual(sheet.cell(row=row, column=14).number_format, '0%')
        # actual semi_annual period 2 should be 1(100%), 4(400%) for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, 1)
        self.assertEqual(sheet.cell(row=row+1, column=14).number_format, '0%')
        self.assertEqual(sheet.cell(row=row+2, column=14).value, 4)
        self.assertEqual(sheet.cell(row=row+2, column=14).number_format, '0%')

        return True

    def tva_period_column_tests(self, sheet):
        # sheet should only have 24 columns (4 quarters * 3 columns = 12 on top of 12 base columns)
        self.assertEqual(sheet.max_column, 24)
        # Numeric / Non-cumulative / No results
        row = 6
        for target_column in range(13, 24, 3):
            # targets should all be 250 for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column).value, 250)
            # targets should all be blank for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
            # actuals should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+1).value, BLANK_CELL)
            # actuals should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, BLANK_CELL)
            # % met should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+2).value, BLANK_CELL)
            # % met should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)

        # Numeric / Non-cumulative / One result
        row = 9
        # targets should all be 250 for indicator row
        self.assertEqual(sheet.cell(row=row, column=13).value, 250)
        # targets should all be blank for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, BLANK_CELL)
        # actuals should be 500 for indicator row
        self.assertEqual(sheet.cell(row=row, column=14).value, 500)
        # actuals should be 100,400 for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, 100)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, 400)
        # % met should be 2 (200%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=15).value, 2)
        self.assertEqual(sheet.cell(row=row, column=15).number_format, '0%')
        # % met should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=15).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=15).value, BLANK_CELL)
        for target_column in range(16, 24, 3):
            # targets should all be 250 for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column).value, 250)
            # targets should all be blank for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
            # actuals should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+1).value, BLANK_CELL)
            # actuals should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, BLANK_CELL)
            # % met should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+2).value, BLANK_CELL)
            # % met should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)

        # Numeric / Non-cumulative / Many results
        row = 12
        # no result column is 19:
        target_column = 19
        # targets should all be 250 for indicator row
        self.assertEqual(sheet.cell(row=row, column=target_column).value, 250)
        # targets should all be blank for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
        # actuals should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=target_column+1).value, BLANK_CELL)
        # actuals should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, BLANK_CELL)
        # % met should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=target_column+2).value, BLANK_CELL)
        # % met should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)
        for target_column in [13, 16, 22]:
            # targets should all be 250 for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column).value, 250)
            # targets should all be blank for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
            # actuals should be 500 for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+1).value, 500)
            # actuals should be 100, 400 for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, 100)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, 400)
            # % met should be 2 (200%) for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+2).value, 2)
            self.assertEqual(sheet.cell(row=row, column=target_column+2).number_format, '0%')
            # % met should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)

        # Numeric / cumulative / No results
        row = 15
        for target_column in range(13, 24, 3):
            # targets should all be 250 for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column).value, 250)
            # targets should all be blank for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
            # actuals should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+1).value, BLANK_CELL)
            # actuals should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, BLANK_CELL)
            # % met should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+2).value, BLANK_CELL)
            # % met should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)

        # Numeric / cumulative / One result
        row = 18
        # targets should all be 250 for indicator row
        self.assertEqual(sheet.cell(row=row, column=13).value, 250)
        # targets should all be blank for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, BLANK_CELL)
        # actuals should be 500 for indicator row
        self.assertEqual(sheet.cell(row=row, column=14).value, 500)
        # actuals should be 100,400 for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, 100)
        self.assertEqual(sheet.cell(row=row+2, column=14).value, 400)
        # % met should be 2 (200%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=15).value, 2)
        self.assertEqual(sheet.cell(row=row, column=15).number_format, '0%')
        # % met should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=15).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=15).value, BLANK_CELL)
        for target_column in range(16, 24, 3):
            # targets should all be 250 for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column).value, 250)
            # targets should all be blank for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
            # actuals should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+1).value, BLANK_CELL)
            # actuals should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, BLANK_CELL)
            # % met should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+2).value, BLANK_CELL)
            # % met should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)

        # Numeric / cumulative / Many results
        row = 21
        # no result column is 19:
        target_column = 19
        # targets should all be 250 for indicator row
        self.assertEqual(sheet.cell(row=row, column=target_column).value, 250)
        # targets should all be blank for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
        # actuals should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=target_column+1).value, BLANK_CELL)
        # actuals should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, BLANK_CELL)
        # % met should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=target_column+2).value, BLANK_CELL)
        # % met should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)
        for c, target_column in enumerate([13, 16, 22]):
            # targets should all be 250 for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column).value, 250)
            # targets should all be blank for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
            # actuals should be (500, 1000, 1500) for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+1).value, 500*(c+1))
            # actuals should be (100,400; 200,800; 300,1200) for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, 100*(c+1))
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, 400*(c+1))
            # % met should be 2(200%), 4(400%), 6(600%) for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+2).value, 2*(c+1))
            self.assertEqual(sheet.cell(row=row, column=target_column+2).number_format, '0%')
            # % met should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)
        
        # Percentage / No results
        row = 24
        for target_column in range(13, 24, 3):
            # targets should all be 2.5 (250%) for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column).value, 2.5)
            self.assertEqual(sheet.cell(row=row, column=target_column).number_format, '0%')
            # targets should all be blank for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
            # actuals should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+1).value, BLANK_CELL)
            # actuals should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, BLANK_CELL)
            # % met should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+2).value, BLANK_CELL)
            # % met should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)
        
        # Percentage / One result
        row = 27
        # targets should all be 2.5 (250%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=13).value, 2.5)
        self.assertEqual(sheet.cell(row=row, column=13).number_format, '0%')
        # targets should all be blank for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=13).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=13).value, BLANK_CELL)
        # actuals should be 5(500%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=14).value, 5)
        self.assertEqual(sheet.cell(row=row, column=14).number_format, '0%')
        # actuals should be 1(100%),4(400%) for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=14).value, 1)
        self.assertEqual(sheet.cell(row=row+1, column=14).number_format, '0%')
        self.assertEqual(sheet.cell(row=row+2, column=14).value, 4)
        self.assertEqual(sheet.cell(row=row+2, column=14).number_format, '0%')
        # % met should be 2 (200%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=15).value, 2)
        self.assertEqual(sheet.cell(row=row, column=15).number_format, '0%')
        # % met should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=15).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=15).value, BLANK_CELL)
        for target_column in range(16, 24, 3):
            # targets should all be 2.5 (250%) for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column).value, 2.5)
            self.assertEqual(sheet.cell(row=row, column=target_column).number_format, '0%')
            # targets should all be blank for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
            # actuals should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+1).value, BLANK_CELL)
            # actuals should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, BLANK_CELL)
            # % met should be – for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+2).value, BLANK_CELL)
            # % met should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)

        # Percentage / Many results
        row = 30
        # no result column is 19:
        target_column = 19
        # targets should all be 2.5 (250%) for indicator row
        self.assertEqual(sheet.cell(row=row, column=target_column).value, 2.5)
        self.assertEqual(sheet.cell(row=row, column=target_column).number_format, '0%')
        # targets should all be blank for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
        # actuals should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=target_column+1).value, BLANK_CELL)
        # actuals should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, BLANK_CELL)
        # % met should be – for indicator row
        self.assertEqual(sheet.cell(row=row, column=target_column+2).value, BLANK_CELL)
        # % met should be – for disaggregation rows:
        self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
        self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)
        for target_column in [13, 16, 22]:
            # targets should all be 2.5 (250%) for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column).value, 2.5)
            self.assertEqual(sheet.cell(row=row, column=target_column).number_format, '0%')
            # targets should all be blank for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column).value, BLANK_CELL)
            # actuals should be 5(500%) for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+1).value, 5)
            self.assertEqual(sheet.cell(row=row, column=target_column+1).number_format, '0%')
            # actuals should be 1(100%), 4(400%) for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).value, 1)
            self.assertEqual(sheet.cell(row=row+1, column=target_column+1).number_format, '0%')
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).value, 4)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+1).number_format, '0%')
            # % met should be 2 (200%) for indicator row
            self.assertEqual(sheet.cell(row=row, column=target_column+2).value, 2)
            self.assertEqual(sheet.cell(row=row, column=target_column+2).number_format, '0%')
            # % met should be – for disaggregation rows:
            self.assertEqual(sheet.cell(row=row+1, column=target_column+2).value, BLANK_CELL)
            self.assertEqual(sheet.cell(row=row+2, column=target_column+2).value, BLANK_CELL)
        
        return True