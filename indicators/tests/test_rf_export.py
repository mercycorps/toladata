"""Tests for the Excel export of the Results Framework for a given program"""

import io
import openpyxl
from django import test
from django.urls import reverse
from factories.workflow_models import (
    RFProgramFactory,
    CountryFactory,
    TolaUserFactory,
    grant_program_access,
    grant_country_access,
    PROGRAM_ROLE_CHOICES,
    COUNTRY_ROLE_CHOICES
)
from factories.indicators_models import LevelFactory

# filename for all rf export downloads:
FILENAME = "Results Framework.xlsx"

# url resolution helper:
get_export_url = lambda program: reverse('rf_export', kwargs={'program': program.pk})


class TestRFExportDownloadPermissions(test.TestCase):
    def test_url_resolves(self):
        program = RFProgramFactory()
        url = get_export_url(program)
        self.assertIsNotNone(url)

    def test_anonymous_user_denied(self):
        program = RFProgramFactory()
        response = self.client.get(get_export_url(program))
        expected_url = f"{reverse('login')}?next={get_export_url(program)}"
        self.assertRedirects(response, expected_url, status_code=302, target_status_code=200)

    def test_logged_in_and_permissioned_user_status(self):
        country = CountryFactory(code="TL", country="TestLand")
        out_country = CountryFactory(code="OT", country="OutLand")
        program = RFProgramFactory()
        program.country.set([country])
        out_program = RFProgramFactory()
        out_program.country.set([country])
        # superuser should have access:
        superuser = TolaUserFactory(superadmin=True)
        self.client.force_login(superuser.user)
        response = self.client.get(get_export_url(program))
        self.assertEqual(response.status_code, 200)
        self.client.logout()
        # users with permission should have access:
        user_1 = TolaUserFactory(mc_staff=True, superadmin=False)
        grant_country_access(user_1, country, role=COUNTRY_ROLE_CHOICES[0][0])
        user_2 = TolaUserFactory(mc_staff=True, superadmin=False)
        grant_country_access(user_2, country, role=COUNTRY_ROLE_CHOICES[1][0])
        user_3 = TolaUserFactory(mc_staff=False, superadmin=False)
        grant_program_access(user_3, program, country, role=PROGRAM_ROLE_CHOICES[0][0])
        user_4 = TolaUserFactory(mc_staff=False, superadmin=False)
        grant_program_access(user_4, program, country, role=PROGRAM_ROLE_CHOICES[1][0])
        user_5 = TolaUserFactory(mc_staff=False, superadmin=False)
        grant_program_access(user_5, program, country, role=PROGRAM_ROLE_CHOICES[2][0])
        for user in [user_1, user_2, user_3, user_4, user_5]:
            self.client.force_login(user.user)
            response = self.client.get(get_export_url(program))
            self.assertEqual(response.status_code, 200)
            self.client.logout()
        out_user_1 = TolaUserFactory(mc_staff=True, superadmin=False)
        grant_country_access(out_user_1, out_country, role=COUNTRY_ROLE_CHOICES[0][0])
        out_user_2 = TolaUserFactory(mc_staff=True, superadmin=False)
        grant_country_access(out_user_1, out_country, role=COUNTRY_ROLE_CHOICES[1][0])
        out_user_3 = TolaUserFactory(mc_staff=False, superadmin=False)
        grant_program_access(out_user_3, out_program, country, role=PROGRAM_ROLE_CHOICES[0][0])
        for user in [out_user_1, out_user_2, out_user_3]:
            self.client.force_login(user.user)
            response = self.client.get(get_export_url(program))
            self.assertEqual(response.status_code, 403)
            self.client.logout()


class TestRFExportStandardMCTiers(test.TestCase):

    @classmethod
    def setUpTestData(cls):
        CountryFactory.reset_sequence()
        cls.program_name = "Standard Program Name"
        cls.program = RFProgramFactory(tiers=True, levels=False, name=cls.program_name)
        cls.goal_level = LevelFactory(program=cls.program, parent=None, name="Goal Level", customsort=1)
        cls.outcome_levels = [
            LevelFactory(program=cls.program, parent=cls.goal_level, name="First Outcome", customsort=1),
            LevelFactory(program=cls.program, parent=cls.goal_level, name="Second Outcome", customsort=2)
        ]
        cls.output_levels = [
            LevelFactory(program=cls.program, parent=cls.outcome_levels[0], name="First Output", customsort=1),
            LevelFactory(program=cls.program, parent=cls.outcome_levels[0], name="Second Output", customsort=2),
            LevelFactory(program=cls.program, parent=cls.outcome_levels[1], name="Third Output", customsort=1),
            LevelFactory(program=cls.program, parent=cls.outcome_levels[1], name="Fourth Output", customsort=2),
        ]
        cls.activity_levels = [
            LevelFactory(program=cls.program, parent=cls.output_levels[0], name="First Activity", customsort=1),
            LevelFactory(program=cls.program, parent=cls.output_levels[0], name="Second Activity", customsort=2),
            LevelFactory(program=cls.program, parent=cls.output_levels[0], name="Third Activity", customsort=3),
            LevelFactory(program=cls.program, parent=cls.output_levels[1], name="Fourth Activity", customsort=1),
            LevelFactory(program=cls.program, parent=cls.output_levels[1], name="Fifth Activity", customsort=2),
            LevelFactory(program=cls.program, parent=cls.output_levels[1], name="Sixth Activity", customsort=3),
            LevelFactory(program=cls.program, parent=cls.output_levels[2], name="Seventh Activity", customsort=1),
            LevelFactory(program=cls.program, parent=cls.output_levels[2], name="Eighth Activity", customsort=2),
            LevelFactory(program=cls.program, parent=cls.output_levels[2], name="Ninth Activity", customsort=3),
            LevelFactory(program=cls.program, parent=cls.output_levels[3], name="Tenth Activity", customsort=1),
            LevelFactory(program=cls.program, parent=cls.output_levels[3], name="Eleventh Activity", customsort=2),
            LevelFactory(program=cls.program, parent=cls.output_levels[3], name="Twelfth Activity", customsort=3),
        ]
        cls.superuser = TolaUserFactory(superadmin=True)

    def setUp(self):
        self.client.force_login(user=self.superuser.user)

    def tearDown(self):
        self.client.logout()

    def test_report_has_correct_format(self):
        response = self.client.get(get_export_url(self.program))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get('Content-Type'), "application/ms-excel")
        self.assertEqual(response.get('Content-Disposition'), f'attachment; filename="{FILENAME}"')
        openpyxl.load_workbook(io.BytesIO(response.content))

    def level_cell_tests(self, cell, span=True):
        self.assertEqual(cell.font.name, 'Calibri')
        self.assertEqual(cell.font.size, 12)
        self.assertTrue(cell.alignment.wrap_text)
        self.assertEqual(cell.fill.patternType, 'solid')
        if span:
            self.assertEqual(cell.alignment.vertical, 'center')
            self.assertEqual(cell.alignment.horizontal, 'center')
        else:
            self.assertEqual(cell.alignment.vertical, 'top')
            self.assertEqual(cell.alignment.horizontal, 'left')
        return True

    def test_report_content(self):
        report = openpyxl.load_workbook(io.BytesIO(self.client.get(get_export_url(self.program)).content))
        worksheet = report.active
        title_cell = worksheet.cell(row=2, column=2)
        self.assertEqual(title_cell.value, "Results Framework")
        self.assertTrue(title_cell.font.bold)
        self.assertEqual(title_cell.font.name, "Calibri")
        self.assertEqual(title_cell.font.size, 18)
        program_name_cell = worksheet.cell(row=3, column=2)
        self.assertEqual(program_name_cell.value, self.program_name)
        self.assertFalse(program_name_cell.font.bold)
        self.assertEqual(program_name_cell.font.name, 'Calibri')
        self.assertEqual(program_name_cell.font.size, 18)
        goal_cell = worksheet.cell(row=5, column=2)
        self.assertEqual(goal_cell.value, "Goal: Goal Level")
        self.assertTrue(self.level_cell_tests(goal_cell))
        outcome_cell1 = worksheet.cell(row=7, column=2)
        self.assertEqual(outcome_cell1.value, "Outcome 1: First Outcome")
        self.assertTrue(self.level_cell_tests(outcome_cell1))
        outcome_cell2 = worksheet.cell(row=7, column=6)
        self.assertEqual(outcome_cell2.value, "Outcome 2: Second Outcome")
        self.assertTrue(self.level_cell_tests(outcome_cell2))
        for output, column in [
                ("Output 1.1: First Output", 2),
                ("Output 1.2: Second Output", 4),
                ("Output 2.1: Third Output", 6),
                ("Output 2.2: Fourth Output", 8),
            ]:
            cell = worksheet.cell(row=9, column=column)
            self.assertEqual(cell.value, output)
            self.assertTrue(self.level_cell_tests(cell, span=False))
        # spot checking activities to avoid writing a novel:
        for row, column, activity in [
                (11, 2, "Activity 1.1.1: First Activity"),
                (13, 4, "Activity 1.2.2: Fifth Activity"),
                (15, 6, "Activity 2.1.3: Ninth Activity"),
                (15, 8, "Activity 2.2.3: Twelfth Activity"),
            ]:
            cell = worksheet.cell(row=row, column=column)
            self.assertEqual(cell.value, activity)
            self.assertTrue(self.level_cell_tests(cell, span=False))
        for merge_range in ['B5:H5', 'B7:D7', 'F7:H7']:
            self.assertIn(merge_range, worksheet.merged_cells)
        for column in ['B', 'D', 'F', 'H']:
            self.assertEqual(worksheet.column_dimensions[column].width, 24.5)
        for column in ['A', 'C', 'E', 'G']:
            self.assertEqual(worksheet.column_dimensions[column].width, 3)


class RFExportTests:
    def test_merged_cells(self):
        for merge_range in self.merged_cells:
            self.assertIn(merge_range, self.worksheet.merged_cells)
        for column in self.used_columns:
            self.assertEqual(self.worksheet.column_dimensions[column].width, 24.5)
        for column in self.border_columns:
            self.assertEqual(self.worksheet.column_dimensions[column].width, 3)
        for row in range(3, self.max_row+1):
            if row % 2 == 0:
                self.assertEqual(self.worksheet.row_dimensions[row].height, 10, row)


class TestCustomMultiTieredRFExport(test.TestCase, RFExportTests):
    merged_cells = ['B5:H5', 'B7:H7', 'B9:F9', 'B11:F11', 'B13:D13', 'B15:D15']
    used_columns = ['B', 'D', 'F', 'H']
    border_columns = ['A', 'C', 'E', 'G']
    max_row = 22

    @classmethod
    def setUpTestData(cls):
        cls.program = RFProgramFactory(tiers=[f"Tier {x}" for x in range(1, 9)], levels=1)
        goal_levels = cls.program.levels.filter(parent=None)
        assert goal_levels.count() == 1
        cls.goal_level = goal_levels.first()
        tier_2_levels = cls.program.levels.filter(parent=cls.goal_level)
        assert tier_2_levels.count() == 1
        cls.tier_2_level = tier_2_levels.first()
        assert cls.tier_2_level.customsort == 1
        # add a second level at tier 3 with no children
        cls.orphan_tier_3_level = LevelFactory(
            program=cls.program, customsort=2, name="Orphan Tier 3 Level", parent=cls.tier_2_level)
        assert cls.program.levels.filter(parent=cls.tier_2_level).count() == 2
        # add a second level at tier 5 with one child at tier 6 (but none at bottom tier)
        cls.tier_4_level = cls.program.levels.filter(parent__parent=cls.tier_2_level).first()
        assert cls.tier_4_level.child_levels.count() == 1
        assert cls.tier_4_level.child_levels.first().child_levels.count() == 1
        cls.tier_6_level = cls.tier_4_level.child_levels.first().child_levels.first()
        cls.new_tier_5_level = LevelFactory(
            program=cls.program, customsort=2, name="Parent with Orphan Tier 6", parent=cls.tier_4_level)
        cls.orphan_tier_6_level = LevelFactory(
            program=cls.program, customsort=1, name="Orphaned Child at Tier 6", parent=cls.new_tier_5_level)
        # add a second tier 7 level (2nd from bottom):
        cls.new_tier_7_level = LevelFactory(
            program=cls.program, customsort=2, name="Bonus Tier 7", parent=cls.tier_6_level)
        # make the 2nd column longer than the first:
        cls.new_tier_8_levels = [
            LevelFactory(
                program=cls.program, customsort=1, name="Child of Bonus Tier 7 1", parent=cls.new_tier_7_level),
            LevelFactory(
                program=cls.program, customsort=2, name="Child of Bonus Tier 7 2", parent=cls.new_tier_7_level)
        ]
        assert cls.program.levels.count() == 14
        cls.superuser = TolaUserFactory(superadmin=True)

    def setUp(self):
        self.client.force_login(user=self.superuser.user)
        self.worksheet = openpyxl.load_workbook(
            io.BytesIO(self.client.get(get_export_url(self.program)).content)
        ).active

    def tearDown(self):
        self.client.logout()

    def test_level_cell_content(self):
        for row, column, tier_name, level in [
                (7, 2, "Tier 2", self.tier_2_level),
                (9, 8, "Tier 3", self.orphan_tier_3_level),
                (11, 2, "Tier 4", self.tier_4_level),
                (13, 6, "Tier 5", self.new_tier_5_level),
                (15, 2, "Tier 6", self.tier_6_level),
                (15, 6, "Tier 6", self.orphan_tier_6_level),
                (17, 4, "Tier 7", self.new_tier_7_level),
                (19, 4, "Tier 8", self.new_tier_8_levels[0]),
                (21, 4, "Tier 8", self.new_tier_8_levels[1])
            ]:
            self.assertEqual(self.worksheet.cell(row=row, column=column).value,
                             f"{tier_name} {level.display_ontology}: {level.name}")


class TestShortTierListExport(test.TestCase, RFExportTests):
    merged_cells = ['B5:D5',]
    used_columns = ['B', 'D']
    border_columns = ['A', 'C', 'E']
    max_row = 11

    @classmethod
    def setUpTestData(cls):
        cls.program = RFProgramFactory(tiers=[f"Tîér {x}" for x in range(1, 4)], levels=2)
        cls.superuser = TolaUserFactory(superadmin=True)

    def setUp(self):
        self.client.force_login(user=self.superuser.user)
        self.worksheet = openpyxl.load_workbook(
            io.BytesIO(self.client.get(get_export_url(self.program)).content)
        ).active

    def tearDown(self):
        self.client.logout()

    def test_level_cell_content(self):
        goal_level = self.program.levels.filter(parent=None).first()
        second_levels = list(goal_level.child_levels.all())
        third_levels = [list(l.child_levels.all()) for l in second_levels]
        for row, column, tier_name, level in [
                (7, 2, "Tîér 2", second_levels[0]),
                (7, 4, "Tîér 2", second_levels[1]),
                (9, 2, "Tîér 3", third_levels[0][0]),
                (11, 2, "Tîér 3", third_levels[0][1]),
                (9, 4, "Tîér 3", third_levels[1][0]),
                (11, 4, "Tîér 3", third_levels[1][1]),
            ]:
            self.assertEqual(self.worksheet.cell(row=row, column=column).value,
                             f"{tier_name} {level.display_ontology}: {level.name}")
        self.assertEqual(self.worksheet.cell(row=5, column=2).value, f"Tîér 1: {goal_level.name}")


class TestTranslations(test.TestCase):

    @classmethod
    def setUpTestData(cls):
        cls.program = RFProgramFactory(tiers=True, levels=2)
        cls.user_en = TolaUserFactory(superadmin=True)
        cls.user_es = TolaUserFactory(superadmin=True, language='es')
        cls.user_fr = TolaUserFactory(superadmin=True, language='fr')

    def setUp(self):
        self.client.force_login(user=self.user_en.user)
        self.worksheet_en = openpyxl.load_workbook(
            io.BytesIO(self.client.get(get_export_url(self.program)).content)
        ).active
        self.client.logout()
        self.client.force_login(user=self.user_es.user)
        self.worksheet_es = openpyxl.load_workbook(
            io.BytesIO(self.client.get(get_export_url(self.program)).content)
        ).active
        self.client.logout()
        self.client.force_login(user=self.user_fr.user)
        self.worksheet_fr = openpyxl.load_workbook(
            io.BytesIO(self.client.get(get_export_url(self.program)).content)
        ).active
        self.client.logout()

    def tearDown(self):
        self.client.logout()

    def test_translated_title(self):
        self.assertEqual(self.worksheet_es.cell(row=2, column=2).value, "Sistema de Resultados")
        self.assertEqual(self.worksheet_fr.cell(row=2, column=2).value, "Cadre de résultats")

    def test_translated_tiers(self):
        for row, column, en, fr, es in [
                (5, 2, "Goal", "But", "Objetivo"),
                (7, 2, "Outcome", "Résultat", "Resultado"),
                (7, 6, "Outcome", "Résultat", "Resultado"),
                (9, 2, "Output", "Extrant", "Salida"),
                (9, 4, "Output", "Extrant", "Salida"),
                (9, 6, "Output", "Extrant", "Salida"),
                (9, 8, "Output", "Extrant", "Salida"),
                (11, 2, "Activity", "Activité", "Actividad"),
            ]:
            self.assertEqual(self.worksheet_en.cell(row=row, column=column).value[:len(en)], en)
            self.assertEqual(self.worksheet_fr.cell(row=row, column=column).value[:len(fr)], fr)
            self.assertEqual(self.worksheet_es.cell(row=row, column=column).value[:len(es)], es)
